"""
TGS Projections Data Extractor — Multi-League
Auto-discovers all "The Sheets *" folders and extracts each league's data
into its own subfolder. Generates a leagues.json manifest for the web app.

Usage:
  python extract_data.py              # Extract ALL leagues
  python extract_data.py BLM          # Extract only the "BLM" league
  python extract_data.py "BLM - Copy" # Extract only "BLM - Copy"
"""

import json
import os
import sys
from datetime import datetime
from openpyxl import load_workbook

# Parent directory containing all "The Sheets *" folders
PARENT_DIR = r"C:\Users\perfe\Desktop\TGS Projections"
OUTPUT_BASE = os.path.join(PARENT_DIR, "tgs-viz", "public", "data")

# The 4 Excel files we look for in each league folder.
# Keys = dataset name, values = (possible filenames, output json name, sheet hint)
# Filenames are checked case-insensitively.
DATASET_DEFS = {
    "hitters": {
        "filenames": ["The Sheet Hitters.xlsx"],
        "output": "hitters.json",
        "sheet_hint": "Hitters",
    },
    "pitchers": {
        "filenames": ["The Sheet Pitchers.xlsx"],
        "output": "pitchers.json",
        "sheet_hint": "Pitchers",
    },
    "hitters_draft": {
        "filenames": ["The Sheet Hitters - Draft.xlsx"],
        "output": "hitters_draft.json",
        "sheet_hint": "Hitters",
    },
    "pitchers_draft": {
        "filenames": ["The Sheet Pitchers - Draft.xlsx", "The Sheet Pitchers - DRAFT.xlsx"],
        "output": "pitchers_draft.json",
        "sheet_hint": "Pitchers",
    },
    "hitters_fa": {
        "filenames": ["The Sheet Hitters - FA.xlsx"],
        "output": "hitters_fa.json",
        "sheet_hint": "Hitters",
    },
    "pitchers_fa": {
        "filenames": ["The Sheet Pitchers - FA.xlsx"],
        "output": "pitchers_fa.json",
        "sheet_hint": "Pitchers",
    },
}


def safe_val(val):
    """Convert Excel cell values to JSON-safe types."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        if isinstance(val, float) and (val != val):  # NaN check
            return None
        return val
    return str(val).strip()


def extract_sheet(filepath, sheet_name, max_rows=None):
    """Extract data from a specific sheet in an Excel file."""
    print(f"  Loading workbook: {os.path.basename(filepath)}...")
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Find the right sheet (case-insensitive exact match first)
    target_sheet = None
    for name in wb.sheetnames:
        if name.lower() == sheet_name.lower():
            target_sheet = name
            break

    if not target_sheet:
        # Try partial match
        for name in wb.sheetnames:
            if sheet_name.lower() in name.lower():
                target_sheet = name
                break

    if not target_sheet:
        print(f"  Available sheets: {wb.sheetnames}")
        print(f"  WARNING: Sheet '{sheet_name}' not found!")
        wb.close()
        return []

    print(f"  Reading sheet: {target_sheet}")
    ws = wb[target_sheet]

    rows = []
    headers = None
    row_count = 0

    for row in ws.iter_rows():
        values = [safe_val(cell.value) for cell in row]

        if headers is None:
            headers = values
            headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(headers)]
            continue

        if all(v is None for v in values):
            continue

        record = {}
        for i, h in enumerate(headers):
            if i < len(values):
                record[h] = values[i]
            else:
                record[h] = None

        rows.append(record)
        row_count += 1

        if max_rows and row_count >= max_rows:
            break

        if row_count % 1000 == 0:
            print(f"    Processed {row_count} rows...")

    wb.close()
    print(f"  Extracted {row_count} rows with {len(headers) if headers else 0} columns")
    return rows


def extract_draft_filter_ids(filepath):
    """Extract draft-eligible and already-drafted player IDs from a draft workbook.

    Reads two sheets:
      1. 'Player List' — rows with Manual='DRAFT' and a Name are draft-eligible.
      2. 'Drafted' — CSV-like rows pasted from StatsPlus with already-drafted IDs.

    Returns (eligible_ids, drafted_ids, dem_map, pitch_map) where pitch_map is
    a dict of player ID -> {FB, FBP, CH, CHP, ...} pitch rating columns.
    """
    # Individual pitch type columns to extract from Player List
    PITCH_COLS = [
        'FB', 'FBP', 'CH', 'CHP', 'CB', 'CBP', 'SL', 'SLP',
        'SI', 'SIP', 'SP', 'SPP', 'CT', 'CTP', 'FO', 'FOP',
        'CC', 'CCP', 'SC', 'SCP', 'KC', 'KCP', 'KN', 'KNP',
    ]

    wb = load_workbook(filepath, read_only=True, data_only=True)

    # --- Player List: get draft-eligible IDs, DEM values, and pitch ratings ---
    eligible = set()
    dem_map = {}
    pitch_map = {}
    if 'Player List' in wb.sheetnames:
        ws = wb['Player List']
        headers = None
        for i, row in enumerate(ws.iter_rows()):
            vals = [cell.value for cell in row]
            if i == 0:
                headers = [str(v).strip() if v else f"Col_{j}" for j, v in enumerate(vals)]
                continue
            if not headers:
                continue
            rec = {headers[j]: vals[j] for j in range(min(len(headers), len(vals)))}

            pid = rec.get('ID')
            if pid is None:
                continue
            pid_str = str(pid).strip()

            # Draft eligibility: any player with a Name in the Player List is draft-eligible
            if rec.get('Name'):
                eligible.add(pid_str)

            # DEM value
            dem = rec.get('DEM')
            if dem is not None:
                dem_map[pid_str] = str(dem).strip()

            # Pitch ratings — store all available columns for this player
            pitch_data = {}
            for col in PITCH_COLS:
                v = rec.get(col)
                if v is not None:
                    pitch_data[col] = safe_val(v)
            if pitch_data:
                pitch_map[pid_str] = pitch_data

        print(f"  Player List: {len(eligible)} draft-eligible players with names")
        print(f"  Player List: {len(dem_map)} players with DEM values")
        print(f"  Player List: {len(pitch_map)} players with pitch ratings")

    # --- Drafted sheet: get already-drafted IDs ---
    drafted = set()
    if 'Drafted' in wb.sheetnames:
        ws = wb['Drafted']
        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                continue  # skip header
            val = row[0].value
            if val and isinstance(val, str) and val.startswith('"'):
                parts = val.replace('"', '').split(',')
                if parts[0].strip():
                    drafted.add(parts[0].strip())
        print(f"  Drafted sheet: {len(drafted)} already-drafted player IDs")

    remaining = eligible - drafted
    print(f"  Remaining draftable: {len(remaining)}")

    wb.close()
    return eligible, drafted, dem_map, pitch_map


def extract_sheet_names(filepath):
    """Get all sheet names from a workbook."""
    wb = load_workbook(filepath, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def find_file_case_insensitive(directory, possible_names):
    """Find a file in directory matching any of the possible names (case-insensitive)."""
    try:
        actual_files = os.listdir(directory)
    except OSError:
        return None

    # Build a lowercase lookup
    lower_map = {f.lower(): f for f in actual_files}

    for name in possible_names:
        actual = lower_map.get(name.lower())
        if actual:
            return os.path.join(directory, actual)

    return None


def discover_leagues():
    """Find all 'The Sheets *' folders in the parent directory."""
    leagues = []
    try:
        entries = os.listdir(PARENT_DIR)
    except OSError as e:
        print(f"ERROR: Cannot read parent directory: {e}")
        return leagues

    for entry in sorted(entries):
        full_path = os.path.join(PARENT_DIR, entry)
        if os.path.isdir(full_path) and entry.startswith("The Sheets"):
            # Derive league ID: "The Sheets BLM" -> "BLM", "The Sheets TGS - Copy" -> "TGS - Copy"
            league_id = entry.replace("The Sheets", "").strip()
            if not league_id:
                league_id = "Default"
            leagues.append({
                "id": league_id,
                "name": league_id,
                "folder": entry,
                "path": full_path,
            })

    return leagues


def extract_league(league, output_base):
    """Extract all datasets for a single league."""
    league_id = league["id"]
    league_path = league["path"]
    league_output = os.path.join(output_base, league_id)
    os.makedirs(league_output, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"LEAGUE: {league_id}")
    print(f"Source: {league_path}")
    print(f"Output: {league_output}")
    print(f"{'='*60}")

    extracted_datasets = []
    all_data = {}

    for ds_key, ds_def in DATASET_DEFS.items():
        filepath = find_file_case_insensitive(league_path, ds_def["filenames"])

        if not filepath:
            print(f"\n  [{ds_key}] No matching file found (looked for: {ds_def['filenames']})")
            continue

        print(f"\n  [{ds_key}] Found: {os.path.basename(filepath)}")

        sheets = extract_sheet_names(filepath)
        print(f"  Sheets: {sheets}")

        data = extract_sheet(filepath, ds_def["sheet_hint"])

        if not data and len(sheets) >= 5:
            print(f"  Trying sheet by index: {sheets[4]}")
            data = extract_sheet(filepath, sheets[4])

        # For draft datasets, filter to only draft-eligible players
        # who haven't been drafted yet (using Player List + Drafted sheets)
        is_draft = ds_key in ("hitters_draft", "pitchers_draft")
        if data and is_draft:
            eligible, drafted, dem_map, pitch_map = extract_draft_filter_ids(filepath)
            draftable = eligible - drafted
            if draftable:
                before = len(data)
                data = [row for row in data if str(row.get("ID", "")).strip() in draftable]
                print(f"  Draft filter: {before} -> {len(data)} players (removed {before - len(data)} non-draftable)")
            else:
                print(f"  WARNING: No draftable IDs found — keeping all data unfiltered")

            # Attach DEM (demand) values from Player List to each player
            if dem_map:
                dem_count = 0
                for row in data:
                    pid = str(row.get("ID", "")).strip()
                    row["DEM"] = dem_map.get(pid, None)
                    if row["DEM"]:
                        dem_count += 1
                print(f"  DEM values attached: {dem_count} players")

            # Attach individual pitch ratings from Player List (pitchers only)
            if pitch_map and ds_key == "pitchers_draft":
                pitch_count = 0
                for row in data:
                    pid = str(row.get("ID", "")).strip()
                    pdata = pitch_map.get(pid)
                    if pdata:
                        row.update(pdata)
                        pitch_count += 1
                print(f"  Pitch ratings merged: {pitch_count} pitchers")

        if data:
            outpath = os.path.join(league_output, ds_def["output"])
            with open(outpath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False)
            size_mb = os.path.getsize(outpath) / (1024 * 1024)
            print(f"  Saved {len(data)} records to {ds_def['output']} ({size_mb:.1f} MB)")
            extracted_datasets.append(ds_key)
            all_data[ds_key] = data
        else:
            print(f"  ERROR: No data extracted for {ds_key}!")

    # Per-league metadata
    meta = {
        "extracted_at": datetime.now().isoformat(),
        "league": league_id,
        "source_dir": league_path,
        "datasets": {}
    }
    for key, data in all_data.items():
        if data:
            meta["datasets"][key] = {
                "count": len(data),
                "columns": list(data[0].keys()) if data else [],
                "sample": data[0] if data else None
            }

    meta_path = os.path.join(league_output, "metadata.json")
    with open(meta_path, 'w', encoding='utf-8') as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)

    return extracted_datasets


def main():
    print("TGS Projections Data Extractor — Multi-League")
    print(f"Parent: {PARENT_DIR}")
    print(f"Output: {OUTPUT_BASE}")

    os.makedirs(OUTPUT_BASE, exist_ok=True)

    # Discover all league folders
    leagues = discover_leagues()

    if not leagues:
        print("\nERROR: No 'The Sheets *' folders found!")
        sys.exit(1)

    print(f"\nDiscovered {len(leagues)} league(s):")
    for lg in leagues:
        print(f"  - {lg['id']} ({lg['folder']})")

    # Optional: filter to a single league if CLI arg provided
    filter_league = None
    if len(sys.argv) > 1:
        filter_league = sys.argv[1]
        matching = [lg for lg in leagues if lg["id"] == filter_league]
        if not matching:
            print(f"\nERROR: League '{filter_league}' not found!")
            print(f"Available: {[lg['id'] for lg in leagues]}")
            sys.exit(1)
        leagues = matching
        print(f"\nFiltering to league: {filter_league}")

    # Extract each league
    league_manifest = []

    for league in leagues:
        datasets = extract_league(league, OUTPUT_BASE)
        league_manifest.append({
            "id": league["id"],
            "name": league["name"],
            "folder": league["folder"],
            "datasets": datasets,
        })

    # Write the leagues manifest (always write full manifest even if filtering)
    # If filtering, merge with existing manifest
    manifest_path = os.path.join(OUTPUT_BASE, "leagues.json")

    if filter_league and os.path.exists(manifest_path):
        # Merge: keep existing entries, update the filtered one
        with open(manifest_path, 'r', encoding='utf-8') as f:
            existing = json.load(f)
        # Remove old entry for this league, add new one
        existing = [e for e in existing if e["id"] != filter_league]
        existing.extend(league_manifest)
        # Sort by id for consistency
        existing.sort(key=lambda x: x["id"])
        league_manifest = existing

    # Sort alphabetically
    league_manifest.sort(key=lambda x: x["id"])

    with open(manifest_path, 'w', encoding='utf-8') as f:
        json.dump(league_manifest, f, indent=2, ensure_ascii=False)

    print(f"\n{'='*60}")
    print("EXTRACTION COMPLETE")
    print(f"{'='*60}")
    print(f"\nLeagues manifest ({manifest_path}):")
    for lg in league_manifest:
        ds_list = ", ".join(lg["datasets"]) if lg["datasets"] else "none"
        print(f"  {lg['id']}: [{ds_list}]")

    print(f"\nTotal leagues: {len(league_manifest)}")


if __name__ == "__main__":
    main()
