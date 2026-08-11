"""
fielding_curves_fit.py — audit D3: replace the LINEAR range->PM% fits with
monotone piecewise (isotonic) curves fitted on opportunity-weighted archive
bucket means, per position per league. Emits calib/<LG>/fielding_curves.json,
the opt-in `fielding` layer consumed by engine/hitters.py via
ratings.live_fielding().

Why: OOTP's per-difficulty catch probabilities hit hard floors/ceilings, so
PM% vs range is an S-curve; the linear fit mis-prices BOTH tails and the
mid-band (verified: RF 60-vs-80 engine +22.6 runs vs sim +0.3; LF 55-vs-60
-11.2; CF 65-vs-75 -11.1). The clamp proposed originally was refuted — the
mid-band is wrong too, so the whole curve is refit monotone.

Method (NOTHING INVENTED):
  1. Rebuild the exact per-position fit pools (top-N by innings, same as
     calibrate.py fielding()); the bivariate/univariate LINEST refit from
     these pools must reproduce constants-latest.json before anything else
     runs (pool fidelity gate).
  2. Partial out the secondary predictor (IF ARM / HT CM) with its OWN fitted
     LINEST coefficient, bucket the adjusted PM% by range rating on 5-point
     rungs weighted by opportunities (pa), and fit MONOTONE non-decreasing
     via weighted PAVA. The curve is the piecewise-linear interpolation
     through the isotonic bucket means (flat beyond the observed support).
  3. Transport: offset = the LIVE position population's innings-weighted mean
     of  curve(RNG) + x2*m2  (populations + innings read straight from the
     league's 25 Metadata.xlsx Fielding Data/Fielding Ratings tabs,
     READ-ONLY). The engine evaluates  curve(RNG) + x2*m2 - offset, so the
     ip-weighted league-mean PM-term is 0 BY CONSTRUCTION — the Phase A
     anchor property survives exactly (the linear anchors only achieved it
     approximately).
  4. Gate: the audit D3 verifier's eligibility-band engine-vs-sim error
     matrix (all 5-rung rating pairs inside the position's eligibility band,
     scaled by the live plays/season and runs-per-play): must be within
     +/- 3 runs. Printed per position, with the linear fit's worst error for
     comparison.

The projection sheets keep their LINEAR PM% constants (sheet formula changes
are out of scope) — engine-side INTENTIONAL divergence, documented in
hitters.py.

Usage:
  python tgs-viz/engine/fielding_curves_fit.py                # both leagues
  python tgs-viz/engine/fielding_curves_fit.py --league TGS
"""
import argparse
import datetime
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import calibrate as C  # noqa: E402
from hitter_tails_fit import pava  # noqa: E402  (same weighted PAVA)

REPO = os.path.dirname(os.path.dirname(HERE))

# pos -> (fielding position code, topN, range col, x2 col (archive Hitters.csv),
#         live x2 source, plays T cell, runs-per-play cell, eligibility min RNG)
POS_META = {
    "1B": (3, 32, "IF RNG", "HT CM", "HT", "T5", "H38", 20),
    "2B": (4, 32, "IF RNG", "IF ARM", "IF ARM", "T8", "H38", 50),
    "3B": (5, 32, "IF RNG", "IF ARM", "IF ARM", "T12", "H38", 40),
    "SS": (6, 32, "IF RNG", "IF ARM", "IF ARM", "T15", "H38", 60),
    "LF": (7, 30, "OF RNG", None, None, "T18", "H39", 50),
    "CF": (8, 30, "OF RNG", None, None, "T22", "H39", 60),
    "RF": (9, 30, "OF RNG", None, None, "T26", "H39", 50),
}
SUPPORT = (20.0, 80.0)


def _norm(h):
    return (str(h) if h is not None else "").replace("\xa0", " ") \
        .replace("▴", "").replace("▾", "").strip()


def _num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return None
    return None


def dollarde(ip):
    try:
        ip = float(ip)
    except (TypeError, ValueError):
        return 0.0
    return int(ip) + (ip - int(ip)) * 10.0 / 3.0


def height_cm(ht):
    """5' 11" -> 5*30.48 + 11*2.54 (identical to the sheet's HT Sort scale)."""
    try:
        feet, rest = str(ht).split("'", 1)
        inches = rest.split('"')[0]
        return float(feet) * 30.48 + float(inches) * 2.54
    except (AttributeError, ValueError):
        return None


def read_live(league):
    """25 Metadata.xlsx (READ-ONLY): per-position live [(id, innings)] from the
    Fielding Data tab + per-id ratings from the Fielding Ratings tab."""
    from openpyxl import load_workbook
    path = os.path.join(REPO, f"The Sheets {league}", "25 Metadata.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)

    ws = wb["Fielding Data"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [_norm(h) for h in rows[1]]
    idcols = [j for j, h in enumerate(hdr) if h == "ID"]
    code2pos = {2: "C", 3: "1B", 4: "2B", 5: "3B", 6: "SS", 7: "LF", 8: "CF", 9: "RF"}
    pos_ip = {}
    for start in idcols:
        cols = {}
        for j in range(start, len(hdr)):
            if hdr[j] in ("", "<--", "-->") and j > start:
                break
            if hdr[j]:
                cols.setdefault(hdr[j], j)
        recs = []
        for r in rows[2:]:
            if start >= len(r) or r[start] in (None, ""):
                continue
            recs.append(r)
        if not recs:
            continue
        pos = code2pos.get(int(_num(recs[0][cols["POS"]]) or 0))
        if not pos:
            continue
        pairs = []
        for r in recs:
            pid = _num(r[cols["ID"]])
            ip = dollarde(r[cols["IP"]]) if cols.get("IP") is not None else 0.0
            if pid is not None and ip > 0:
                pairs.append((int(pid), ip))
        pos_ip[pos] = pairs

    ws = wb["Fielding Ratings"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [_norm(h) for h in rows[1]]
    idx = {h: j for j, h in enumerate(hdr) if h}
    ratings = {}
    for r in rows[2:]:
        pid = _num(r[idx["ID"]]) if idx.get("ID") is not None and idx["ID"] < len(r) else None
        if pid is None:
            continue
        rec = {}
        for col in ("IF RNG", "IF ARM", "OF RNG"):
            if idx.get(col) is not None and idx[col] < len(r):
                rec[col] = _num(r[idx[col]])
        if idx.get("HT") is not None and idx["HT"] < len(r):
            rec["HT"] = height_cm(r[idx["HT"]])
        ratings[int(pid)] = rec
    wb.close()
    return pos_ip, ratings


def interp(knots, r):
    if r <= knots[0][0]:
        return knots[0][1]
    if r >= knots[-1][0]:
        return knots[-1][1]
    for (r1, v1), (r2, v2) in zip(knots, knots[1:]):
        if r1 <= r <= r2:
            return v1 + (v2 - v1) * (r - r1) / (r2 - r1)
    return 0.0


def fit_league(league):
    csv_dir = os.path.join(HERE, "calib", league)
    consts = json.load(open(os.path.join(csv_dir, "constants-latest.json")))["fielding"]
    dpx = json.load(open(os.path.join(HERE, "extracted",
                                      f"{league}_hitters_datapoints.json")))

    print(f"=== {league}: archive pools ===")
    bat = C.load_rows(os.path.join(csv_dir, "Batting.csv"))
    pit = C.load_rows(os.path.join(csv_dir, "Pitching.csv"))
    fld = C.add_fld_derived(C.load_rows(os.path.join(csv_dir, "Fielding.csv")))
    hitters = {r["ID"]: r for r in C.load_rows(os.path.join(csv_dir, "Hitters.csv"))
               if isinstance(r.get("ID"), float)}
    _, _, fld = C.drop_partial_final_season(bat, pit, fld)
    del bat, pit

    print(f"=== {league}: live population (25 Metadata.xlsx, read-only) ===")
    pos_ip, live_ratings = read_live(league)

    out = {"league": league,
           "fitted_at": datetime.datetime.now().isoformat(timespec="seconds"),
           "source": f"engine/fielding_curves_fit.py over calib/{league} archive "
                     "+ 25 Metadata.xlsx live population (offsets)",
           "positions": {}}
    worst_fid = 0.0
    gate_worst = {}
    for pos, (code, topn, rngcol, x2col, x2live, tcell, rppcell, elig_min) in POS_META.items():
        agg, _ = C.group_sum(fld, "player_id", C.FLD_STATS,
                             where=lambda r, c=code: r.get("position") == c)
        vis = sorted(agg, key=lambda i: (-agg[i]["ip"], i))[:topn]
        gt = C.totals(agg, vis, C.FLD_STATS)
        pmgt = gt["pm"] / gt["pa"]
        rngb = C.rating_mean(hitters, vis, rngcol)
        ys = [agg[i]["pm"] / agg[i]["pa"] - pmgt for i in vis]
        xs1 = [hitters[i][rngcol] - rngb for i in vis]
        if x2col:
            x2b = C.rating_mean(hitters, vis, x2col)
            xs2 = [hitters[i][x2col] - x2b for i in vis]
            b, m1, m2 = C.linest(ys, [xs1, xs2])
        else:
            x2b = None
            b, m1 = C.linest(ys, [xs1])
            m2 = None
        # pool fidelity vs constants-latest.json
        want = consts[f"{pos} PM%"]
        for got, w in ((b, want["intercept"]), (m1, want["x"]),
                       (m2, want.get("x2"))):
            if w is None:
                continue
            worst_fid = max(worst_fid, abs(got - w) / (1e-9 + abs(w)))

        # opportunity-weighted adjusted buckets -> weighted PAVA
        aggb = {}
        for i in vis:
            r = hitters[i][rngcol]
            y = agg[i]["pm"] / agg[i]["pa"]
            if m2 is not None:
                y -= m2 * (hitters[i][x2col] - x2b)
            rung = round(r / 5) * 5
            a = aggb.setdefault(rung, [0, 0.0, 0.0])
            a[0] += 1
            a[1] += agg[i]["pa"]
            a[2] += y * agg[i]["pa"]
        rungs = sorted(aggb)
        emp = {r: aggb[r][2] / aggb[r][1] for r in rungs}
        iso = pava([emp[r] for r in rungs], [aggb[r][1] for r in rungs], True)
        knots = [[float(r), v - pmgt] for r, v in zip(rungs, iso)]

        # live transport offset: ip-weighted mean of curve(RNG) + x2*m2 over
        # the live position population (rated players; innings from the
        # metadata Fielding Data tab)
        wsum = acc = 0.0
        skipped_ip = 0.0
        for pid, ip in pos_ip.get(pos, []):
            rec = live_ratings.get(pid) or {}
            r = rec.get(rngcol if rngcol in ("IF RNG", "OF RNG") else rngcol)
            x2v = rec.get(x2live) if x2live else None
            if r is None or (x2live and x2v is None):
                skipped_ip += ip
                continue
            v = interp(knots, r) + (x2v * m2 if m2 is not None else 0.0)
            wsum += ip
            acc += ip * v
        offset = acc / wsum if wsum else 0.0
        cover = wsum / (wsum + skipped_ip) if (wsum + skipped_ip) else 0.0

        # ---- gate: eligibility-band engine-vs-sim error matrix ----
        T = float(dpx[tcell])
        rpp = float(dpx[rppcell])
        band = [r for r in rungs if elig_min <= r <= SUPPORT[1]]
        worst = worst_lin = 0.0
        worst_pair = None
        for a_i, r1 in enumerate(band):
            for r2 in band[a_i + 1:]:
                sim = (emp[r2] - emp[r1]) * T * rpp
                eng = (interp(knots, r2) - interp(knots, r1)) * T * rpp
                lin = (m1 * (r2 - r1)) * T * rpp
                if abs(eng - sim) > worst:
                    worst, worst_pair = abs(eng - sim), (r1, r2)
                worst_lin = max(worst_lin, abs(lin - sim))
        gate_worst[pos] = (worst, worst_pair, worst_lin)

        out["positions"][pos] = {
            "rng": rngcol, "x2": x2live, "m2": m2, "offset": offset,
            "knots": knots, "gt_pm": pmgt,
            "live_ip_coverage": cover, "plays_cell": tcell, "rpp_cell": rppcell,
            "elig_min_rng": elig_min,
            "report": {str(int(r)): {"n": aggb[r][0], "pa": aggb[r][1],
                                     "emp_adj": emp[r], "iso": v[1] + pmgt}
                       for r, v in zip(rungs, knots)},
        }
        print(f"  {pos}: n_pool={len(vis)} m2={'%.6g' % m2 if m2 is not None else '-'} "
              f"offset={offset:+.5f} liveIPcov={cover:.1%}")
        for r in rungs:
            print(f"      r={r:3.0f} n={aggb[r][0]:3d} pa={aggb[r][1]:9.0f}  "
                  f"emp={emp[r]:.4f}  iso={interp(knots, r) + pmgt:.4f}  "
                  f"lin={pmgt + m1 * (r - rngb) + b:.4f}")
        w, wp, wl = gate_worst[pos]
        print(f"      band>={elig_min}: worst |engine-sim| {w:.2f} runs "
              f"(pair {wp})   linear fit's worst: {wl:.2f} runs")
    print(f"  pool fidelity vs constants-latest.json: worst rel diff "
          f"{worst_fid:.2e} ({'OK' if worst_fid < 1e-6 else '!! POOLS DIFFER'})")
    out["gate_matrix_worst_runs"] = {p: {"piecewise": g[0], "pair": list(g[1] or ()),
                                         "linear": g[2]} for p, g in gate_worst.items()}
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--league", choices=["TGS", "BLM"])
    a = ap.parse_args()
    for lg in ([a.league] if a.league else ["TGS", "BLM"]):
        res = fit_league(lg)
        path = os.path.join(HERE, "calib", lg, "fielding_curves.json")
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(res, fh, indent=1)
        print(f"wrote {path}\n")


if __name__ == "__main__":
    main()
