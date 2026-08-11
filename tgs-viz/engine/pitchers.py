"""
Pitcher projection engine -- a faithful, literal port of "The Sheet" Pitchers tab.

It computes every output column (per-BF counting stats -> wOBA-against ->
RA/9 -> WAA/WAR) from a pitcher's rating inputs plus the league's lifted
Data Points / Filters / Ballparks constants. Four role blocks are produced:
  SP   (Starter present-day)     gate: Starter  AND Eligible
  RP   (Reliever present-day)    gate: Eligible
  P    (Starter potential)       gate: Starter P AND Eligible AND Age<24
  P RP (Reliever potential)      gate: Eligible AND Age<24

NOTHING here writes to any workbook. Constants are read read-only.

Run directly to VALIDATE against the sheet's own cached values:
    python tgs-viz/engine/pitchers.py [LEAGUE]      (default TGS)
It computes from rating inputs and reports the max abs diff per output column.
This is the fidelity gate: the port is only trusted when diffs are ~0.
"""
import math, os, sys, re, json
from openpyxl import load_workbook

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ---------- constants loader (read-only) ----------
def _detect_park_row(path):
    """The Ballparks AA park-factor cell sits at a different row per league
    (TGS row 41, BLM row 39). Read it from the wOBA vR formula's Ballparks ref."""
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb["Pitchers"]
    hdr = frow = None
    for i, row in enumerate(ws.iter_rows()):
        v = [c.value for c in row]
        if i == 0:
            hdr = [str(x).strip() if x is not None else "" for x in v]
        elif i == 1:
            frow = v
            break
    wb.close()
    m = re.search(r"Ballparks!\$AA\$(\d+)", str(frow[hdr.index("wOBA vR")]))
    return int(m.group(1)) if m else 41


def scan_consts(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    dp, filt = {}, {}
    for row in wb["Data Points"].iter_rows():
        for c in row:
            if c.value is not None:
                dp[c.coordinate] = c.value
    for row in wb["Filters"].iter_rows():
        for c in row:
            if c.value is not None:
                filt[c.coordinate] = c.value
    bp = {}
    for row in wb["Ballparks"].iter_rows(min_row=30, max_row=50):
        for c in row:
            if c.value is not None:
                bp[c.coordinate] = c.value
    wb.close()
    row = _detect_park_row(path)
    park_aa = float(bp[f"AA{row}"]) if f"AA{row}" in bp and bp[f"AA{row}"] not in ("", None) else 1.0
    return dp, filt, park_aa


def num(v):
    """Coerce a cell value to float, or None if blank/non-numeric."""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s == "-":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def load_scurves(path):
    """Load a scurve_fit.py scurves-preview.json -> the `scurves` dict compute()
    accepts: {"SP": {block: params}, "RP": {block: params}}. OPT-IN ONLY — the
    live pipeline never passes this, so it keeps today's two-segment output."""
    data = json.load(open(path, encoding="utf-8"))
    return {role: data["roles"][role]["blocks"] for role in ("SP", "RP")}


def load_currency(path):
    """Load a currency_fit.py calib/<LEAGUE>/currency.json (audit D2/D9): the
    archive-FITTED cross-type currency layer. OPT-IN ONLY — same contract as
    load_scurves: the live pipeline passes it, the sheet-fidelity validator
    below never does, so validation keeps matching the workbook bit-for-bit."""
    return json.load(open(path, encoding="utf-8"))


def compute(p, dp, filt, park_aa, scurves=None, currency=None):
    """p: dict of rating inputs + meta (incl. _row). Returns dict of computed outputs.

    Faithful port: every branch / quirk mirrors the sheet's array formulas.

    scurves (OPT-IN, audit D1): when given ({"SP"/"RP": {"SO"/"uBB"/"HR"/"HHR":
    {A,B,k,m,offset,support}}}), those four rate blocks are evaluated as the
    fitted+transported logistic  A + B/(1+e^(-k(r-m))) + offset  (rating clamped
    to the fitted support) instead of the two-segment anchor lines. Everything
    downstream (XBH/3B split, SBAT, wOBA weights, RA/9, WAA) is unchanged.
    When scurves is None (default) the output is bit-identical to the sheet.

    currency (OPT-IN, audit D2 + D9 — INTENTIONAL divergence from the sheet
    when passed): a currency_fit.py dict (calib/<LEAGUE>/currency.json). Two
    effects here:
      * ra9_exponent {"SP": e, "RP": e} — the wOBA->RA/9 conversion becomes
        (wOBA/lg)^e instead of the sheet's ^2. FITTED per league per role from
        the clone archives (log-log, BF-weighted; octile calibration table is
        flat under the fit and tilted 0.97->1.04 under ^2). The sheet's ^2
        under-spreads pitcher run impact ~12-15% — the pitcher half of the
        hitter<->pitcher currency skew (audit D2). League-average RA/9 is
        unchanged (the ratio is 1 at the anchor).
      * pitcher_cells {"H30": rpw, ...} — Data Points VALUE overrides. H30 is
        runs-per-win: the sheet's 9*(R/IP)*1.5+3 tangent formula sits below the
        archive's own fitted wins-on-run-diff regression; one FITTED RPW per
        league ships (mid-range spec = GM of the two OLS directions, audit D9).
    When currency is None (default) the output is bit-identical to the sheet.
    """
    if currency:
        cells = currency.get("pitcher_cells") or {}
        if cells:
            dp = {**dp, **cells}
    exps = (currency or {}).get("ra9_exponent") or {}
    ra9_exp = {"SP": float(exps.get("SP", 2.0)), "RP": float(exps.get("RP", 2.0))}
    g = lambda k: float(dp[k])          # Data Points (strict)
    # blank-tolerant Data Points accessor (Excel treats blank SUMPRODUCT cells as 0)
    g0 = lambda k: float(dp[k]) if k in dp and isinstance(dp.get(k), (int, float)) else 0.0
    f = lambda k: float(filt[k]) if k in filt and filt[k] not in ("", None) else None

    T = (p.get("T") or "").strip()      # throws R/L/S
    POS = (p.get("POS") or "").strip()
    is_sp = POS == "SP"
    row = p.get("_row", 2)
    EPS = row / 10000000.0

    H31, H32 = g("H31"), g("H32")        # BF (SP), BF RP
    out = {}

    # SUMPRODUCT helpers over a 4-wide constant vector (x^0..x^3); blanks -> 0
    def sp4(x, b, c, d, e):
        return g0(b) + g0(c) * x + g0(d) * x ** 2 + g0(e) * x ** 3

    # ---------------- generic role stat-line builder ----------------
    # cfg carries all the row/col anchors + league-adj cells that differ
    # between SP and RP blocks. handed picks HR/H-HR handedness multipliers.
    def statline(CON, STU, HRR, PBABIP, BF, cfg, handed):
        # opt-in fitted S-curves for this cfg's regression block (audit D1)
        sc = (scurves or {}).get(cfg.get("role")) or {}

        def sc_rate(blk, r):
            c = sc[blk]
            lo, hi = c.get("support", (20.0, 80.0))
            rr = min(max(r, lo), hi)
            v = c["A"] + c["B"] / (1.0 + math.exp(-c["k"] * (rr - c["m"]))) + c["offset"]
            kn = c.get("knot")
            if kn:
                # elite-K hinge (D1 tail, SO block): the fitted 75->80 K% jump
                # the logistic saturates below. Slope is live-fitted (hinge_hi=80)
                # or archive-transported (hinge_hi=narrowed support) — see
                # scurve_fit.py; s >= 0 keeps it monotone. Uses the RAW rating
                # (capped at kn["hi"]) so the hinge can keep climbing where the
                # drift-narrowed logistic support has already saturated.
                v += kn["s"] * max(0.0, min(r, kn["hi"]) - kn["r"])
            return v

        # HBP
        HBP = g(cfg["HBP_k"]) * BF
        # uBB (CON): anchor, slope hi/lo, int hi/lo, +Kadj
        if "uBB" in sc:
            uBB = sc_rate("uBB", CON) * (BF - HBP)
        elif CON >= 50:
            uBB = ((CON - g(cfg["uBB_anchor"])) * g(cfg["uBB_sh"]) + g(cfg["uBB_ih"]) + g(cfg["uBB_k"])) * (BF - HBP)
        else:
            uBB = ((CON - g(cfg["uBB_anchor"])) * g(cfg["uBB_sl"]) + g(cfg["uBB_il"]) + g(cfg["uBB_k"])) * (BF - HBP)
        uBB = max(uBB, 0.0)
        # SO (STU): POS adjusts the rating used. The sheet's RP/P-RP <50 branch
        # uses the *bare* STU (cfg stu_lo_adj flags it) — that was audit bug B6;
        # all cfgs now set stu_lo_adj=True so the adjustment holds in both branches.
        stu_adj = STU + cfg["stu_delta"] if is_sp else STU + cfg["stu_delta_rp"]
        if "SO" in sc:
            SO = sc_rate("SO", stu_adj) * (BF - uBB - HBP)
        elif stu_adj >= 50:
            SO = ((stu_adj - g(cfg["SO_anchor"])) * g(cfg["SO_sh"]) + g(cfg["SO_ih"]) + g(cfg["SO_k"])) * (BF - uBB - HBP)
        else:
            stu_lo = stu_adj if cfg["stu_lo_adj"] else STU
            SO = ((stu_lo - g(cfg["SO_anchor"])) * g(cfg["SO_sl"]) + g(cfg["SO_il"]) + g(cfg["SO_k"])) * (BF - uBB - HBP)
        SO = max(SO, 0.0)
        # HR (HRR): + handedness multiplier
        hr_mult = handed["hr"]
        if "HR" in sc:
            HR = sc_rate("HR", HRR) * (BF - uBB - HBP) * hr_mult
        elif HRR >= 50:
            HR = ((HRR - g(cfg["HR_anchor"])) * g(cfg["HR_sh"]) + g(cfg["HR_ih"]) + g(cfg["HR_k"])) * (BF - uBB - HBP) * hr_mult
        else:
            HR = ((HRR - g(cfg["HR_anchor"])) * g(cfg["HR_sl"]) + g(cfg["HR_il"]) + g(cfg["HR_k"])) * (BF - uBB - HBP) * hr_mult
        HR = max(HR, 0.0)
        # H-HR (PBABIP): + babip handedness multiplier
        bab_mult = handed["bab"]
        rem = BF - HBP - uBB - SO - HR
        if "HHR" in sc:
            HHR = sc_rate("HHR", PBABIP) * rem * bab_mult
        elif PBABIP >= 50:
            HHR = ((PBABIP - g(cfg["HH_anchor"])) * g(cfg["HH_sh"]) + g(cfg["HH_ih"]) + g(cfg["HH_k"])) * rem * bab_mult
        else:
            HHR = ((PBABIP - g(cfg["HH_anchor"])) * g(cfg["HH_sl"]) + g(cfg["HH_il"]) + g(cfg["HH_k"])) * rem * bab_mult
        HHR = max(HHR, 0.0)
        # XBH-HR = H-HR * Kxbh * Filters!E9 ; 3B = XBH * K3b * Filters!E10
        XBH = HHR * g(cfg["xbh_k"]) * f("E9")
        T3B = XBH * g(cfg["t3b_k"]) * f("E10")
        D2B = XBH - T3B
        S1B = HHR - XBH
        # SBAT / SB% (HLD): SUMPRODUCT((HLD-anchor)^{0..3}, vec) + Kadj
        HLD = p["HLD"]
        SBAT = max((sp4(HLD - g(cfg["hld_anchor"]), "B25", "C25", "D25", "E25") + g(cfg["sbat_k"]))
                   * (uBB + HBP + S1B), 0.0)
        SBpct = min(sp4(HLD - g(cfg["hld_anchor"]), "B23", "C23", "D23", "E23") + g(cfg["sbpct_k"]), 1.0)
        SB = SBpct * SBAT
        CS = SBAT - SB
        # wOBA: HBP,uBB,1B,2B,3B,HR,SB,CS weighted, / BF / park
        w = cfg["woba_w"]  # list of 8 DP cells
        wOBA = (HBP * g(w[0]) + uBB * g(w[1]) + S1B * g(w[2]) + D2B * g(w[3]) + T3B * g(w[4])
                + HR * g(w[5]) + SB * g(w[6]) + CS * g(w[7])) / BF / park_aa
        # RA/9 = (wOBA/lg-role-wOBA)^exp * RA9base. The sheet hardcodes exp=2;
        # the currency layer (audit D2) replaces it with the archive-fitted
        # per-role exponent. cfg["scale"] is the role's league wOBA (I31/I29).
        RA9 = (wOBA / g(cfg["scale"])) ** ra9_exp[cfg["role"]] * g(cfg["ra9_base"])
        return dict(HBP=HBP, uBB=uBB, SO=SO, HR=HR, HHR=HHR, XBH=XBH, T3B=T3B, D2B=D2B,
                    S1B=S1B, SBAT=SBAT, SBpct=SBpct, SB=SB, CS=CS, wOBA=wOBA, RA9=RA9)

    def waa(ra9, ip, ra9_base, waa_const):
        return ((ra9_base - ra9) * (ip / 9.0)) / waa_const + EPS

    # ================= SP block (present-day starter) =================
    SP_CFG = dict(
        role="SP",
        HBP_k="K10",
        uBB_anchor="H5", uBB_sh="C3", uBB_ih="B3", uBB_sl="E3", uBB_il="D3", uBB_k="K3",
        SO_anchor="H2", SO_sh="C7", SO_ih="B7", SO_sl="E7", SO_il="D7", SO_k="K5",
        stu_delta=0, stu_delta_rp=-5, stu_lo_adj=True,
        HR_anchor="H3", HR_sh="C5", HR_ih="B5", HR_sl="E5", HR_il="D5", HR_k="K4",
        HH_anchor="H4", HH_sh="C9", HH_ih="B9", HH_sl="E9", HH_il="D9", HH_k="K6",
        xbh_k="K7", t3b_k="K8",
        hld_anchor="I2", sbat_k="K12", sbpct_k="K9",
        woba_w=["H12", "H13", "H14", "H15", "H16", "H17", "H18", "H19"],
        scale="I31", ra9_base="H41",
    )
    # handedness: HR uses Filters C11 (vR) / D11 (vL); H-HR uses C8 (vR) / D8 (vL)
    sp_R = statline(p["CON vR"], p["STU vR"], p["HRR vR"], p["PBABIP vR"], H31, SP_CFG,
                    handed=dict(hr=f("C11"), bab=f("C8")))
    sp_L = statline(p["CON vL"], p["STU vL"], p["HRR vL"], p["PBABIP vL"], H31, SP_CFG,
                    handed=dict(hr=f("D11"), bab=f("D8")))

    def emit(prefix, sR, sL, woba_share_func, ip, ra9_base, waa_const, scale, exp=2.0):
        for suf, s in (("vR", sR), ("vL", sL)):
            out[f"HBP {suf}{prefix}"] = s["HBP"]; out[f"uBB {suf}{prefix}"] = s["uBB"]
            out[f"SO {suf}{prefix}"] = s["SO"]; out[f"HR {suf}{prefix}"] = s["HR"]
            out[f"H-HR {suf}{prefix}"] = s["HHR"]; out[f"XBH-HR {suf}{prefix}"] = s["XBH"]
            out[f"3B {suf}{prefix}"] = s["T3B"]; out[f"2B {suf}{prefix}"] = s["D2B"]
            out[f"1B {suf}{prefix}"] = s["S1B"]; out[f"SBAT {suf}{prefix}"] = s["SBAT"]
            out[f"SB% {suf}{prefix}"] = s["SBpct"]; out[f"SB {suf}{prefix}"] = s["SB"]
            out[f"CS {suf}{prefix}"] = s["CS"]; out[f"wOBA {suf}{prefix}"] = s["wOBA"]
            out[f"RA/9 {suf}{prefix}"] = s["RA9"]
            out[f"WAA {suf}{prefix}"] = waa(s["RA9"], ip, ra9_base, waa_const)
        # wtd by throws-hand platoon share
        wtd_woba = woba_share_func(sR["wOBA"], sL["wOBA"])
        out[f"wOBA wtd{prefix}"] = wtd_woba
        ra9_wtd = (wtd_woba / scale) ** exp * ra9_base   # exp: currency layer (D2)
        out[f"RA/9 wtd{prefix}"] = ra9_wtd
        out[f"WAA wtd{prefix}"] = waa(ra9_wtd, ip, ra9_base, waa_const)

    # platoon weighting by T (throws): R->H24, L->H23, S->H25 ; wtd = vL*(1-share)+vR*share
    def share_by_T(vr, vl):
        share = {"R": g("H24"), "L": g("H23"), "S": g("H25")}.get(T, g("H24"))
        return vl * (1 - share) + vr * share

    emit("", sp_R, sp_L, share_by_T, g("H33"), g("H41"), g("H30"), g("I31"), ra9_exp["SP"])

    # Starter gate (sheet 'Starter' rule): a pitcher who doesn't qualify to start
    # gets NO starter projection — the sheet leaves those cells blank.
    spp = p.get("SP P Pitch") or 0
    npitch = p.get("Pitches") or 0
    stm = p.get("STM") or 0
    STARTER_MIN_STM = 35   # user-tuned: the sheet's 'Starter' rule uses 40; lowered to 35 so
                           # 35-stamina arms still qualify to start (engine intentionally diverges
                           # from the sheet here — change the sheet 'Starter'/'Starter P' formulas
                           # to >=35 too if you ever extract straight from the workbook).
    starter = ((spp >= 3) or (spp >= 2 and npitch >= 3) or (spp >= 1 and npitch >= 5)) and stm >= STARTER_MIN_STM
    out["Starter"] = starter
    if not starter:
        for suf in ("vR", "vL"):
            for c in ("HBP", "uBB", "SO", "HR", "H-HR", "XBH-HR", "3B", "2B", "1B",
                      "SBAT", "SB%", "SB", "CS", "wOBA", "RA/9", "WAA"):
                out[f"{c} {suf}"] = None
        out["wOBA wtd"] = out["RA/9 wtd"] = out["WAA wtd"] = None

    # ================= RP block (present-day reliever) =================
    RP_CFG = dict(
        role="RP",
        HBP_k="K22",
        uBB_anchor="H10", uBB_sh="C14", uBB_ih="B14", uBB_sl="E14", uBB_il="D14", uBB_k="K15",
        SO_anchor="H7", SO_sh="C18", SO_ih="B18", SO_sl="E18", SO_il="D18", SO_k="K17",
        # audit B6 (INTENTIONAL divergence from the sheet): the sheet's RP <50 branch
        # drops the +5 starter-STU bonus (stu_lo_adj was False here) while its SP-block
        # mirror keeps the -5 in both branches, proving intent. A starter at STU 40 lost
        # 5·E18 of K% and his relief floor cliffed at the 50 crossing; the engine now
        # applies the adjusted STU in BOTH branches.
        stu_delta=5, stu_delta_rp=0, stu_lo_adj=True,
        HR_anchor="H8", HR_sh="C16", HR_ih="B16", HR_sl="E16", HR_il="D16", HR_k="K16",
        HH_anchor="H9", HH_sh="C20", HH_ih="B20", HH_sl="E20", HH_il="D20", HH_k="K18",
        xbh_k="K19", t3b_k="K20",
        hld_anchor="I7", sbat_k="K24", sbpct_k="K9",
        woba_w=["K29", "K30", "K31", "K32", "K33", "K34", "K35", "K36"],
        scale="I29", ra9_base="H42",
    )
    rp_R = statline(p["CON vR"], p["STU vR"], p["HRR vR"], p["PBABIP vR"], H32, RP_CFG,
                    handed=dict(hr=f("C11"), bab=f("C8")))
    # audit B11 (INTENTIONAL divergence from the sheet): the sheet's RP vL HR/H-HR
    # formulas reference the vR handedness filters (C11/C8); the engine uses the true
    # vL multipliers D11/D8 — same as the SP block's vL line (+3.8% HR-vL in BLM fixed).
    rp_L = statline(p["CON vL"], p["STU vL"], p["HRR vL"], p["PBABIP vL"], H32, RP_CFG,
                    handed=dict(hr=f("D11"), bab=f("D8")))
    emit(" RP", rp_R, rp_L, share_by_T, g("H34"), g("H42"), g("H30"), g("I29"), ra9_exp["RP"])

    # ================= P block (starter potential, single line) =================
    # Uses the SP regression layout but P-ratings and a *blended* handedness mult.
    def blended_hr():
        if T == "R":
            return f("C11") * g("H24") + f("D11") * (1 - g("H24"))
        if T == "L":
            return f("C11") * g("H23") + f("D11") * (1 - g("H23"))
        return f("E11")

    def blended_bab():
        if T == "R":
            return f("C8") * g("H24") + f("D8") * (1 - g("H24"))
        if T == "L":
            return f("C8") * g("H23") + f("D8") * (1 - g("H23"))
        return f("E11")  # sheet quirk: switch-throwers fall back to E11 here

    have_p = all(p.get(k) is not None for k in ("CON P", "STU P", "HRR P", "PBABIP P"))
    if have_p:
        sp_P = statline(p["CON P"], p["STU P"], p["HRR P"], p["PBABIP P"], H31, SP_CFG,
                        handed=dict(hr=blended_hr(), bab=blended_bab()))
        for k, src in (("HBP P", "HBP"), ("uBB P", "uBB"), ("SO P", "SO"), ("HR P", "HR"),
                       ("H-HR P", "HHR"), ("XBH-HR P", "XBH"), ("3B P", "T3B"), ("2B P", "D2B"),
                       ("1B P", "S1B"), ("SBAT P", "SBAT"), ("SB% P", "SBpct"), ("SB P", "SB"),
                       ("CS P", "CS"), ("wOBA P", "wOBA"), ("RA/9 P", "RA9")):
            out[k] = sp_P[src]
        out["WAP"] = waa(sp_P["RA9"], g("H33"), g("H41"), g("H30"))
        if not starter:   # non-starter -> no starter-potential projection either
            for k in ("HBP P", "uBB P", "SO P", "HR P", "H-HR P", "XBH-HR P", "3B P", "2B P",
                      "1B P", "SBAT P", "SB% P", "SB P", "CS P", "wOBA P", "RA/9 P", "WAP"):
                out[k] = None

        # P RP block (reliever potential)
        rp_P = statline(p["CON P"], p["STU P"], p["HRR P"], p["PBABIP P"], H32, RP_CFG,
                        handed=dict(hr=blended_hr(), bab=blended_bab()))
        for k, src in (("HBP P RP", "HBP"), ("uBB P RP", "uBB"), ("SO P RP", "SO"), ("HR P RP", "HR"),
                       ("H-HR P RP", "HHR"), ("XBH-HR P RP", "XBH"), ("3B P RP", "T3B"), ("2B P RP", "D2B"),
                       ("1B P RP", "S1B"), ("SBAT P RP", "SBAT"), ("SB% P RP", "SBpct"), ("SB P RP", "SB"),
                       ("CS P RP", "CS"), ("wOBA P RP", "wOBA"), ("RA/9 P RP", "RA9")):
            out[k] = rp_P[src]
        out["WAP RP"] = waa(rp_P["RA9"], g("H34"), g("H42"), g("H30"))
    return out


# ---------- validation ----------
RATING_COLS = ["STU P", "HRR P", "PBABIP P", "CON P", "STU vR", "HRR vR", "PBABIP vR", "CON vR",
               "STU vL", "HRR vL", "PBABIP vL", "CON vL", "STM", "HLD", "Pitches",
               "SP Pitch", "SP P Pitch", "B", "T", "POS", "Name", "Age"]
META = {"B", "T", "POS", "Name"}
OPTIONAL = {"STU P", "HRR P", "PBABIP P", "CON P", "SP P Pitch"}

# Output columns to validate (every computed numeric column in the sheet)
SP_COLS = ["HBP vR", "uBB vR", "SO vR", "HR vR", "H-HR vR", "XBH-HR vR", "3B vR", "2B vR", "1B vR",
           "SBAT vR", "SB% vR", "SB vR", "CS vR", "wOBA vR", "RA/9 vR", "WAA vR",
           "HBP vL", "uBB vL", "SO vL", "HR vL", "H-HR vL", "XBH-HR vL", "3B vL", "2B vL", "1B vL",
           "SBAT vL", "SB% vL", "SB vL", "CS vL", "wOBA vL", "RA/9 vL", "WAA vL",
           "wOBA wtd", "RA/9 wtd", "WAA wtd"]
RP_COLS = [c + " RP" for c in
           ["HBP vR", "uBB vR", "SO vR", "HR vR", "H-HR vR", "XBH-HR vR", "3B vR", "2B vR", "1B vR",
            "SBAT vR", "SB% vR", "SB vR", "CS vR", "wOBA vR", "RA/9 vR", "WAA vR",
            "HBP vL", "uBB vL", "SO vL", "HR vL", "H-HR vL", "XBH-HR vL", "3B vL", "2B vL", "1B vL",
            "SBAT vL", "SB% vL", "SB vL", "CS vL", "wOBA vL", "RA/9 vL", "WAA vL",
            "wOBA wtd", "RA/9 wtd", "WAA wtd"]]
P_COLS = ["HBP P", "uBB P", "SO P", "HR P", "H-HR P", "XBH-HR P", "3B P", "2B P", "1B P",
          "SBAT P", "SB% P", "SB P", "CS P", "wOBA P", "RA/9 P", "WAP"]
PRP_COLS = ["HBP P RP", "uBB P RP", "SO P RP", "HR P RP", "H-HR P RP", "XBH-HR P RP", "3B P RP",
            "2B P RP", "1B P RP", "SBAT P RP", "SB% P RP", "SB P RP", "CS P RP", "wOBA P RP",
            "RA/9 P RP", "WAP RP"]
CHECK_COLS = SP_COLS + RP_COLS + P_COLS + PRP_COLS


def main():
    league = sys.argv[1] if len(sys.argv) > 1 else "TGS"
    path = os.path.join(REPO, f"The Sheets {league}", "The Sheet Pitchers.xlsx")
    dp, filt, park_aa = scan_consts(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Pitchers"]
    header = None
    diffs = {c: 0.0 for c in CHECK_COLS}
    counts = {c: 0 for c in CHECK_COLS}
    worst = {c: None for c in CHECK_COLS}
    n = 0
    for ri, row in enumerate(ws.iter_rows()):
        vals = [c.value for c in row]
        if header is None:
            header = [str(v).strip() if v is not None else "" for v in vals]
            idx = {h: i for i, h in enumerate(header)}
            continue
        rec = {h: vals[i] for h, i in idx.items() if i < len(vals)}
        # row number in the sheet (header is row 1 -> first data row is 2)
        sheet_row = ri + 1
        if not rec.get("Name") or rec.get("Eligible") not in (True, 1, "TRUE"):
            continue
        p = {"_row": sheet_row}
        ok = True
        for col in RATING_COLS:
            v = rec.get(col)
            if col in META:
                p[col] = v
            else:
                nv = num(v)
                p[col] = nv
                if nv is None and col not in OPTIONAL:
                    ok = False
        if not ok:
            continue
        try:
            out = compute(p, dp, filt, park_aa)
        except Exception:
            continue
        n += 1
        for col in CHECK_COLS:
            mine = out.get(col)
            sheet = num(rec.get(col))
            if mine is None or sheet is None:
                continue
            d = abs(mine - sheet)
            diffs[col] = max(diffs[col], d)
            counts[col] += 1
            if worst[col] is None or d > worst[col][0]:
                worst[col] = (d, rec.get("Name"), sheet, mine)
    wb.close()
    print(f"Validated {n} eligible pitchers in {league}  (park AA={park_aa})")
    print(f"{'column':14} {'n':>5} {'maxAbsDiff':>14}   worst-case (name: sheet vs mine)")
    worst_overall = 0.0
    for col in CHECK_COLS:
        w = worst[col]
        wt = f"{w[1]}: {w[2]:.6g} vs {w[3]:.6g}" if w else "-"
        flag = "  <-- MISMATCH" if diffs[col] > 1e-9 else ""
        worst_overall = max(worst_overall, diffs[col])
        print(f"{col:16} {counts[col]:5} {diffs[col]:14.6e}   {wt}{flag}")
    print(f"\nWORST OVERALL maxAbsDiff across all columns: {worst_overall:.6e}")
    print("PASS" if worst_overall < 1e-9 else "FAIL (see mismatches above)")


if __name__ == "__main__":
    main()
