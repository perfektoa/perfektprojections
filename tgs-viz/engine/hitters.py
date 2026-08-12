"""
Hitter projection engine — a faithful, literal port of "The Sheet" Hitters tab.

It computes every output column (counting stats -> wOBA/BatR -> fielding RunsP ->
baserunning -> WAA per position -> Max WAA) from the player's rating inputs plus
the league's lifted Data Points / Filters / Ballparks constants.

NOTHING here writes to any workbook. Constants are read read-only.

Run directly to VALIDATE against the sheet's own cached values:
    python tgs-viz/engine/hitters.py [LEAGUE]      (default TGS)
It computes from rating inputs and reports the max abs diff per output column.
This is the fidelity gate: the port is only trusted when diffs are ~0.
"""
import os, sys, json, re
from openpyxl import load_workbook

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ---------- constants loader (read-only) ----------
def _detect_park_base_row(path):
    """The Ballparks summary block sits at a different row per league
    (TGS row 40, BLM row 39). Read it from the wOBA formula's Ballparks ref."""
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb["Hitters"]
    hdr = frow = None
    for i, row in enumerate(ws.iter_rows()):
        v = [c.value for c in row]
        if i == 0:
            hdr = [str(x).strip() if x is not None else "" for x in v]
        elif i == 1:
            frow = v
            break
    wb.close()
    m = re.search(r"Ballparks!\$AA\$(\d+)", str(frow[hdr.index("wOBA vR")]))
    return int(m.group(1)) if m else 40


def scan_consts(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    dp, filt, bp = {}, {}, {}
    for row in wb["Data Points"].iter_rows():
        for c in row:
            if c.value is not None:
                dp[c.coordinate] = c.value
    for row in wb["Filters"].iter_rows():
        for c in row:
            if c.value is not None:
                filt[c.coordinate] = c.value
    for row in wb["Ballparks"].iter_rows(min_row=30, max_row=50):
        for c in row:
            if c.value is not None:
                bp[c.coordinate] = c.value
    wb.close()
    base = _detect_park_base_row(path)
    away = base + 3  # LH-batter ("away") row is 3 below the RH/home row
    fnum = lambda k: float(bp[k]) if k in bp and bp[k] not in ("", None) else None
    park = {
        "AA": fnum(f"AA{base}"), "AB": fnum(f"AB{base}"),
        "AG_home": fnum(f"AG{base}"), "AG_away": fnum(f"AG{away}"),
        "AH_home": fnum(f"AH{base}"), "AH_away": fnum(f"AH{away}"),
        "AI_home": fnum(f"AI{base}"), "AI_away": fnum(f"AI{away}"),
    }
    return dp, filt, park


def num(v):
    """Coerce a cell value to float, or None if blank/non-numeric."""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s == "-":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def piece(r, anchor, slope_hi, int_hi, slope_lo, int_lo):
    """Piecewise-linear regression: breakpoint at rating>=50, centered on anchor."""
    if r >= 50:
        return (r - anchor) * slope_hi + int_hi
    return (r - anchor) * slope_lo + int_lo


# League-specific SB% saturation cap (see the comment at its use site): measured from
# each league's calibration archive (career SB% by STE bucket, regulars).
SB_CAP = {"TGS": 0.88, "BLM": 0.888}


def _interp_knots(knots, r):
    """Piecewise-linear eval of sorted [[rating, value], ...]; flat beyond ends."""
    if r <= knots[0][0]:
        return knots[0][1]
    if r >= knots[-1][0]:
        return knots[-1][1]
    for (r1, v1), (r2, v2) in zip(knots, knots[1:]):
        if r1 <= r <= r2:
            return v1 + (v2 - v1) * (r - r1) / (r2 - r1)
    return 0.0


def compute(p, dp, filt, park, league="TGS", currency=None, tails=None, fielding=None):
    """p: dict of rating inputs + meta. Returns dict of computed outputs.
    `league` selects league-specific engine constants (currently the SB% cap).

    tails (OPT-IN, audit D4 — INTENTIONAL divergence from the sheet when
    passed): calib/<LEAGUE>/hitter_tails.json from hitter_tails_fit.py.
    Archive-MEASURED monotone corrections to the two-segment rate fits in the
    five proven-broken tail regions only (SO at K 65-80, HR at POW 70-80,
    BABIP at BA 20-35, XBH at both GAP tails, 3B/XBH at SPE 20-45): an
    additive rate delta that is identically 0 outside those regions, lands on the PA-weighted
    isotonic archive bucket means inside them, and holds its last value beyond
    rating 80. The sheet keeps its two-line formulas (formula changes are out
    of scope) — the SB%-cap precedent.

    fielding (OPT-IN, audit D3 — INTENTIONAL divergence from the sheet when
    passed): calib/<LEAGUE>/fielding_curves.json from fielding_curves_fit.py.
    Replaces the LINEAR range->PM% term per position with a monotone piecewise
    (isotonic) curve fitted on opportunity-weighted archive bucket means; the
    stored `offset` is the live position population's innings-weighted mean of
    (curve + x2*m2), so the league-mean PM-runs stays ~0 by construction (the
    Phase A anchor property). E%/DP/ARM/C blocks keep the sheet's linear math.

    currency (OPT-IN, audit D9 — INTENTIONAL divergence from the sheet when
    passed): a currency_fit.py dict (calib/<LEAGUE>/currency.json). On the
    hitter side only hitter_cells applies — today that is H30 (runs-per-win):
    one FITTED RPW per league from the archive wins-on-run-diff regression
    (mid-range spec), replacing the sheet's 9*(R/IP)*1.5+3 tangent value.
    NOTE — audit D2's hitter run-value replacement was fitted and then
    REFUSED BY ITS OWN GATE: the team-clone-year regression of actual runs on
    OFF gives slope 0.979 (TGS) / 0.979 (BLM) with the AS-BUILT weights
    (already inside the 1.00±0.03 gate; live-2044 frame: 1.052±0.04), while
    substituting the archive-fitted per-event run values into the wOBA
    construction moves the slope AWAY from 1 (0.944/0.962). So the wOBA
    weights H12..H17/H20/H29 deliberately stay the sheet's own — see
    currency_fit.py run_values report for the full numbers.
    When currency is None (default) the output is bit-identical to the sheet.
    """
    if currency:
        cells = currency.get("hitter_cells") or {}
        if cells:
            dp = {**dp, **cells}
    tblocks = (tails or {}).get("blocks") or {}

    def tadj(blk, r):
        c = tblocks.get(blk)
        return _interp_knots(c["knots"], r) if c and c.get("knots") else 0.0

    fpos = (fielding or {}).get("positions") or {}

    def pm_pw(pos, rng, x2val):
        c = fpos[pos]
        v = _interp_knots(c["knots"], rng) - c["offset"]
        if c.get("m2") is not None and x2val is not None:
            v += x2val * c["m2"]
        return v
    g = lambda k: float(dp[k])          # Data Points (strict)
    g0 = lambda k: float(dp[k]) if k in dp and dp[k] not in ("", None) else 0.0  # blank -> 0
    f = lambda k: float(filt[k]) if k in filt and filt[k] not in ("", None) else None
    B = (p.get("B") or "").strip()      # bats R/L/S
    PA = g("H31")
    out = {}

    # ---- per-split counting stats ----
    def split_stats(suffix, EYE, POW, K, BA, GAP, SPE):
        def ha(col):
            home, away = park[col + "_home"], park[col + "_away"]
            return (home if B == "R" else away) if suffix == "vR" else (away if B == "L" else home)
        ah, ag, ai = ha("AH"), ha("AG"), ha("AI")
        HBP = g("H37") * PA
        uBB = max((piece(EYE, g("H2"), g("C3"), g("B3"), g("E3"), g("D3")) + g("C33")) * (PA - HBP), 0.0)
        # HR handedness multiplier (depends on split + bats)
        if suffix == "vR":
            hr_hand = f("C11") if B == "R" else f("D11")
        else:
            hr_hand = f("D11") if B == "L" else f("C11")
        # audit D4: tadj() adds the archive-measured tail delta (0 outside the
        # audited regions) to the HR/SO/H-HR/XBH rate fits.
        HR = max((piece(POW, g("H3"), g("C5"), g("B5"), g("E5"), g("D5")) + g("C34")
                  + tadj("HR", POW)) * (PA - HBP - uBB) * hr_hand, 0.0)
        SO = max((piece(K, g("H4"), g("C7"), g("B7"), g("E7"), g("D7")) + g("C35")
                  + tadj("SO", K)) * (PA - HBP - uBB), 0.0)
        # H-HR (contact, BA rating); >=50 adds ballpark, <50 applies handedness
        rem = PA - HBP - uBB - HR - SO
        if BA >= 50:
            HHR = (BA - g("H5")) * g("C9") + g("B9") + g("C36") + tadj("HHR", BA)
            HHR = HHR * rem + ah
        else:
            HHR = (BA - g("H5")) * g("E9") + g("D9") + g("C36") + tadj("HHR", BA)
            c3 = filt.get("C3")
            if c3 in (None, ""):
                hand = 1.0
            else:
                hand = {"R": f("C8"), "L": f("D8")}.get(B, f("C8") * g("H25") + f("D8") * (1 - g("H25")))
            HHR = HHR * rem * hand
        HHR = max(HHR, 0.0)
        # XBH-HR (GAP rating); >=50 adds ballpark AG, <50 * Filters C9
        if GAP >= 50:
            XBH = ((GAP - g("H6")) * g("C11") + g("B11") + g("C37") + tadj("XBH", GAP)) * HHR + ag
        else:
            XBH = ((GAP - g("H6")) * g("E11") + g("D11") + g("C37") + tadj("XBH", GAP)) * HHR * f("C9")
        XBH = max(XBH, 0.0)
        # 3B (SPE rating); >=50 adds ballpark AI, <50 * Filters C10
        # tadj("T3B") is identically 0 at and above SPE 50 (region is lo/50)
        if SPE >= 50:
            T3B = ((SPE - g("H7")) * g("C13") + g("B13") + g("C38") + tadj("T3B", SPE)) * XBH + ai
        else:
            T3B = ((SPE - g("H7")) * g("E13") + g("D13") + g("C38") + tadj("T3B", SPE)) * XBH * f("C10")
        T3B = max(T3B, 0.0)
        D2B = XBH - T3B
        S1B = HHR - XBH
        return dict(HBP=HBP, uBB=uBB, HR=HR, SO=SO, HHR=HHR, XBH=XBH, T3B=T3B, D2B=D2B, S1B=S1B)

    sR = split_stats("vR", p["EYE vR"], p["POW vR"], p["K vR"], p["BA vR"], p["GAP vR"], p["SPE"])
    sL = split_stats("vL", p["EYE vL"], p["POW vL"], p["K vL"], p["BA vL"], p["GAP vL"], p["SPE"])
    for suf, s in (("vR", sR), ("vL", sL)):
        out[f"HBP {suf}"] = s["HBP"]; out[f"uBB {suf}"] = s["uBB"]; out[f"HR {suf}"] = s["HR"]
        out[f"SO {suf}"] = s["SO"]; out[f"H-HR {suf}"] = s["HHR"]; out[f"XBH-HR {suf}"] = s["XBH"]
        out[f"3B {suf}"] = s["T3B"]; out[f"2B {suf}"] = s["D2B"]; out[f"1B {suf}"] = s["S1B"]

    # ---- wOBA / OBP / BatR ----
    def woba(s):
        return (s["HBP"] * g("H12") + s["uBB"] * g("H13") + s["S1B"] * g("H14") + s["D2B"] * g("H15")
                + s["T3B"] * g("H16") + s["HR"] * g("H17")) / PA / park["AA"]

    def obp(s):
        return (s["HBP"] + s["uBB"] + s["HR"] + s["HHR"]) / PA

    out["wOBA vR"], out["wOBA vL"] = woba(sR), woba(sL)
    out["OBP vR"], out["OBP vL"] = obp(sR), obp(sL)
    # split-weighted (wtd) by bats
    share = {"R": g("H24"), "L": g("H23"), "S": g("H25")}.get(B, g("H24"))
    wtd = lambda vr, vl: vl * (1 - share) + vr * share
    out["wOBA wtd"] = wtd(out["wOBA vR"], out["wOBA vL"])
    out["OBP wtd"] = wtd(out["OBP vR"], out["OBP vL"])
    for suf in ("vR", "vL", "wtd"):
        out[f"BatR {suf}"] = ((out[f"wOBA {suf}"] - g("H29")) / g("H20")) * PA

    # DH wOBA (small SO-based discount) + DH BatR
    def dh_woba(s):
        return ((s["HBP"] * g("H12") + s["uBB"] * g("H13") + s["S1B"] * g("H14") + s["D2B"] * g("H15")
                 + s["T3B"] * g("H16")) * 0.98 + s["HR"] * g("H17")) / (PA - s["SO"] * 0.02) / park["AA"]
    out["DH wOBA vR"], out["DH wOBA vL"] = dh_woba(sR), dh_woba(sL)
    out["DH wOBA wtd"] = wtd(out["DH wOBA vR"], out["DH wOBA vL"])
    for suf in ("vR", "vL", "wtd"):
        out[f"DH BatR {suf}"] = ((out[f"DH wOBA {suf}"] - g("H29")) / g("H20")) * PA

    # ---- baserunning ----
    # SB% is itself a cubic in STE (capped at 80) — compute it (don't read it),
    # so the engine needs only ratings + age, nothing pre-computed by the sheet.
    ste_c = min(p["STE"], 80.0)
    SBpct = max(sum((ste_c - g("H8")) ** k * g0(c + "17") for k, c in zip(range(4), "BCDE")) + g("C39"), 0.0)
    # Saturation cap (INTENTIONAL divergence from the sheet): the fitted SB% line keeps
    # climbing through the top of the STE scale, but in the sim success plateaus —
    # measured from the calibration archives (career SB% by STE bucket, regulars):
    # TGS/OOTP26 75->.879 80->.880; BLM/OOTP27 75->.886 80->.888. Uncapped, an 80-STE
    # runner projects ~.94 and wSB overpays him ~3 runs (confirmed vs real career data:
    # Burchfield 486/81 = .857). League-specific per those empirics (audit B1):
    # TGS .880 / BLM .888. Validation vs the sheet will show expected diffs on
    # SB%-derived columns (wSB/UBR/BSR/WAA) for STE >= ~74; everything else stays exact.
    SBpct = min(SBpct, SB_CAP.get(league, 0.88))
    out["SB%"] = SBpct
    # audit B11: clamp rating INPUTS at 80 where live ratings exceed calibration support
    # (mirrors the STE cap above; BLM has 47 players with RUN>80 and STE-85 burners).
    run_c = min(p["RUN"], 80.0)
    def sbat(s):
        # audit B1: the SBA cubic (B15..E15) is fitted as DEVIATION from the archive
        # league attempt rate since the refit, so adding the live base C41 here counts
        # the league rate exactly ONCE (the old raw-level fit double-counted it —
        # +0.4..+2.0 wSB runs per fast player). STE input capped at 80 (support ends there).
        return max((sum((ste_c - g("H8")) ** k * g0(c + "15") for k, c in zip(range(4), "BCDE")) + g("C41"))
                   * (s["S1B"] + s["uBB"] + s["HBP"]), 0.0)
    def ubr(s):
        # audit B9 (INTENTIONAL divergence from the sheet, which subtracts C40): the UBR
        # cubic is fitted as a deviation from the archive GT rate, so the live league base
        # C40 must be ADDED back — (cubic + C40), verified live-frame: league-mean projected
        # UBR then matches the live C40·bases (the old sign gave every hitter ~+0.5..1.0
        # phantom runs/600 PA because C40 is negative).
        cubic = sum((run_c - g("H9")) ** k * g0(c + "19") for k, c in zip(range(4), "BCDE"))
        SBAT = sbat(s); SBn = SBpct * SBAT; CSn = SBAT - SBn
        wsb_pos = max(SBn * 0.2 + CSn * g("H35"), 0.0) - g("H36") * (s["S1B"] + s["uBB"] + s["HBP"])
        return (cubic + g("C40")) * ((s["uBB"] + s["S1B"] + s["HBP"]) * 3 + s["D2B"] * 2 + s["T3B"]
                                     - (CSn * 3 - SBn if wsb_pos > 0 else 0))
    def wsb(s):
        SBAT = sbat(s); SBn = SBpct * SBAT; CSn = SBAT - SBn
        return max(SBn * 0.2 + CSn * g("H35"), 0.0) - g("H36") * (s["S1B"] + s["uBB"] + s["HBP"])
    out["wSB vR"], out["wSB vL"] = wsb(sR), wsb(sL)
    out["UBR vR"], out["UBR vL"] = ubr(sR), ubr(sL)
    out["wSB wtd"] = wtd(out["wSB vR"], out["wSB vL"])
    out["UBR wtd"] = wtd(out["UBR vR"], out["UBR vL"])
    for suf in ("vR", "vL", "wtd"):
        out[f"BSR {suf}"] = out[f"UBR {suf}"] + out[f"wSB {suf}"]

    # ---- fielding RunsP per position ----
    IFR, IFE, IFA, TDP = p["IF RNG"], p["IF ERR"], p["IF ARM"], p["TDP"]
    OFR, OFE, OFA, HTS = p["OF RNG"], p["OF ERR"], p["OF ARM"], p["HT Sort"]
    CFRM, CARM = p["C FRM"], p["C ARM"]
    rp = {}
    # audit D3: when `fielding` is passed, the LINEAR (RNG-anchor)*slope+K term
    # is replaced by pm_pw() — the monotone piecewise archive curve, renormalized
    # so the live position population's ip-weighted mean is 0. All downstream
    # math (EAA coupling, DP, ARM, runs-per-play, WAA) is unchanged.
    # 1B
    if "1B" in fpos:
        pmaa = pm_pw("1B", IFR, HTS) * g("T5")
    else:
        pmaa = (((IFR - g("P9")) * g("L9")) + ((HTS - g("Q9")) * g("M9")) + g("K9")) * g("T5")
    eaa = (((IFE - g("P11")) * g("L11")) + g("K11")) * g("H33")
    rp["1B"] = (pmaa - eaa) * g("H38")
    # 2B
    if "2B" in fpos:
        pmaa = pm_pw("2B", IFR, IFA) * g("T8")
    else:
        pmaa = ((((IFR - g("P13")) * g("L13")) + ((IFA - g("Q13")) * g("M13"))) + g("K13")) * g("T8")
    eaa = (((IFE - g("P15")) * g("L15")) + g("K15")) * (pmaa + (g("T8") * g("T9")))
    dpaa = ((TDP - g("P17")) * g("L17") + g("K17")) * g("H33")
    rp["2B"] = (pmaa - eaa + dpaa) * g("H38")
    # 3B
    if "3B" in fpos:
        pmaa = pm_pw("3B", IFR, IFA) * g("T12")
    else:
        pmaa = (((IFR - g("P19")) * g("L19")) + ((IFA - g("Q19")) * g("M19")) + g("K19")) * g("T12")
    eaa = (((IFE - g("P21")) * g("L21")) + g("K21")) * (pmaa + (g("T12") * g("T13")))
    rp["3B"] = (pmaa - eaa) * g("H38")
    # SS
    if "SS" in fpos:
        pmaa = pm_pw("SS", IFR, IFA) * g("T15")
    else:
        pmaa = (((IFR - g("P23")) * g("L23")) + ((IFA - g("Q23")) * g("M23")) + g("K23")) * g("T15")
    eaa = (((IFE - g("P25")) * g("L25")) + g("K25")) * (pmaa + (g("T15") * g("T16")))
    dpaa = ((TDP - g("Q25")) * g("L45") + g("K45")) * g("H33")
    rp["SS"] = (pmaa - eaa + dpaa) * g("H38")
    # LF/CF/RF
    for pos, (pr, lr, kr, tr, per, ler, ker) in {
        "LF": ("P27", "L27", "K27", "T18", "P31", "L31", "K31"),
        "CF": ("P33", "L33", "K33", "T22", "P37", "L37", "K37"),
        "RF": ("P39", "L39", "K39", "T26", "P43", "L43", "K43"),
    }.items():
        ti = tr  # T col for this OF pos
        if pos in fpos:
            pmaa = pm_pw(pos, OFR, None) * g(ti)
        else:
            pmaa = (((OFR - g(pr)) * g(lr)) + g(kr)) * g(ti)
        # EAA uses pos-specific P/L/K (offset by 2 rows from PMAA row) and (PMAA + T*T)
        eaa_p = {"LF": ("P29", "L29", "K29", "T18", "T19"), "CF": ("P35", "L35", "K35", "T22", "T23"),
                 "RF": ("P41", "L41", "K41", "T26", "T27")}[pos]
        eaa = (((OFE - g(eaa_p[0])) * g(eaa_p[1])) + g(eaa_p[2])) * (pmaa + (g(eaa_p[3]) * g(eaa_p[4])))
        armaa = (((OFA - g(per)) * g(ler)) + g(ker)) * g(ti)
        rp[pos] = (pmaa - eaa) * g("H39") + armaa
    # Catcher
    c_frmaa = (g0("K3") + (CFRM - g("P3")) * g0("L3")) * g("H34")
    c_sba = ((CARM - g("P5")) * g("L5") + g("K5")) * g("H34") + g("T3")
    c_rto = max(0.0, (g0("K7") + (CARM - g("P5")) * g0("L7")) + g("T4"))
    c_cs = c_rto * c_sba
    c_armr = (c_cs * -(0.2 + g("H35"))) - ((g("T3") * g("T4")) * -(0.2 + g("H35")))
    rp["C"] = c_armr + c_frmaa  # C PMAA is blank -> 0
    for _pos, _val in rp.items():
        out[f"{_pos} RunsP"] = _val
    out["C FRMAA"], out["C SBA"], out["C ArmR"] = c_frmaa, c_sba, c_armr

    # ---- WAA per position ----
    H30, H32 = g("H30"), g("H32")
    posadj = {"C": "W2", "1B": "W3", "2B": "W4", "3B": "W5", "SS": "W6",
              "LF": "W7", "CF": "W8", "RF": "W9", "DH": "W10"}
    for suf in ("vR", "vL", "wtd"):
        BSR = out[f"BSR {suf}"]; BatR = out[f"BatR {suf}"]
        # catcher: batting via H32 scaling + framing offset. audit B8 (INTENTIONAL
        # divergence from the sheet, which adds full BSR): BSR is a 600-PA (H31)
        # quantity while catcher batting is scaled to H32 (500 PA) — scale BSR too.
        out[f"C WAA {suf}"] = (rp["C"] + BSR * (H32 / PA) + ((out[f"wOBA {suf}"] - g("H29")) / g("H20") * H32)
                               + (park["AB"] / PA * H32) + g("W2")) / H30
        for pos in ("1B", "2B", "3B", "SS", "LF", "CF", "RF"):
            out[f"{pos} WAA {suf}"] = (rp[pos] + BSR + BatR + g(posadj[pos])) / H30
        out[f"DH WAA {suf}"] = (BSR * 0.98 + out[f"DH BatR {suf}"] + g("W10")) / H30
    # Max WAA across eligible positions (eligibility from rating thresholds)
    elig = {
        "C": CFRM >= 45, "1B": HTS > 179 and IFR > 20,
        "2B": IFR >= 50 and p.get("T") == "R" and TDP >= 45,
        "3B": IFR >= 40 and IFA >= 50 and p.get("T") == "R",
        "SS": IFR >= 60 and IFA >= 50 and p.get("T") == "R",
        "LF": OFR >= 50, "CF": OFR >= 60, "RF": OFR >= 50, "DH": True,
    }
    out["_eligible_pos"] = {k for k, v in elig.items() if v}
    for pos in ("C", "1B", "2B", "3B", "SS", "LF", "CF", "RF"):
        out[f"{pos} Eligible"] = elig[pos]   # the app/org-builder slot hitters by these
    for suf in ("vR", "vL", "wtd"):
        vals = [out[f"{pos} WAA {suf}"] for pos in elig if elig[pos]]
        out[f"Max WAA {suf}"] = max(vals) if vals else None

    # ---- POTENTIAL (prospect ceiling): split-aware line from P-ratings ----
    # OOTP publishes potential WITHOUT splits, so a P rating is one number per skill.
    # It reads on the vR basis (measured over fully-developed hitters, where potential
    # must equal current: mean abs error vs vR 1.20 TGS / 2.37 BLM, vs the platoon-
    # weighted average 1.31 / 2.39, vs vL 2.13 / 3.41 — vR wins in both leagues).
    # The current line is built twice (vR and vL) and weighted, so it banks the
    # player's platoon advantage; a one-line potential banks none of it. For a maxed
    # player that difference IS the whole gap, which is how a ceiling landed under a
    # floor. So give the potential line the player's OWN measured platoon shape —
    # potential vL = P + (current vL - current vR) — and run both lines through the
    # same split_stats/woba/wsb/ubr path the current line uses, weighted by the same
    # platoon share. Nothing is fitted or tuned here; the gap is read off his ratings.
    EYEp, POWp, Kp, HTp, GAPp = (p.get(k) for k in ("EYE P", "POW P", "K P", "HT P", "GAP P"))
    out["MAX WAA P"] = None
    if None not in (EYEp, POWp, Kp, HTp, GAPp):
        def vl_of(pot, cur_vR, cur_vL):
            # Fallback to today's shapeless behaviour when a current split is missing.
            if cur_vR is None or cur_vL is None:
                return pot
            # Carrying the platoon gap can INVENT a rating the model was never fitted on:
            # TGS publishes no hitting rating above 80 at all (0 of 102,495 slots), yet the
            # derived vL reached 85 and 90, and those slots landed on 4 of the TGS and 9 of
            # the BLM draft top-25 — the most visible rankings resting on the least
            # supported arithmetic. Cap at the B11/B1 support end (80, same as the RUN and
            # STE input clamps above), but never below a rating this player actually
            # carries, so BLM's genuine 85s and 90s are not clipped.
            cap = max(80.0, cur_vR, cur_vL, pot)
            return min(pot + (cur_vL - cur_vR), cap)

        SPE = p["SPE"]
        pR = split_stats("vR", EYEp, POWp, Kp, HTp, GAPp, SPE)
        pL = split_stats("vL",
                         vl_of(EYEp, p["EYE vR"], p["EYE vL"]),
                         vl_of(POWp, p["POW vR"], p["POW vL"]),
                         vl_of(Kp, p["K vR"], p["K vL"]),
                         vl_of(HTp, p["BA vR"], p["BA vL"]),
                         vl_of(GAPp, p["GAP vR"], p["GAP vL"]),
                         SPE)
        out["wOBA P"] = wobaP = wtd(woba(pR), woba(pL))
        out["BatR P"] = BatRp = ((wobaP - g("H29")) / g("H20")) * PA
        dhwobaP = wtd(dh_woba(pR), dh_woba(pL))
        DHBatRp = ((dhwobaP - g("H29")) / g("H20")) * PA
        BSRp = wtd(ubr(pR) + wsb(pR), ubr(pL) + wsb(pL))
        # audit B8: BSR scaled to the catcher's H32 PA basis (see C WAA above)
        out["C WAA P"] = (rp["C"] + BSRp * (H32 / PA) + ((wobaP - g("H29")) / g("H20") * H32)
                          + (park["AB"] / PA * H32) + g("W2")) / H30
        for pos, w in (("1B", "W3"), ("2B", "W4"), ("3B", "W5"), ("SS", "W6"),
                       ("LF", "W7"), ("CF", "W8"), ("RF", "W9")):
            out[f"{pos} WAA P"] = (rp[pos] + BSRp + BatRp + g(w)) / H30
        out["DH WAA P"] = (BSRp * 0.98 + DHBatRp + g("W10")) / H30
        pvals = [out[f"{pos} WAA P"] for pos in ("C", "1B", "2B", "3B", "SS", "LF", "CF", "RF", "DH")
                 if pos in out["_eligible_pos"]]
        out["MAX WAA P"] = max(pvals) if pvals else None
    return out


# ---------- validation ----------
RATING_COLS = ["BA vR", "GAP vR", "POW vR", "EYE vR", "K vR", "BA vL", "GAP vL", "POW vL",
               "EYE vL", "K vL", "EYE P", "POW P", "K P", "HT P", "GAP P",
               "SPE", "SR", "STE", "RUN", "SB%", "IF RNG", "IF ERR", "IF ARM",
               "TDP", "OF RNG", "OF ERR", "OF ARM", "C ABI", "C FRM", "C ARM", "HT Sort",
               "B", "T", "Age", "Name", "POS"]
OPTIONAL = {"SB%", "EYE P", "POW P", "K P", "HT P", "GAP P"}  # may be blank for non-prospects

CHECK_COLS = ["SB%", "HBP vR", "uBB vR", "HR vR", "SO vR", "H-HR vR", "XBH-HR vR", "3B vR", "2B vR", "1B vR",
              "wOBA vR", "wOBA vL", "wOBA wtd", "OBP wtd", "BatR wtd", "wSB vR", "UBR vR", "BSR wtd",
              "C RunsP", "1B RunsP", "2B RunsP", "3B RunsP", "SS RunsP", "LF RunsP", "CF RunsP", "RF RunsP",
              "1B WAA wtd", "2B WAA wtd", "SS WAA wtd", "CF WAA wtd", "C WAA wtd", "DH WAA wtd", "Max WAA wtd",
              "wOBA P", "BatR P", "MAX WAA P"]
# Map a few CHECK names whose RunsP we expose under rp{} — handled below.


def main():
    league = sys.argv[1] if len(sys.argv) > 1 else "TGS"
    path = os.path.join(REPO, f"The Sheets {league}", "The Sheet Hitters.xlsx")
    dp, filt, park = scan_consts(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Hitters"]
    header = None
    diffs = {c: 0.0 for c in CHECK_COLS}
    counts = {c: 0 for c in CHECK_COLS}
    worst = {c: None for c in CHECK_COLS}
    n = 0
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        if header is None:
            header = [str(v).strip() if v is not None else "" for v in vals]
            idx = {h: i for i, h in enumerate(header)}
            continue
        rec = {h: vals[i] for h, i in idx.items() if i < len(vals)}
        if not rec.get("Name") or rec.get("Eligible") not in (True, 1, "TRUE"):
            continue
        # build input dict (numeric ratings)
        p = {}
        ok = True
        for col in RATING_COLS:
            v = rec.get(col)
            if col in ("B", "T", "Name", "POS"):
                p[col] = v
            else:
                nv = num(v)
                p[col] = nv
                if nv is None and col not in OPTIONAL:
                    ok = False
        if not ok:
            continue
        try:
            out = compute(p, dp, filt, park, league)
        except Exception as e:
            continue
        n += 1
        # also expose RunsP under "<POS> RunsP"
        for pos in ("C", "1B", "2B", "3B", "SS", "LF", "CF", "RF"):
            pass
        for col in CHECK_COLS:
            mine = out.get(col)
            if col.endswith("RunsP"):
                mine = out.get(col)  # not in out; compute path stored rp -> add below
            sheet = num(rec.get(col))
            if mine is None or sheet is None:
                continue
            d = abs(mine - sheet)
            diffs[col] = max(diffs[col], d)
            counts[col] += 1
            if worst[col] is None or d > worst[col][0]:
                worst[col] = (d, rec.get("Name"), sheet, mine)
    wb.close()
    print(f"Validated {n} eligible hitters in {league}")
    print(f"{'column':14} {'n':>5} {'maxAbsDiff':>14}   worst-case (name: sheet vs mine)")
    for col in CHECK_COLS:
        w = worst[col]
        wt = f"{w[1]}: {w[2]:.5f} vs {w[3]:.5f}" if w else "-"
        print(f"{col:14} {counts[col]:5} {diffs[col]:14.6e}   {wt}")


if __name__ == "__main__":
    main()
