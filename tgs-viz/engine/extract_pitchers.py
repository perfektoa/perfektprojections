"""
READ-ONLY extraction tool for reverse-engineering the PITCHERS tab of
"The Sheet" workbooks. Mirrors extract_sheet.py but targets the Pitchers sheet.

It NEVER writes to the workbooks. It loads them read-only and dumps:
  - every Pitchers column's header->row-2 formula (array formulas resolved to text)
  - the full Data Points constants (coordinate -> value) as JSON
  - Filters and a slice of Ballparks
into tgs-viz/engine/extracted/ so the math can be ported to Python.

Usage:  python tgs-viz/engine/extract_pitchers.py [LEAGUE]   (default TGS)
"""
import os, sys, json
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUTDIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "extracted")


def resolve(v):
    return v.text if isinstance(v, ArrayFormula) else v


def dump_formulas(path, sheet, out_txt):
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[sheet]
    hdr, frow = None, None
    for i, row in enumerate(ws.iter_rows()):
        vals = [resolve(c.value) for c in row]
        if i == 0:
            hdr = vals
        elif i == 1:
            frow = vals
            break
    wb.close()
    with open(out_txt, "w", encoding="utf-8") as f:
        for idx in range(len(hdr)):
            h = hdr[idx]
            fo = frow[idx] if frow and idx < len(frow) else None
            if h is None and fo is None:
                continue
            f.write(f"[{idx}] {h!r} => {fo!r}\n")
    return hdr, frow


def scan_values(path, sheet, min_row=None, max_row=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    d = {}
    kw = {}
    if min_row:
        kw["min_row"] = min_row
    if max_row:
        kw["max_row"] = max_row
    for row in ws.iter_rows(**kw):
        for c in row:
            if c.value is not None:
                d[c.coordinate] = c.value
    wb.close()
    return d


def main():
    league = sys.argv[1] if len(sys.argv) > 1 else "TGS"
    base = os.path.join(REPO, f"The Sheets {league}")
    ppath = os.path.join(base, "The Sheet Pitchers.xlsx")
    os.makedirs(OUTDIR, exist_ok=True)

    hdr, frow = dump_formulas(ppath, "Pitchers", os.path.join(OUTDIR, f"{league}_pitchers_formulas.txt"))
    dp = scan_values(ppath, "Data Points")
    with open(os.path.join(OUTDIR, f"{league}_pitchers_datapoints.json"), "w") as f:
        json.dump(dp, f, indent=0, default=str, sort_keys=True)
    filt = scan_values(ppath, "Filters")
    with open(os.path.join(OUTDIR, f"{league}_pitchers_filters.json"), "w") as f:
        json.dump(filt, f, indent=0, default=str, sort_keys=True)

    hmap = {h: i for i, h in enumerate(hdr) if h}
    print(f"=== {league} PITCHERS: {len(hmap)} named columns ===")
    for i, h in enumerate(hdr):
        if h:
            print(f"  [{i}] {h}")
    print(f"\nWrote: {OUTDIR}\\{league}_pitchers_formulas.txt, _datapoints.json, _filters.json")


if __name__ == "__main__":
    main()
