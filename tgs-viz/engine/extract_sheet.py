"""
READ-ONLY extraction tool for reverse-engineering "The Sheet" workbooks.

It NEVER writes to the workbooks. It loads them read-only and dumps:
  - every Hitters/Pitchers column's formula (array formulas resolved to text)
  - the full Data Points constants (coordinate -> value) as JSON
  - relevant Filters / Ballparks cells
into tgs-viz/engine/extracted/ so we can port the math to Python and lift the
league-specific constants.

Usage:  python tgs-viz/engine/extract_sheet.py [LEAGUE]   (default TGS)
"""
import os, sys, json
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUTDIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "extracted")


def resolve(v):
    return v.text if isinstance(v, ArrayFormula) else v


def dump_formulas(path, sheet, out_txt):
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[sheet]
    hdr, frow = None, None
    for i, row in enumerate(ws.iter_rows()):
        vals = [resolve(c.value) for c in row]
        if i == 0:
            hdr = vals
        elif i == 1:
            frow = vals
            break
    wb.close()
    with open(out_txt, "w", encoding="utf-8") as f:
        for idx in range(len(hdr)):
            h = hdr[idx]
            fo = frow[idx] if frow and idx < len(frow) else None
            if h is None and fo is None:
                continue
            f.write(f"[{idx}] {h!r} => {fo!r}\n")
    return hdr, frow


def scan_values(path, sheet, min_row=None, max_row=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    d = {}
    kw = {}
    if min_row:
        kw["min_row"] = min_row
    if max_row:
        kw["max_row"] = max_row
    for row in ws.iter_rows(**kw):
        for c in row:
            if c.value is not None:
                d[c.coordinate] = c.value
    wb.close()
    return d


def main():
    league = sys.argv[1] if len(sys.argv) > 1 else "TGS"
    base = os.path.join(REPO, f"The Sheets {league}")
    hpath = os.path.join(base, "The Sheet Hitters.xlsx")
    os.makedirs(OUTDIR, exist_ok=True)

    hdr, frow = dump_formulas(hpath, "Hitters", os.path.join(OUTDIR, f"{league}_hitters_formulas.txt"))
    dp = scan_values(hpath, "Data Points")
    with open(os.path.join(OUTDIR, f"{league}_hitters_datapoints.json"), "w") as f:
        json.dump(dp, f, indent=0, default=str, sort_keys=True)
    filt = scan_values(hpath, "Filters")
    bp = scan_values(hpath, "Ballparks", min_row=36, max_row=46)

    hmap = {h: i for i, h in enumerate(hdr) if h}

    print(f"=== {league} HITTERS: array formulas of interest ===")
    for col in ["H-HR vR", "wOBA wtd", "OBP wtd", "wSB wtd", "UBR vR", "BSR wtd",
                "Max WAA wtd", "MAX WAA P", "PA"]:
        i = hmap.get(col)
        if i is not None:
            print(f"\n--- {col} [{i}] ---\n{frow[i]!r}")

    print("\n\n=== REGRESSION CONSTANTS (Data Points) ===")
    # stat -> (rating-anchor cell, regression row in B/C/D/E, league-adj cell)
    specs = {
        "uBB(EYE)": ("H2", 3, "C33"),
        "HR(POW)":  ("H3", 5, "C34"),
        "SO(K)":    ("H4", 7, "C35"),
        "H-HR(BA)": ("H5", 9, "C36"),
        "XBH(GAP)": ("H6", 11, "C37"),
        "3B(SPE)":  ("H7", 13, "C38"),
    }
    for name, (anchor, r, adj) in specs.items():
        print(f"  {name:10} anchor {anchor}={dp.get(anchor)}  "
              f"above[B{r}={dp.get('B'+str(r))}, C{r}={dp.get('C'+str(r))}]  "
              f"below[D{r}={dp.get('D'+str(r))}, E{r}={dp.get('E'+str(r))}]  "
              f"{adj}={dp.get(adj)}")
    print("\n  HBP rate H37 =", dp.get("H37"))
    print("  wOBA weights H12:H17 =", [dp.get(f"H{r}") for r in range(12, 18)])
    print("  lgwOBA H29 =", dp.get("H29"), " wOBAscale H20 =", dp.get("H20"),
          " R/W H30 =", dp.get("H30"), " PA H31 =", dp.get("H31"), " C-PA H32 =", dp.get("H32"))
    print("  posAdj W2:W8 =", [dp.get(f"W{r}") for r in range(2, 9)])
    print("  fielding RunsP conv H38 =", dp.get("H38"), " H39 =", dp.get("H39"))
    print("  baserunning H35 =", dp.get("H35"), " H36 =", dp.get("H36"))

    print("\n=== FILTERS handedness/park multipliers (C9:D11) ===")
    for r in (9, 10, 11):
        print(f"  C{r}={filt.get('C'+str(r))}  D{r}={filt.get('D'+str(r))}")
    print("\n=== BALLPARKS row 40/43 (selected park factors) ===")
    for col in ["AA", "AB", "AG", "AI"]:
        print(f"  {col}40={bp.get(col+'40')}  {col}43={bp.get(col+'43')}")
    print(f"\nWrote: {OUTDIR}\\{league}_hitters_formulas.txt and _datapoints.json")


if __name__ == "__main__":
    main()
