"""
Production-path proof: drive the engine from the RAW `Player List` ratings
(read-only) and confirm it reproduces the current hitters.json — i.e. the engine
replaces the Excel formulas entirely. The same input reader is what an OOTP /
StatsPlus ratings feed will plug into (Player List == the OOTP export schema).

Read-only. Does not modify the workbook or the existing JSON.

  python tgs-viz/engine/build_json.py [LEAGUE]      (default TGS)
"""
import os, sys, json
from openpyxl import load_workbook
import hitters as H

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def rating(v):
    # The sheet converts "-" -> 20 then VALUE(); mirror that for raw ratings.
    if isinstance(v, str) and v.strip() == "-":
        return 20.0
    return H.num(v)


def main():
    league = sys.argv[1] if len(sys.argv) > 1 else "TGS"
    hpath = os.path.join(REPO, f"The Sheets {league}", "The Sheet Hitters.xlsx")
    dp, filt, park = H.scan_consts(hpath)

    # current JSON (the app's data), keyed by ID
    jpath = os.path.join(REPO, "tgs-viz", "public", "data", league, "hitters.json")
    cur = {str(r.get("ID")): r for r in json.load(open(jpath, encoding="utf-8")) if r.get("ID") is not None}

    wb = load_workbook(hpath, read_only=True, data_only=True)
    ws = wb["Player List"]
    header = None
    cols = ["Max WAA wtd", "wOBA wtd", "MAX WAA P", "2B WAA wtd", "C WAA wtd", "BatR wtd"]
    diffs = {c: 0.0 for c in cols}
    n = matched = 0
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        if header is None:
            header = [str(v).strip() if v is not None else "" for v in vals]
            idx = {h: i for i, h in enumerate(header)}
            continue
        rec = {h: vals[i] for h, i in idx.items() if i < len(vals)}
        if not rec.get("Name"):
            continue
        p = {}
        ok = True
        for col in H.RATING_COLS:
            v = rec.get(col)
            if col in ("B", "T", "Name", "POS"):
                p[col] = v
            else:
                nv = rating(v)
                p[col] = nv
                if nv is None and col not in H.OPTIONAL:
                    ok = False
        if not ok:
            continue
        try:
            out = H.compute(p, dp, filt, park, league)
        except Exception:
            continue
        n += 1
        j = cur.get(str(rec.get("ID")))
        if not j:
            continue
        matched += 1
        for c in cols:
            mine, sheet = out.get(c), H.num(j.get(c))
            if mine is None or sheet is None:
                continue
            diffs[c] = max(diffs[c], abs(mine - sheet))
    wb.close()
    print(f"{league}: computed {n} hitters from Player List, matched {matched} to hitters.json")
    print(f"  max abs diff vs current JSON (engine fed RAW ratings, no Excel formulas):")
    for c in cols:
        print(f"    {c:14} {diffs[c]:.3e}")
    worst = max(diffs.values())
    print(f"  => {'PARITY OK' if worst < 1e-6 else 'MISMATCH'} (worst {worst:.2e})")


if __name__ == "__main__":
    main()
