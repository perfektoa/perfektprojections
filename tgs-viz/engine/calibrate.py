"""
calibrate.py — compute every "Data Points" regression constant in Python,
replacing the 25 Regressions.xlsx workbook (its pivots + LINEST/weighted-OLS sheets).

The workbook pooled OOTP clone-sim dumps (Batting/Fielding/Pitching tables) with the
baseline ratings tables (Hitters/Pitchers), aggregated per player via Power Pivot, and
fit rating->stat regressions. This module reproduces that math EXACTLY — including the
workbook's as-built quirks (documented inline) — so it can be validated cell-for-cell
against the workbook before replacing it.

Data sources (CSV):
  Batting.csv / Pitching.csv / Fielding.csv : OOTP dump rows (many seasons per player)
  Hitters.csv / Pitchers.csv                : baseline ratings tables (one row per player)

Modes:
  python calibrate.py --csv-dir DIR --validate expected_dp.json   # reproduce workbook
  python calibrate.py --csv-dir DIR --json out.json               # emit constants

Regression model (hitting/pitching, per block):
  points  = players passing the block's rating filter
  x       = player rating − the centre the CONSUMER subtracts: the sheet's Data Points
            anchor for the pitching rate blocks (audit B12), else the mean rating over
            the pivot's visible players ("GT delta")
  y       = player stat-rate − rate over the pivot's visible-player sums    ("GT delta")
  weight  = player PA (hitting) / BF (pitching), always from the base pivot
  fit     = weighted least squares:  slope = (Σw·Σwxy − Σwx·Σwy)/(Σw·Σwx² − (Σwx)²)
                                     intercept = (Σwy − slope·Σwx)/Σw
Fielding: unweighted LINEST (1 or 2 predictors) per position block.

"Visible players" (visual totals) per pivot:
  hitting pivot1 : split_id=3 rows, players with Σpa ≥ 10
  hitting pivot2 : split_id=1 rows, top 288 players by Σpa
  pitching SP    : split_id=1 rows, players with Σgs ≥ 250
  pitching RP    : split_id=1 AND gs=0 rows, players with Σg ≥ 100
  fielding       : per position, top 32 (IF) / 30 (OF, C) players by fielding-row count
"""
import csv, json, math, os, sys, argparse
from collections import defaultdict

# ---------------------------------------------------------------- tiny linear algebra


def wls(pts):
    """Weighted least squares over (x, y, w) tuples — the workbook's normal equations."""
    sw = swx = swy = swxy = swx2 = 0.0
    for x, y, w in pts:
        sw += w; swx = swx + x * w; swy += y * w; swxy += x * y * w; swx2 += x * x * w
    slope = (sw * swxy - swx * swy) / (sw * swx2 - swx ** 2)
    intercept = (swy - slope * swx) / sw
    return intercept, slope


def linest(ys, xcols):
    """Unweighted OLS of y on one or two predictor columns + constant (Excel LINEST).
    Returns (intercept, m1, m2?) with m1 = slope of xcols[0]."""
    n = len(ys)
    k = len(xcols)
    # normal equations over design [x1, (x2,) 1]
    cols = list(xcols) + [[1.0] * n]
    m = len(cols)
    A = [[sum(cols[i][r] * cols[j][r] for r in range(n)) for j in range(m)] for i in range(m)]
    b = [sum(cols[i][r] * ys[r] for r in range(n)) for i in range(m)]
    # gaussian elimination with partial pivoting
    for col in range(m):
        piv = max(range(col, m), key=lambda r: abs(A[r][col]))
        A[col], A[piv] = A[piv], A[col]
        b[col], b[piv] = b[piv], b[col]
        for r in range(col + 1, m):
            f = A[r][col] / A[col][col]
            for c in range(col, m):
                A[r][c] -= f * A[col][c]
            b[r] -= f * b[col]
    sol = [0.0] * m
    for r in range(m - 1, -1, -1):
        sol[r] = (b[r] - sum(A[r][c] * sol[c] for c in range(r + 1, m))) / A[r][r]
    if k == 1:
        return sol[1], sol[0]              # intercept, m1
    return sol[2], sol[0], sol[1]          # intercept, m1, m2

# ---------------------------------------------------------------- data loading


def load_rows(path):
    """CSV -> list of dicts with numeric coercion; rows lacking a numeric player_id/ID dropped."""
    out = []
    with open(path, newline="", encoding="utf-8") as f:
        rd = csv.reader(f)
        hdr = next(rd)
        idcol = 0
        for row in rd:
            if not row or row[idcol] in ("", None):
                continue
            d = {}
            for h, v in zip(hdr, row):
                if v == "":
                    d[h] = None
                    continue
                try:
                    d[h] = float(v)
                except ValueError:
                    d[h] = v
            out.append(d)
    return out


def group_sum(rows, keyfield, fields, where=None):
    """SUM(fields) grouped by keyfield over rows passing `where` (the pivot aggregation)."""
    agg = defaultdict(lambda: defaultdict(float))
    cnt = defaultdict(int)
    for r in rows:
        if where and not where(r):
            continue
        k = r.get(keyfield)
        if k is None:
            continue
        a = agg[k]
        for f in fields:
            v = r.get(f)
            if isinstance(v, (int, float)):
                a[f] += v
        cnt[k] += 1
    return agg, cnt


def totals(agg, ids, fields):
    """Visual-totals Grand Total: sums over the visible player set only."""
    t = {f: 0.0 for f in fields}
    for i in ids:
        a = agg[i]
        for f in fields:
            t[f] += a.get(f, 0.0)
    return t


def rating_mean(ratings, ids, col):
    """GT of 'Average of <rating>' = unweighted mean over the visible players (verified)."""
    vals = [ratings[i][col] for i in ids if i in ratings and isinstance(ratings[i].get(col), (int, float))]
    return sum(vals) / len(vals)

# ---------------------------------------------------------------- hitting

HIT_FIELDS = ["pa", "ab", "ibb", "hp", "bb", "hr", "k", "h", "d", "t", "sb", "cs", "sf", "ubr"]


def _div(n, d, iferror=None):
    if d == 0:
        if iferror is not None:
            return iferror
        raise ZeroDivisionError("unexpected 0 denominator in a non-IFERROR block")
    return n / d


def hitting(bat_rows, hitters):
    # pivot1: split_id=3, players with sum(pa) >= 10
    p1, _ = group_sum(bat_rows, "player_id", HIT_FIELDS, where=lambda r: r.get("split_id") == 3)
    vis1 = sorted(i for i, a in p1.items() if a["pa"] >= 10)
    gt1 = totals(p1, vis1, HIT_FIELDS)
    # pivot2: split_id=1, top 288 players by sum(pa)
    p2, _ = group_sum(bat_rows, "player_id", HIT_FIELDS, where=lambda r: r.get("split_id") == 1)
    vis2 = sorted(p2, key=lambda i: (-p2[i]["pa"], i))[:288]
    gt2 = totals(p2, vis2, HIT_FIELDS)

    # y-rate definitions over an aggregate dict (used for both player and GT rows)
    def ubb(a):   return _div(a["bb"] - a["ibb"], a["pa"] - a["ibb"] - a["hp"])
    def hrp(a):   return _div(a["hr"], a["pa"] - a["ibb"] - a["bb"] - a["hp"])
    def kp(a):    return _div(a["k"], a["pa"] - a["ibb"] - a["bb"] - a["hp"])
    def babip(a): return _div(a["h"] - a["hr"], a["ab"] - a["hr"] - a["k"] + a["sf"])
    def xbh(a):   return _div(a["d"] + a["t"], a["h"] - a["hr"])
    def trip(a):  return _div(a["t"], a["d"] + a["t"], iferror=None)  # IFERROR handled per-block
    def sbpct(a): return _div(a["sb"], a["sb"] + a["cs"], iferror=None)
    def sba(a):   return _div(a["sb"] + a["cs"], a["h"] - a["hr"] - a["t"] - a["d"] + a["bb"] + a["hp"], iferror=None)
    def ubr_rate(a):
        base = ((a["h"] - a["d"] - a["t"] - a["hr"]) + a["bb"] + a["hp"]) * 3 \
               + a["d"] * 2 + a["t"] * 2 - a["sb"] - a["cs"] * 3
        return _div(a["ubr"], base, iferror=None)

    rat = lambda col: {i: hitters[i][col] for i in hitters if isinstance(hitters[i].get(col), (int, float))}

    def block(vis, agg, gt, xcol, yfn, filt, weight_from_p1=False, iferror0=False, subtract_gt=True):
        xr = rat(xcol)
        xbase = rating_mean(hitters, vis, xcol)
        ybase = yfn(gt) if subtract_gt else 0.0
        pts = []
        for i in vis:
            if i not in xr or not filt(xr[i]):
                continue
            if weight_from_p1:
                w = p1[i]["pa"] if (i in p1 and i in set(vis1_set)) else None
                if w is None:
                    w = 0.0  # IFERROR(...,0) on the weight (UBR block)
            else:
                w = agg[i]["pa"]
            try:
                y = yfn(agg[i]) - ybase
            except ZeroDivisionError:
                if not iferror0:
                    raise
                y = 0.0
            pts.append((xr[i] - xbase, y, w))
        return wls(pts)

    vis1_set = set(vis1)
    out = {}
    # ---- pivot1 blocks (weight = own pa). lo-filters reproduce the workbook verbatim
    #      except audit fix B10: the workbook's lGAP block filtered on SPE<50 and its
    #      lSPE block on GAP<50 (swapped columns). Here each lo block filters on its
    #      OWN rating, matching every other lo block. (lEye <=45 vs <50 is kept — all
    #      ratings are multiples of 5, so it is a verified no-op.)
    spe = rat("SPE"); gap = rat("GAP vR")
    out["hEYE"]   = block(vis1, p1, gt1, "EYE vR", ubb,   lambda x: x >= 50)
    out["lEye"]   = block(vis1, p1, gt1, "EYE vR", ubb,   lambda x: x <= 45)
    out["hPOW"]   = block(vis1, p1, gt1, "POW vR", hrp,   lambda x: x >= 50)
    out["lPOW"]   = block(vis1, p1, gt1, "POW vR", hrp,   lambda x: x < 50)
    out["hK's"]   = block(vis1, p1, gt1, "K vR",   kp,    lambda x: x >= 50)
    out["lK's"]   = block(vis1, p1, gt1, "K vR",   kp,    lambda x: x < 50)
    out["hBABIP"] = block(vis1, p1, gt1, "BA vR",  babip, lambda x: x >= 50)
    out["lBABIP"] = block(vis1, p1, gt1, "BA vR",  babip, lambda x: x < 50)
    out["hGAP"]   = block(vis1, p1, gt1, "GAP vR", xbh,   lambda x: x >= 50)
    out["lGAP"]   = block(vis1, p1, gt1, "GAP vR", xbh,   lambda x: x < 50)   # B10: was filtered on SPE
    out["hSPE"]   = block(vis1, p1, gt1, "SPE", trip, lambda x: x >= 50, iferror0=True)
    out["lSPE"]   = block(vis1, p1, gt1, "SPE", trip, lambda x: x < 50,  iferror0=True)  # B10: was filtered on GAP

    # ---- pivot2 blocks (split_id=1 stats; weight = pivot1 pa)
    def block_p2(xcol, yfn, fcol, filt, subtract_gt=True, weight_iferror0=False):
        xr, fr = rat(xcol), rat(fcol)
        xbase = rating_mean(hitters, vis2, xcol)
        ybase = yfn(gt2) if subtract_gt else 0.0
        pts = []
        for i in vis2:
            if i not in xr or i not in fr or not filt(fr[i]):
                continue
            if i in vis1_set:
                w = p1[i]["pa"]
            elif weight_iferror0:
                w = 0.0
            else:
                raise KeyError(f"player {i} in pivot2 block but not in pivot1 (workbook would #N/A)")
            try:
                y = yfn(p2[i]) - ybase
            except ZeroDivisionError:
                y = 0.0  # all pivot2 y-INDEX cells are IFERROR(...,0)
            pts.append((xr[i] - xbase, y, w))
        return wls(pts)

    # audit B1: the SBA block is fitted as DEVIATION from the pivot's grand-total rate
    # (subtract_gt=True) like every other block. The old raw-level fit (subtract_gt=False)
    # baked the archive league attempt rate into the intercept, and the sheet/engine then
    # added the live base C41 on top — double-counting the league rate (+0.4..+2.0 wSB
    # runs per fast player). Slope is unchanged; only the intercept drops by the GT rate.
    out["hSBA"] = block_p2("STE", sba,      "STE", lambda f: f >= 55, subtract_gt=True)
    # audit B12: the SB% block selected its pool on SPE while regressing on STE — the
    # same swapped-column copy-drag as B10, and the odd one out here: hSBA and UBR both
    # filter on their own predictor. Stolen-base SUCCESS is a function of stealing
    # ability, not raw speed (user directive), so the pool is now the players the fit is
    # actually about. A fast runner who cannot steal no longer sets the success curve.
    out["SB%"]  = block_p2("STE", sbpct,    "STE", lambda f: f >= 55)
    out["UBR"]  = block_p2("RUN", ubr_rate, "RUN", lambda f: True, weight_iferror0=True)
    return out

# ---------------------------------------------------------------- pitching

PIT_FIELDS = ["bf", "g", "gs", "sb", "cs", "hp", "bb", "k", "hra", "ha", "sa", "da", "ta", "iw", "ab", "sf"]

# audit B12 — WHERE THE FITTED INTERCEPT IS BASED.
# The Sheet Pitchers evaluates every one of these lines as
#       intercept + slope * (rating - Data Points anchor) + league-rate K
# so the intercept has to be the line's value AT THAT ANCHOR. The anchors are
# the league's own BF-weighted mean ratings, kept in 25 Metadata.xlsx and pushed
# to Data Points H2..H10 by ingest/sync_datapoints.py ('panchor:'). Centring the
# fit on the pivot's unweighted mean rating instead left the intercept expressed
# about a DIFFERENT population (1.3-7.4 rating points away); because the hi and
# lo branches have different slopes the common shift moved them by different
# amounts, tearing the two segments apart at rating 50 (wrong-direction rungs)
# and pushing the projected league rate off its own archive Grand Total by up to
# 2.8pp of K%. WLS is invariant to a shift of x, so centring here leaves every
# fitted line bit-identical and changes only the intercept it is stated with.
# NOT covered: the HLD (SB%/SBA) block, whose consumer anchor is the separate
# Data Points I2/I7 "Hold" cell and whose consumer is a cubic SUMPRODUCT.
PITCH_ANCHOR_SECTIONS = {"Starter Ratings": "SP", "Reliever Ratings": "RP"}
PITCH_ANCHOR_COLS = {"STU": "STU vR", "HRA": "HRR vR",
                     "pBABIP": "PBABIP vR", "CON": "CON vR"}
LEAGUE_DIRS = {"TGS": "The Sheets TGS", "BLM": "The Sheets BLM"}
REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def pitch_anchors(league):
    """{'SP'/'RP': {rating column: anchor rating}} read READ-ONLY out of that
    league's own 25 Metadata.xlsx 'Data Points' (labels col S, values col T —
    the same layout ingest/sync_datapoints.py pushes into the Pitchers sheet).
    Raises rather than guessing: a fit centred on anything but the consumer's
    anchor is the defect this guards against."""
    from openpyxl import load_workbook
    path = os.path.join(REPO, LEAGUE_DIRS[league], "25 Metadata.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data Points"]
    out, role = {"SP": {}, "RP": {}}, None
    for lab, val in ws.iter_rows(min_col=19, max_col=20, values_only=True):   # S, T
        lab = lab.strip() if isinstance(lab, str) else lab
        if lab in PITCH_ANCHOR_SECTIONS:
            role = PITCH_ANCHOR_SECTIONS[lab]
        elif role and lab in PITCH_ANCHOR_COLS and isinstance(val, (int, float)):
            out[role][PITCH_ANCHOR_COLS[lab]] = float(val)
    wb.close()
    for r, cols in out.items():
        missing = set(PITCH_ANCHOR_COLS.values()) - set(cols)
        if missing:
            raise KeyError(f"{league} 25 Metadata.xlsx 'Data Points' is missing the "
                           f"{r} anchor(s) for {sorted(missing)} (labels col S / values col T)")
    return out


def pitching(pit_rows, pitchers, role, anchors):
    if role == "SP":
        agg, _ = group_sum(pit_rows, "player_id", PIT_FIELDS, where=lambda r: r.get("split_id") == 1)
        vis = sorted(i for i, a in agg.items() if a["gs"] >= 250)
    else:  # RP: relief rows only (gs=0), players with sum(g) >= 100
        agg, _ = group_sum(pit_rows, "player_id", PIT_FIELDS,
                           where=lambda r: r.get("split_id") == 1 and r.get("gs") == 0)
        vis = sorted(i for i, a in agg.items() if a["g"] >= 100)
    gt = totals(agg, vis, PIT_FIELDS)

    def ubb(a):   return _div(a["bb"] - a["iw"], a["bf"] - a["iw"] - a["hp"])
    def hrp(a):   return _div(a["hra"], a["bf"] - a["bb"] - a["hp"])
    def kp(a):    return _div(a["k"], a["bf"] - a["hp"] - a["bb"])
    def babip(a): return _div(a["ha"] - a["hra"], a["ab"] - a["hra"] - a["k"] + a["sf"])
    def sba(a):   return _div(a["sb"] + a["cs"], a["bb"] + a["hp"] + a["sa"])
    def sbp(a):   return _div(a["sb"], a["sb"] + a["cs"])

    def block_pts(xcol, yfn, filt):
        xr = {i: pitchers[i][xcol] for i in pitchers if isinstance(pitchers[i].get(xcol), (int, float))}
        # audit B12: centre on the anchor the sheet subtracts, not the pivot mean
        xbase = anchors[xcol] if xcol in anchors else rating_mean(pitchers, vis, xcol)
        ybase = yfn(gt)
        return [(xr[i] - xbase, yfn(agg[i]) - ybase, agg[i]["bf"])
                for i in vis if i in xr and filt(xr[i])]

    def block(xcol, yfn, filt):
        return wls(block_pts(xcol, yfn, filt))

    out = {}
    out["hCON"]   = block("CON vR",    ubb,   lambda x: x >= 50)
    out["lCON"]   = block("CON vR",    ubb,   lambda x: x < 50)
    out["hHRR"]   = block("HRR vR",    hrp,   lambda x: x >= 50)
    out["lHRR"]   = block("HRR vR",    hrp,   lambda x: x < 50)
    out["hSTU"]   = block("STU vR",    kp,    lambda x: x >= 50)
    out["lSTU"]   = block("STU vR",    kp,    lambda x: x < 50)
    out["hBABIP"] = block("PBABIP vR", babip, lambda x: x >= 20)   # >=20 = whole sample
    # audit B7: the workbook's lo-BABIP rows were byte-copies of the hi fit (its ">=20"
    # filter spans the whole sample). Emit the REAL lo-block points (x<50); compute_all
    # pools them across SP+RP before fitting — the SP lo pool alone is only ~10 players.
    out["_lBABIP_pts"] = block_pts("PBABIP vR", babip, lambda x: x < 50)
    if role == "SP":  # SB%/SBA regressions are consumed from the SP sheet for both roles
        out["SB%"] = block("HLD", sbp, lambda x: True)
        out["SBA"] = block("HLD", sba, lambda x: True)
    return out

# ---------------------------------------------------------------- fielding

FLD_STATS = ["ip", "dp", "sba", "rto", "e", "zr", "framing", "arm", "pa", "pm", "er"]
POS = {"P": 1, "C": 2, "1B": 3, "2B": 4, "3B": 5, "SS": 6, "LF": 7, "CF": 8, "RF": 9}


def fielding(fld_rows, pit_rows, hitters):
    # --- per-position player pivots: top-N players by fielding-row count at that position
    def pos_pivot(pos_ids, topn):
        if isinstance(pos_ids, int):
            pos_ids = [pos_ids]
        agg, cnt = group_sum(fld_rows, "player_id", FLD_STATS + ["team_id"],
                             where=lambda r: r.get("position") in pos_ids)
        # pivot row filter: top-N players by Sum of ip at this position (iMeasureHier 388)
        vis = sorted(agg, key=lambda i: (-agg[i]["ip"], i))[:topn]
        gt = totals(agg, vis, FLD_STATS)
        return agg, vis, gt

    # --- double-play blocks (audit D8, REPLACED WHOLE — never fix the 25-row
    # quirk in isolation): the workbook's attribution chain (pitcher GIDP minus
    # 3B/OF DPs, split by a PM-share heuristic, then per-player LINESTs of the
    # pivot teams' attributed rate on TDP — the SS one spanning only 25 rows as
    # built) undervalued DP skill 30-50%, with the 25-row quirk accidentally
    # compensating part of the SS error. Replacement: ONE bivariate team
    # regression, each position's slope controlling for the other's.
    #   cell = (team_id, year)  [clone-pooled, same convention as currency_fit]
    #   y    = split-1 team GIDP per DOLLARDE-clean inning pitched
    #   x1,x2= the cell's innings-weighted mean TDP of its 2B / SS fielders
    #          (players without a TDP rating excluded from weight and sum)
    #   fit  = OLS with intercept on centered x and y -> intercept ~ 0, so the
    #          engine's ((TDP - live anchor)*slope + K)*H33 stays a deviation
    #          and the position-average TDP contribution stays ~ 0.
    #
    # audit D8b — DP OPPORTUNITY EXPOSURE (the fix for the wrong-signed BLM slopes).
    # y is GIDP per INNING, but a GIDP needs a runner on first AND a ground ball, and
    # neither is constant across cells. A team's mean middle-infield TDP is strongly
    # NEGATIVELY correlated with the baserunners its pitchers allow (defensive ratings
    # co-move, so high-TDP infields are also high-range infields that suppress hits).
    # Measured in BLM's own archive: corr(TDP_2B, baserunners/IP) = -0.39 while
    # corr(baserunners/IP, dp/IP) = +0.50. Regressing dp/IP on TDP alone therefore
    # carries a negative omitted-variable bias through the opportunity channel — in BLM
    # big enough to invert the sign (a better TDP rating projecting FEWER double plays).
    # The bias is a property of the fit, not of a league, so the controls are always on
    # and each league's size is whatever its own archive says it is.
    # Controlling for the cell's own DP opportunity
    #   BR/IP = (hits allowed + BB + HBP - HR)/IP     GB/IP = ground balls/IP
    # removes it. Both controls are properties of the opposing offence / pitching staff,
    # so the reported TDP coefficients keep their dp-per-inning units and the engine's
    # application form is untouched; controlling for BR/IP additionally stops the DP
    # block double-counting infield RANGE, which already has its own PM% block.
    def _dollarde(ip):
        return int(ip) + (ip - int(ip)) * 10.0 / 3.0

    def _ols_se(ys, xcols):
        """OLS with intercept over k predictors; returns (intercept, coefs, ses)."""
        n, k = len(ys), len(xcols)
        cols = list(xcols) + [[1.0] * n]
        m = k + 1
        A = [[sum(cols[i][r] * cols[j][r] for r in range(n)) for j in range(m)] for i in range(m)]
        rhs = [sum(cols[i][r] * ys[r] for r in range(n)) for i in range(m)]
        M = [row[:] + [1.0 if i == j else 0.0 for j in range(m)] for i, row in enumerate(A)]
        for c in range(m):
            p = max(range(c, m), key=lambda r: abs(M[r][c]))
            M[c], M[p] = M[p], M[c]
            pv = M[c][c]
            M[c] = [v / pv for v in M[c]]
            for r in range(m):
                if r != c and M[r][c]:
                    f = M[r][c]
                    M[r] = [M[r][i] - f * M[c][i] for i in range(2 * m)]
        inv = [row[m:] for row in M]
        sol = [sum(inv[i][j] * rhs[j] for j in range(m)) for i in range(m)]
        res = [ys[r] - sum(cols[i][r] * sol[i] for i in range(m)) for r in range(n)]
        s2 = sum(v * v for v in res) / (n - m)
        ses = [(s2 * inv[i][i]) ** 0.5 for i in range(m)]
        return sol[k], sol[:k], ses[:k]

    def team_year_tdp(code):
        sip, sipt = defaultdict(float), defaultdict(float)
        for r in fld_rows:
            if r.get("position") != code:
                continue
            i, ip = r.get("player_id"), r.get("ip")
            tdp = hitters.get(i, {}).get("TDP")
            if not isinstance(tdp, (int, float)) or not isinstance(ip, (int, float)) or ip <= 0:
                continue
            key = (r.get("team_id"), r.get("year"))
            sip[key] += ip
            sipt[key] += ip * tdp
        return {k: sipt[k] / sip[k] for k in sip if sip[k] > 0}

    x2b_t, xss_t = team_year_tdp(4), team_year_tdp(6)
    dp_t, ip_t = defaultdict(float), defaultdict(float)
    opp_t = defaultdict(lambda: defaultdict(float))
    for r in pit_rows:
        if r.get("split_id") != 1:
            continue
        key = (r.get("team_id"), r.get("year"))
        dp_t[key] += r.get("dp") or 0.0
        ip_t[key] += _dollarde(r.get("ip") or 0.0)
        for f in ("ha", "bb", "hp", "hra", "gb"):
            v = r.get(f)
            if isinstance(v, (int, float)):
                opp_t[key][f] += v
    cells = sorted(k for k in dp_t if k in x2b_t and k in xss_t and ip_t[k] > 0)
    yv = [dp_t[k] / ip_t[k] for k in cells]
    ybar = sum(yv) / len(yv)
    # DP-opportunity exposure per inning (audit D8b) — same cells, same denominator as y
    br_v = [(opp_t[k]["ha"] + opp_t[k]["bb"] + opp_t[k]["hp"] - opp_t[k]["hra"]) / ip_t[k] for k in cells]
    gb_v = [opp_t[k]["gb"] / ip_t[k] for k in cells]
    xcols_raw = [[x2b_t[k] for k in cells], [xss_t[k] for k in cells], br_v, gb_v]
    xnames = ["TDP_2B", "TDP_SS", "BR/IP", "GB/IP"]
    if not any(br_v) or not any(gb_v):
        # opportunity columns unavailable in this dump vintage — fall back to the bare
        # bivariate fit rather than silently regressing on a column of zeros
        xcols_raw, xnames = xcols_raw[:2], xnames[:2]
        print("  D8b WARNING: no ha/bb/gb columns in the pitching dump — DP fit runs "
              "WITHOUT the opportunity controls and its slopes may be sign-unstable")
    xcols = [[v - sum(c) / len(c) for v in c] for c in xcols_raw]  # centred, as above
    dp_b, coefs, ses = _ols_se([y - ybar for y in yv], xcols)
    dp_m2b, dp_mss = coefs[0], coefs[1]
    dp_prov = ("bivariate-team-regression+opportunity-controls" if len(xnames) == 4
               else "bivariate-team-regression")
    print(f"  D8 DP bivariate team regression [{dp_prov}]: n={len(cells)} team-year cells; "
          f"intercept {dp_b:.2e}; league {ybar:.5f} dp/inning")
    for nm, c, se in zip(xnames, coefs, ses):
        flag = ""
        if nm.startswith("TDP"):
            flag = "   <-- WRONG SIGN" if c < 0 else ("   (CI covers 0)" if abs(c) < 1.96 * se else "")
        print(f"      {nm:8s} {c:+.6g}  se {se:.4g}  t {c / se:+.2f}{flag}")

    R = lambda i, col: hitters[i][col]

    def dev_series(vis, agg, gt, ycols, xspec):
        """Build y list + x column lists for a LINEST block."""
        ys, xs = [], [[] for _ in xspec]
        for i in vis:
            ys.append(ycols(i))
            for j, (col, base) in enumerate(xspec):
                xs[j].append(R(i, col) - base)
        return ys, xs

    out = {}

    # ---------- infield blocks
    # E% fits (audit B2+B5): the workbook baselined 2B/SS E% on the GT's zr/ip instead
    # of e/ip (B2 — the sole cause of the −4.63/−2.33 run live means at SS/2B), and it
    # fitted every IF E% per-INNING while the engine applies the 2B/3B/SS slope per
    # PLAY-MADE (B5 — error skill compressed ~3x). Fix: 2B/3B/SS are fitted as
    # y = e/pm (errors per play made, mirroring the OF E% blocks) against their OWN
    # pivot's GT and rating means (SS no longer borrows the 2B pivot's x-baseline).
    # 1B keeps the per-inning fit — its engine formula multiplies by H33 innings.
    for pos, topn, x2col, e_den in [
        ("1B", 32, "HT CM",  "ip"),
        ("2B", 32, "IF ARM", "pm"),
        ("3B", 32, "IF ARM", "pm"),
        ("SS", 32, "IF ARM", "pm"),
    ]:
        agg, vis, gt = pos_pivot(POS[pos], topn)
        rngb = rating_mean(hitters, vis, "IF RNG")
        x2b  = rating_mean(hitters, vis, x2col)
        # PM%: y = pm/pa − GT, x1 = IF RNG dev, x2 = HT CM (1B) / IF ARM dev
        pmgt = gt["pm"] / gt["pa"]
        ys, xs = dev_series(vis, agg, gt,
                            lambda i: agg[i]["pm"] / agg[i]["pa"] - pmgt,
                            [("IF RNG", rngb), (x2col, x2b)])
        b, m1, m2 = linest(ys, xs)
        out[f"{pos} PM%"] = {"intercept": b, "x": m1, "x2": m2}
        # E%: y = e/ip (1B) or e/pm (2B/3B/SS) − same-ratio GT ; x = IF ERR dev (own pivot)
        egt = gt["e"] / gt[e_den]
        errb = rating_mean(hitters, vis, "IF ERR")
        ys, xs = dev_series(vis, agg, gt,
                            lambda i: agg[i]["e"] / agg[i][e_den] - egt,
                            [("IF ERR", errb)])
        b, m1 = linest(ys, xs)
        out[f"{pos} E%"] = {"intercept": b, "x": m1}
        # DP blocks (2B and SS only) — audit D8: slopes from the ONE bivariate
        # team regression above; intercept ~0 keeps the deviation form.
        if pos in ("2B", "SS"):
            out[f"{pos} DP"] = {"intercept": dp_b,
                                "x": dp_m2b if pos == "2B" else dp_mss,
                                "_fit": dp_prov,
                                "_se": ses[0] if pos == "2B" else ses[1],
                                "_n": len(cells)}

    # ---------- outfield blocks
    for pos in ("LF", "CF", "RF"):
        agg, vis, gt = pos_pivot(POS[pos], 30)
        for name, ycol, dencol, xcol in [
            ("PM%", "pm", "pa", "OF RNG"),
            ("E%",  "e",  "pm", "OF ERR"),
            ("ARM%", "arm", "pa", "OF ARM"),
        ]:
            xb = rating_mean(hitters, vis, xcol)
            ygt = gt[ycol] / gt[dencol]
            ys = [agg[i][ycol] / agg[i][dencol] - ygt for i in vis]
            xs = [[R(i, xcol) - xb for i in vis]]
            b, m1 = linest(ys, xs)
            out[f"{pos} {name}"] = {"intercept": b, "x": m1}

    # ---------- catcher blocks
    agg, vis, gt = pos_pivot(POS["C"], 30)
    for name, yfn, xcol in [
        ("FRM", lambda i: agg[i]["framing"] / agg[i]["ip"] - gt["framing"] / gt["ip"], "C FRM"),
        ("SBA", lambda i: agg[i]["sba"] / agg[i]["ip"] - gt["sba"] / gt["ip"],          "C ARM"),
        ("RTO", lambda i: agg[i]["rto"] / agg[i]["sba"] - gt["rto"] / gt["sba"],        "C ARM"),
    ]:
        xb = rating_mean(hitters, vis, xcol)
        ys = [yfn(i) for i in vis]
        xs = [[R(i, xcol) - xb for i in vis]]
        b, m1 = linest(ys, xs)
        out[f"C {name}"] = {"intercept": b, "x": m1}
    return out

# ---------------------------------------------------------------- driver / validation


def add_fld_derived(fld):
    """The workbook's Fielding table adds pa=Σopps_0..5 and pm=Σopps_made_0..5."""
    for r in fld:
        if "pa" not in r or r.get("pa") is None:
            r["pa"] = sum(r.get(f"opps_{i}") or 0 for i in range(6))
            r["pm"] = sum(r.get(f"opps_made_{i}") or 0 for i in range(6))
    return fld


def load_from_dumps(clone_dirs, ratings_dir, already_archived=None):
    """Pool the career dump CSVs of many clone leagues (each = one complete sample).
    Clones named in `already_archived` are skipped — their rows live in the archive CSVs
    (prevents double-counting when archived clones haven't been deleted yet)."""
    import glob as _glob
    bat, pit, fld = [], [], []
    used = []
    skipped = []
    already_archived = already_archived or set()
    # completeness bar = the max dump year across ALL clones (archived ones included) —
    # otherwise archiving the complete clones would lower the bar and let partial clones in
    def _last_year(dd):
        ys = [int(os.path.basename(y).split("_")[1])
              for y in _glob.glob(os.path.join(dd, "dump", "dump_*_yearly"))]
        return max(ys) if ys else 0
    final_all = max((_last_year(d) for d in clone_dirs), default=0)
    clone_dirs = [d for d in clone_dirs
                  if os.path.basename(d) not in already_archived] \
        or []
    for d in clone_dirs:
        yearly = sorted(_glob.glob(os.path.join(d, "dump", "dump_*_yearly")))
        if not yearly:
            continue
        # completeness rule (same as the ingester): only clones simmed through the final
        # season pool; a clone stopped early would under-weight late-career seasons.
        years = sorted(int(os.path.basename(y).split("_")[1]) for y in yearly)
        if years[-1] < final_all:
            skipped.append(os.path.basename(d))
            continue
        cdir = os.path.join(yearly[-1], "csv")   # career files: last dump has all seasons
        b = os.path.join(cdir, "players_career_batting_stats.csv")
        p = os.path.join(cdir, "players_career_pitching_stats.csv")
        f = os.path.join(cdir, "players_career_fielding_stats.csv")
        if not (os.path.exists(b) and os.path.exists(p) and os.path.exists(f)):
            continue
        bat += load_rows(b); pit += load_rows(p); fld += load_rows(f)
        used.append(os.path.basename(d))
    if skipped:
        print(f"  skipped incomplete clones: {', '.join(skipped)}")
    print(f"  pooled {len(used)} clones: {', '.join(used)}")
    load_from_dumps.used = used
    print(f"  rows: batting={len(bat):,} pitching={len(pit):,} fielding={len(fld):,}")
    hitters = {r["ID"]: r for r in load_rows(os.path.join(ratings_dir, "Hitters.csv")) if isinstance(r.get("ID"), float)}
    pitchers = {r["ID"]: r for r in load_rows(os.path.join(ratings_dir, "Pitchers.csv")) if isinstance(r.get("ID"), float)}
    return bat, pit, add_fld_derived(fld), hitters, pitchers


def drop_partial_final_season(bat, pit, fld):
    """audit B11: a clone pool can contain a partially-simmed final season (BLM had a
    partial 2027 = ~1.3% of pooled PA) which tilts every weighted fit toward a small
    lopsided sample. Drop any FINAL year whose batting-PA total is <50% of the median
    complete year — from the FIT pool only (the archive CSVs keep every row)."""
    pa_by_year = defaultdict(float)
    for r in bat:
        y, pa = r.get("year"), r.get("pa")
        if isinstance(y, (int, float)) and isinstance(pa, (int, float)):
            pa_by_year[y] += pa
    years = sorted(pa_by_year)
    dropped = []
    while len(years) >= 2:
        rest = sorted(pa_by_year[y] for y in years[:-1])
        med = rest[len(rest) // 2]
        if pa_by_year[years[-1]] < 0.5 * med:
            dropped.append(years.pop())
        else:
            break
    if not dropped:
        return bat, pit, fld
    keep = set(years)
    ok = lambda r: not isinstance(r.get("year"), (int, float)) or r["year"] in keep
    print(f"  fit pool: dropped partial final season(s) {sorted(int(y) for y in dropped)} "
          f"({sum(pa_by_year[y] for y in dropped):,.0f} PA vs median complete year {med:,.0f})")
    return [r for r in bat if ok(r)], [r for r in pit if ok(r)], [r for r in fld if ok(r)]


def compute_all(csv_dir=None, dumps=None, ratings_dir=None, league=None):
    """csv_dir and dumps may BOTH be given: the archived table CSVs are pooled together with
    any clone dumps on disk (e.g. BLM's historical sample + newly-simmed clones).

    league picks the 25 Metadata.xlsx whose anchors the pitching fits are centred on
    (audit B12). Only that league's own workbook is ever read."""
    bat, pit, fld = [], [], []
    hitters = pitchers = None
    if csv_dir:
        if os.path.exists(os.path.join(csv_dir, "Batting.csv")):
            bat = load_rows(os.path.join(csv_dir, "Batting.csv"))
            pit = load_rows(os.path.join(csv_dir, "Pitching.csv"))
            fld = add_fld_derived(load_rows(os.path.join(csv_dir, "Fielding.csv")))
            print(f"  archive: batting={len(bat):,} pitching={len(pit):,} fielding={len(fld):,}")
        else:
            print(f"  (no archived data CSVs in {csv_dir} yet - using clone dumps only)")
        hitters = {r["ID"]: r for r in load_rows(os.path.join(csv_dir, "Hitters.csv")) if isinstance(r.get("ID"), float)}
        pitchers = {r["ID"]: r for r in load_rows(os.path.join(csv_dir, "Pitchers.csv")) if isinstance(r.get("ID"), float)}
    if dumps:
        import glob as _glob
        clone_dirs = sorted(_glob.glob(dumps))
        already = set()
        manifest = os.path.join(csv_dir, "archived_clones.txt") if csv_dir else None
        if manifest and os.path.exists(manifest):
            already = {ln.strip() for ln in open(manifest) if ln.strip()}
            hit = [d for d in clone_dirs if os.path.basename(d) in already]
            if hit:
                print(f"  {len(hit)} clone(s) already in the archive - not re-pooled: "
                      + ", ".join(os.path.basename(d) for d in hit))
        b2, p2, f2, h2, pt2 = load_from_dumps(clone_dirs, ratings_dir, already_archived=already)
        bat += b2; pit += p2; fld += f2
        hitters = hitters or h2
        pitchers = pitchers or pt2
        compute_all.newly_pooled = list(getattr(load_from_dumps, "used", []))
    compute_all.pooled = dict(bat=bat, pit=pit, fld=fld)   # archive keeps EVERY row
    bat_f, pit_f, fld_f = drop_partial_final_season(bat, pit, fld)
    anchors = pitch_anchors(league)
    sp = pitching(pit_f, pitchers, "SP", anchors["SP"])
    rp = pitching(pit_f, pitchers, "RP", anchors["RP"])
    # audit B7: fit the real lo-BABIP block from the pooled SP+RP x<50 points (per-role
    # pools are too thin — SP alone has ~10 lo players); both roles share the fit.
    lo_pts = sp.pop("_lBABIP_pts", []) + rp.pop("_lBABIP_pts", [])
    if len(lo_pts) >= 3:
        lo_fit = wls(lo_pts)
        sp["lBABIP"] = lo_fit
        rp["lBABIP"] = lo_fit
        print(f"  lBABIP pooled fit: {len(lo_pts)} SP+RP points -> b={lo_fit[0]:.6g} m={lo_fit[1]:.6g}")
    res = {
        "hitting": hitting(bat_f, hitters),
        "pitching_sp": sp,
        "pitching_rp": rp,
        "fielding": fielding(fld_f, pit_f, hitters),
    }
    return res


def validate(computed, expected_path):
    exp = json.load(open(expected_path))
    ok = bad = 0
    rows = []

    def cmp(name, got, want):
        nonlocal ok, bad
        if want is None or got is None:
            return
        tol = 1e-9 + 1e-6 * abs(want)
        good = abs(got - want) <= tol
        ok += good; bad += (not good)
        rows.append((("PASS" if good else "FAIL"), name, want, got))

    # hitting: hi/lo pairs live in one workbook row per h-label
    hit = computed["hitting"]
    pair = {"hEYE": "lEye", "hPOW": "lPOW", "hK's": "lK's", "hBABIP": "lBABIP",
            "hGAP": "lGAP", "hSPE": "lSPE"}
    for lab, e in exp.get("hitting", {}).items():
        hi = hit.get(lab)
        if hi:
            cmp(f"hit {lab} b", hi[0], e["hi_intercept"]); cmp(f"hit {lab} m", hi[1], e["hi_slope"])
        lo = hit.get(pair.get(lab, ""))
        if lo:
            cmp(f"hit {pair[lab]} b", lo[0], e["lo_intercept"]); cmp(f"hit {pair[lab]} m", lo[1], e["lo_slope"])

    for blk, key in [("pitching_sp", "pitching_sp"), ("pitching_rp", "pitching_rp")]:
        comp = computed[blk]
        for lab, e in exp.get(key, {}).items():
            hi = comp.get(lab)
            if hi:
                cmp(f"{blk} {lab} b", hi[0], e["hi_intercept"]); cmp(f"{blk} {lab} m", hi[1], e["hi_slope"])
            lo = comp.get("l" + lab[1:]) if lab.startswith("h") else None
            if lo and lab != "hBABIP":   # lBABIP mirrors hBABIP in the workbook
                cmp(f"{blk} l{lab[1:]} b", lo[0], e["lo_intercept"]); cmp(f"{blk} l{lab[1:]} m", lo[1], e["lo_slope"])
    for lab, e in exp.get("pitching_shared", {}).items():
        hi = computed["pitching_sp"].get(lab)
        if hi:
            cmp(f"shared {lab} b", hi[0], e["hi_intercept"]); cmp(f"shared {lab} m", hi[1], e["hi_slope"])

    for lab, e in exp.get("fielding", {}).items():
        got = computed["fielding"].get(lab)
        if not isinstance(got, dict):
            continue
        cmp(f"fld {lab} b", got.get("intercept"), e.get("intercept"))
        cmp(f"fld {lab} x", got.get("x"), e.get("x"))
        if e.get("x2") is not None:
            cmp(f"fld {lab} x2", got.get("x2"), e.get("x2"))

    width = max(len(r[1]) for r in rows) if rows else 20
    for st, name, want, got in rows:
        note = "" if st == "PASS" else f"   (want {want:.12g}, got {got:.12g}, rel {abs(got-want)/max(1e-30,abs(want)):.2e})"
        print(f"  {st}  {name:<{width}}{note}")
    print(f"\n  {ok} PASS / {bad} FAIL")
    return bad == 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv-dir", help="dir with Batting/Pitching/Fielding/Hitters/Pitchers.csv")
    ap.add_argument("--dumps", help="glob of clone .lg dirs (e.g. 'C:/OOTP 26/data/saved_games/0tgs*.lg')")
    ap.add_argument("--ratings-dir", help="dir with Hitters.csv/Pitchers.csv (required with --dumps)")
    ap.add_argument("--league", choices=sorted(LEAGUE_DIRS),
                    help="league whose 25 Metadata.xlsx supplies the pitching fit anchors "
                         "(default: the --csv-dir folder name)")
    ap.add_argument("--validate", help="expected_dp.json to compare against")
    ap.add_argument("--json", help="write computed constants to this path")
    ap.add_argument("--archive-dir", help="after a successful compute, write the pooled data rows "
                    "to CSVs here and record the pooled clone names in archived_clones.txt — "
                    "archived clone folders are then safe to delete (see ootp/cleanup_clones.py)")
    a = ap.parse_args()
    if not a.csv_dir and not a.dumps:
        ap.error("need --csv-dir or --dumps (or both to pool archive + new clones)")
    league = a.league or os.path.basename(os.path.normpath(a.csv_dir or a.ratings_dir or ""))
    if league not in LEAGUE_DIRS:
        ap.error(f"can't tell which league this is from {a.csv_dir!r} - pass --league "
                 f"{'|'.join(sorted(LEAGUE_DIRS))} (it selects the metadata anchors the "
                 f"pitching fits are centred on)")
    res = compute_all(csv_dir=a.csv_dir, dumps=a.dumps,
                      ratings_dir=a.ratings_dir or a.csv_dir, league=league)

    if a.archive_dir:
        # secure the pooled sample so the clone folders are no longer the only copy
        pooled = compute_all.pooled
        newly = getattr(compute_all, "newly_pooled", [])
        os.makedirs(a.archive_dir, exist_ok=True)
        for name, rows in [("Batting", pooled["bat"]), ("Pitching", pooled["pit"]), ("Fielding", pooled["fld"])]:
            if not rows:
                continue
            cols = list(rows[0].keys())
            tmp = os.path.join(a.archive_dir, name + ".csv.tmp")
            with open(tmp, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(cols)
                for r in rows:
                    w.writerow(["" if r.get(c) is None else r.get(c) for c in cols])
            os.replace(tmp, os.path.join(a.archive_dir, name + ".csv"))
        manifest = os.path.join(a.archive_dir, "archived_clones.txt")
        have = set()
        if os.path.exists(manifest):
            have = {ln.strip() for ln in open(manifest) if ln.strip()}
        with open(manifest, "a", encoding="utf-8") as f:
            for n in newly:
                if n not in have:
                    f.write(n + "\n")
        print(f"  archive updated: {a.archive_dir}  (batting={len(pooled['bat']):,} rows; "
              f"{len(newly)} newly archived clone(s))")
    if a.json:
        flat = {k: ({kk: list(vv) if isinstance(vv, tuple) else vv for kk, vv in v.items()}) for k, v in res.items()}
        json.dump(flat, open(a.json, "w"), indent=1)
        print(f"wrote {a.json}")
    if a.validate:
        sys.exit(0 if validate(res, a.validate) else 1)


if __name__ == "__main__":
    main()
