"""
scurve_fit.py — D1 rebuild (PREVIEW): fitted logistic S-curves for the four
pitching rate blocks (SO~STU, uBB~CON, HR~HRR, BABIP~PBABIP), SP and RP,
per league, replacing the two-segment linear curves.

Method (audit D1, verifier-corrected transport adapted to a continuous curve):
  1. Empirical pools from the calib archive (same pivot semantics as calibrate.py:
     SP = split_id 1, sum(gs)>=250; RP = split_id 1 & gs=0 rows, sum(g)>=100;
     career rate per player, weight = BF, x = the vR rating).
  2. Fit y = A + B/(1+exp(-k(r-m))) by BF-weighted least squares. No scipy on
     this machine -> constrained grid over (k,m) with the closed-form WLS solve
     for (A,B) at each gridpoint, then two local refinement rounds. Asymptotes
     are bounded near the observed bucket range so 2-5-player tail buckets can't
     run the curve away; direction (sign of B) is enforced per block.
  3. Transport to the live frame:
        live_rate(r) = fitted(r) - E_BFwtd,live[fitted(r)] + league_actual_rate
     with the live population and league rates read straight from the league's
     25 Metadata.xlsx (SP/RP Data + SP/RP Ratings tabs, READ-ONLY), blended
     vR/vL by the role's OVR-vR BF share exactly like the sheet's anchors.
     By construction the BF-weighted projected league rate over the live
     population equals the league's actual rate.
  4. Emit calib/<LG>/scurves-preview.json (a NEW side file — nothing live is
     touched). pitchers.py evaluates these curves only when the JSON is
     explicitly passed (opt-in); the live pipeline is unchanged.

Usage:
  python tgs-viz/engine/scurve_fit.py --league TGS
  python tgs-viz/engine/scurve_fit.py --league BLM
"""
import argparse
import json
import math
import os
import sys
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import calibrate as C  # noqa: E402  (pool machinery reused verbatim)

REPO = os.path.dirname(os.path.dirname(HERE))

# block key -> (vR rating col, vL rating col, direction, y-rate fn name)
BLOCKS = {
    "SO":  ("STU vR", "STU vL", +1),
    "uBB": ("CON vR", "CON vL", -1),
    "HR":  ("HRR vR", "HRR vL", -1),
    "HHR": ("PBABIP vR", "PBABIP vL", -1),
}
SUPPORT = (20.0, 80.0)


# ---------------------------------------------------------------- rates (mirror calibrate.py)
def rate_fns():
    def ubb(a):   return (a["bb"] - a["iw"]) / (a["bf"] - a["iw"] - a["hp"])
    def hrp(a):   return a["hra"] / (a["bf"] - a["bb"] - a["hp"])
    def kp(a):    return a["k"] / (a["bf"] - a["hp"] - a["bb"])
    def babip(a): return (a["ha"] - a["hra"]) / (a["ab"] - a["hra"] - a["k"] + a["sf"])
    return {"SO": kp, "uBB": ubb, "HR": hrp, "HHR": babip}


# ---------------------------------------------------------------- archive pools
def build_pools(csv_dir, anchors):
    """Archive CSVs -> per-role {block: [(rating, rate, bf), ...]} + pool metadata.

    anchors: calibrate.pitch_anchors(league) — the two-segment constants are stated
    about the sheet's Data Points anchor (audit B12), so the two-line reference and
    the pool-fidelity refit below must use that same centre. The S-curve fit itself
    never touches xbar."""
    bat = C.load_rows(os.path.join(csv_dir, "Batting.csv"))
    pit = C.load_rows(os.path.join(csv_dir, "Pitching.csv"))
    pitchers = {r["ID"]: r for r in C.load_rows(os.path.join(csv_dir, "Pitchers.csv"))
                if isinstance(r.get("ID"), float)}
    bat, pit, _ = C.drop_partial_final_season(bat, pit, [])
    fns = rate_fns()
    pools = {}
    for role in ("SP", "RP"):
        if role == "SP":
            agg, _ = C.group_sum(pit, "player_id", C.PIT_FIELDS,
                                 where=lambda r: r.get("split_id") == 1)
            vis = sorted(i for i, a in agg.items() if a["gs"] >= 250)
        else:
            agg, _ = C.group_sum(pit, "player_id", C.PIT_FIELDS,
                                 where=lambda r: r.get("split_id") == 1 and r.get("gs") == 0)
            vis = sorted(i for i, a in agg.items() if a["g"] >= 100)
        gt = C.totals(agg, vis, C.PIT_FIELDS)
        blocks = {}
        for blk, (xcol, _, _) in BLOCKS.items():
            yfn = fns[blk]
            pts = []
            for i in vis:
                p = pitchers.get(i)
                if not p or not isinstance(p.get(xcol), (int, float)):
                    continue
                pts.append((float(p[xcol]), yfn(agg[i]), agg[i]["bf"]))
            blocks[blk] = pts
        pools[role] = {
            "blocks": blocks,
            "gt_rate": {blk: fns[blk](gt) for blk in BLOCKS},
            "xbar": {blk: anchors[role][BLOCKS[blk][0]] for blk in BLOCKS},
            "n": len(vis),
        }
    return pools


def bucket_table(pts):
    """(rating, rate, bf) points -> {rung: (n, bf, bf-wtd mean rate)} on 5-point rungs."""
    agg = defaultdict(lambda: [0, 0.0, 0.0])
    for r, y, w in pts:
        rung = round(r / 5) * 5
        a = agg[rung]
        a[0] += 1
        a[1] += w
        a[2] += y * w
    return {rung: (n, bf, wy / bf) for rung, (n, bf, wy) in sorted(agg.items()) if bf > 0}


# ---------------------------------------------------------------- logistic fit
def logistic(r, A, B, k, m):
    return A + B / (1.0 + math.exp(-k * (r - m)))


def _solve_AB(pts, k, m):
    """Closed-form WLS for y = A + B*s(r) at fixed (k, m)."""
    sw = sws = swy = swss = swsy = 0.0
    for r, y, w in pts:
        s = 1.0 / (1.0 + math.exp(-k * (r - m)))
        sw += w; sws += s * w; swy += y * w; swss += s * s * w; swsy += s * y * w
    det = sw * swss - sws * sws
    if abs(det) < 1e-12:
        return None
    B = (sw * swsy - sws * swy) / det
    A = (swy - B * sws) / sw
    return A, B


def _sse(pts, A, B, k, m):
    return sum(w * (y - logistic(r, A, B, k, m)) ** 2 for r, y, w in pts)


def fit_logistic(pts, direction, buckets):
    """BF-weighted constrained fit. Asymptotes bounded near the observed bucket
    range (tail buckets are 2-5 players -> restraint mandatory); sign(B) must
    match the block's known direction. Grid over (k, m) + closed-form (A, B),
    then two shrinking refinement rounds."""
    bvals = [v[2] for v in buckets.values()]
    bmin, bmax = min(bvals), max(bvals)
    span = max(bmax - bmin, 1e-4)
    lo_bound = max(bmin - 1.5 * span, 1e-4)
    hi_bound = min(bmax + 1.5 * span, 0.7)

    def ok(A, B):
        if direction > 0 and B <= 0:
            return False
        if direction < 0 and B >= 0:
            return False
        asy = (A, A + B)   # r -> -inf, r -> +inf
        return (lo_bound <= min(asy)) and (max(asy) <= hi_bound)

    def scan(ks, ms):
        best = None
        for k in ks:
            for m in ms:
                sol = _solve_AB(pts, k, m)
                if not sol or not ok(*sol):
                    continue
                A, B = sol
                e = _sse(pts, A, B, k, m)
                if best is None or e < best[0]:
                    best = (e, A, B, k, m)
        return best

    ks = [0.01 * 1.23 ** i for i in range(24)]            # 0.01 .. ~1.0
    ms = [10 + 2.5 * i for i in range(37)]                # 10 .. 100
    best = scan(ks, ms)
    if best is None:                                      # bounds too tight -> relax once
        lo_bound, hi_bound = 1e-4, 0.7
        best = scan(ks, ms)
    for shrink in (0.35, 0.12):
        _, A, B, k, m = best
        ks = [k * (1 + shrink * (i - 4) / 4.0) for i in range(9)]
        ms = [m + 12 * shrink * (i - 4) / 4.0 for i in range(9)]
        cand = scan([x for x in ks if x > 1e-4], ms)
        if cand and cand[0] < best[0]:
            best = cand
    e, A, B, k, m = best
    return {"A": A, "B": B, "k": k, "m": m, "sse": e}


def fit_elite_knot(pts, f_logi, buckets, region_start=70.0):
    """Elite-K knot (user-approved D1 tail extension): the bounded logistic
    saturates below the sim's real 75->80 K% jump (clone SP emp +3.9 pts vs
    ~+0.5 projected; RP +4.6 vs +0.8). Add ONE hinge  s*max(0, r-K0)  on top
    of the fitted logistic. Nothing invented:
      * K0 candidates are the OBSERVED archive bucket rungs >= region_start
        (70), so the 20..70 mid-range is untouched by construction;
      * s is the BF-weighted through-origin WLS coefficient of the logistic's
        own residuals on (r-K0) over the players above K0;
      * monotonicity holds by s >= 0 (a negative fit -> no knot);
      * K0 is picked by BF-weighted SSE over the r >= region_start points.
    Returns {"r": K0, "s": s} in the ARCHIVE frame, or None (no knot)."""
    cands = [r for r in sorted(buckets) if region_start <= r < SUPPORT[1]]
    tail = [(r, y, w) for r, y, w in pts if r >= region_start]
    if not tail or not cands:
        return None
    base = sum(w * (y - f_logi(r)) ** 2 for r, y, w in tail)
    best = None
    for K0 in cands:
        num = den = 0.0
        for r, y, w in pts:
            if r > K0:
                num += w * (y - f_logi(r)) * (r - K0)
                den += w * (r - K0) ** 2
        if den <= 0:
            continue
        s = num / den
        if s <= 0:
            continue
        sse = sum(w * (y - f_logi(r) - s * max(0.0, r - K0)) ** 2 for r, y, w in tail)
        if best is None or sse < best[0]:
            best = (sse, K0, s)
    if best is None or best[0] >= base:
        return None
    return {"r": best[1], "s": best[2], "sse_tail": best[0], "sse_tail_no_knot": base}


def two_line(r, consts, xbar, gt):
    """The CURRENT two-segment curve, archive frame: gt + b + m*(r - xbar)."""
    b_hi, m_hi = consts["hi"]
    b_lo, m_lo = consts["lo"]
    if r >= 50:
        return gt + b_hi + m_hi * (r - xbar)
    return gt + b_lo + m_lo * (r - xbar)


# ---------------------------------------------------------------- live metadata (read-only)
def _num(v):
    """Metadata input tabs are pasted text — coerce '740' -> 740.0; else None."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _read_table(ws, hdr_row=2):
    """One metadata input tab -> (headers list, rows list of tuple). Headers are
    normalized (sort-arrow decorations like 'BF\\xa0▴' -> 'BF')."""
    rows = list(ws.iter_rows(values_only=True))
    hdr = []
    for h in rows[hdr_row - 1]:
        s = str(h).strip() if h is not None else ""
        s = s.replace("\xa0", " ").replace("▴", "").replace("▾", "").strip()
        hdr.append(s)
    return hdr, rows[hdr_row:]


def read_metadata(league):
    """25 Metadata.xlsx SP/RP Data + Ratings tabs. READ-ONLY. Returns per role:
    league actual rates + the two ratings tables (list of dicts) + OVR-vR share."""
    from openpyxl import load_workbook
    path = os.path.join(REPO, f"The Sheets {league}", "25 Metadata.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    out = {}
    for role in ("SP", "RP"):
        hdr, rows = _read_table(wb[f"{role} Data"])
        idx = {h: i for i, h in enumerate(hdr) if h}
        T = defaultdict(float)
        per_id = {}
        nrows = 0
        for r in rows:
            if r[idx["ID"]] is None:
                continue
            nrows += 1
            rec = {}
            for c in ("BF", "AB", "1B", "2B", "3B", "HR", "BB", "IBB", "K", "HP", "SF"):
                v = _num(r[idx[c]]) or 0.0
                T[c] += v
                rec[c] = v
            per_id[_num(r[idx["ID"]])] = rec
        rates = {
            "uBB": (T["BB"] - T["IBB"]) / (T["BF"] - T["IBB"] - T["HP"]),
            "HR":  T["HR"] / (T["BF"] - T["HP"] - T["BB"]),
            "SO":  T["K"] / (T["BF"] - T["HP"] - T["BB"]),
            "HHR": (T["1B"] + T["2B"] + T["3B"]) / (T["AB"] - T["HR"] + T["SF"] - T["K"]),
        }
        # ratings tabs: vR table = cols 1..14, vL table = cols 18..31 (both carry
        # all rating cols; BF is BF vs that side) — same layout both leagues.
        hdr2, rows2 = _read_table(wb[f"{role} Ratings"])
        def tab(css):
            cols = {hdr2[i]: i for i in css if i < len(hdr2) and hdr2[i]}
            recs = []
            for r in rows2:
                vid = _num(r[cols["ID"]]) if cols["ID"] < len(r) else None
                bf = _num(r[cols["BF"]]) if cols["BF"] < len(r) else None
                if vid is None or bf is None:
                    continue
                rec = {h: _num(r[i]) for h, i in cols.items()}
                rec["ID"], rec["BF"] = vid, bf
                recs.append(rec)
            return recs
        vr, vl = tab(range(0, 14)), tab(range(17, 31))
        bf_vr = sum(x["BF"] for x in vr)
        bf_vl = sum(x["BF"] for x in vl)
        out[role] = {
            "rates": rates, "vr": vr, "vl": vl,
            "share": bf_vr / (bf_vr + bf_vl),
            "stat_rows": nrows, "T": dict(T), "per_id": per_id,
        }
    wb.close()
    return out


def live_points(mrole, blk, xcol, xcol_vl):
    """Per-player live points (blended rating, season rate, weight=denominator)
    for the transport's live-bucket calibration (audit D1: 'calibrate the
    transport against the live-league K%-by-STU buckets')."""
    share = mrole["share"]
    rat_vr = {x["ID"]: x.get(xcol) for x in mrole["vr"]}
    rat_vl = {x["ID"]: x.get(xcol_vl) for x in mrole["vl"]}
    pts = []
    for pid, s in mrole["per_id"].items():
        rv, rl = rat_vr.get(pid), rat_vl.get(pid)
        if rv is None and rl is None:
            continue
        r = (share * rv + (1 - share) * rl) if (rv is not None and rl is not None) \
            else (rv if rv is not None else rl)
        if blk == "SO":
            den, y_num = s["BF"] - s["HP"] - s["BB"], s["K"]
        elif blk == "uBB":
            den, y_num = s["BF"] - s["IBB"] - s["HP"], s["BB"] - s["IBB"]
        elif blk == "HR":
            den, y_num = s["BF"] - s["HP"] - s["BB"], s["HR"]
        else:  # HHR
            den, y_num = s["AB"] - s["HR"] + s["SF"] - s["K"], s["1B"] + s["2B"] + s["3B"]
        if den <= 0:
            continue
        pts.append((float(r), y_num / den, den))
    return pts


def calibrate_scale(f, pts):
    """Fit the live-vs-clone rating-scale drift: y ~ f(a*(r-50)+50+c) + e0 over the
    live per-player points (e0 absorbed analytically; the exact-mean offset
    replaces it later). Returns (a, c, sse, sse_at_identity). Falls back to
    (1, 0) when the improvement is negligible (<1% relative)."""
    def sse_at(a, c):
        sw = swy = 0.0
        vals = []
        for r, y, w in pts:
            rr = min(max(r, SUPPORT[0]), SUPPORT[1])
            # verifier's deep-tail clamp: the drift map is an affine fit over the
            # mid-range live points — its archive-equivalent rating is only
            # trusted inside the clone support, so clamp to [20,80] before
            # evaluating the clone-shape logistic.
            ar = min(max(a * (rr - 50.0) + 50.0 + c, SUPPORT[0]), SUPPORT[1])
            v = f(ar)
            vals.append((v, y, w))
            sw += w
            swy += w * (y - v)
        e0 = swy / sw
        return sum(w * (y - v - e0) ** 2 for v, y, w in vals)

    base = sse_at(1.0, 0.0)
    best = (base, 1.0, 0.0)
    a_grid = [0.6 + 0.05 * i for i in range(21)]           # 0.6 .. 1.6
    c_grid = [-15 + 2.5 * i for i in range(13)]            # -15 .. +15
    for a in a_grid:
        for c in c_grid:
            e = sse_at(a, c)
            if e < best[0]:
                best = (e, a, c)
    for shrink in (0.4, 0.15):
        _, a0, c0 = best
        for a in [max(0.6, min(1.6, a0 + 0.1 * shrink * (i - 4))) for i in range(9)]:
            for c in [c0 + 5 * shrink * (i - 4) for i in range(9)]:
                e = sse_at(a, c)
                if e < best[0]:
                    best = (e, a, c)
    e, a, c = best
    if base - e < 0.01 * base:      # not a real improvement -> keep the pure clone shape
        return 1.0, 0.0, base, base
    # a solution pinned at the scale bounds on a weak improvement is the optimizer
    # chasing noise in an unidentifiable direction -> keep the clone shape
    if (abs(a - 0.6) < 1e-9 or abs(a - 1.6) < 1e-9) and base - e < 0.03 * base:
        return 1.0, 0.0, base, base
    return a, c, e, base


def live_mean(fitfn, recs, col):
    """BF-weighted mean of fitfn(rating) over one ratings table (support-clamped)."""
    sw = swy = 0.0
    for x in recs:
        r = x.get(col)
        if not isinstance(r, (int, float)):
            continue
        rr = min(max(float(r), SUPPORT[0]), SUPPORT[1])
        w = float(x["BF"])
        sw += w
        swy += w * fitfn(rr)
    return swy / sw


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--league", required=True, choices=["TGS", "BLM"])
    ap.add_argument("--out", help="output json (default calib/<LG>/scurves-preview.json)")
    a = ap.parse_args()
    lg = a.league
    csv_dir = os.path.join(HERE, "calib", lg)
    out_path = a.out or os.path.join(csv_dir, "scurves-preview.json")

    print(f"== {lg}: archive pools ==")
    pools = build_pools(csv_dir, C.pitch_anchors(lg))

    # current two-segment constants for the honesty comparison
    consts = json.load(open(os.path.join(csv_dir, "constants-latest.json")))
    seg_key = {"SO": "STU", "uBB": "CON", "HR": "HRR", "HHR": "BABIP"}

    # pool-fidelity check: refit the two-segment WLS from THESE pools; it must
    # reproduce constants-latest.json (proves the pools = the live fit pools).
    worst = 0.0
    for role in ("SP", "RP"):
        role_c = consts[f"pitching_{role.lower()}"]
        for blk, key in seg_key.items():
            pts = pools[role]["blocks"][blk]
            xbar, gtr = pools[role]["xbar"][blk], pools[role]["gt_rate"][blk]
            for name, filt in (("h" + key, lambda x: x >= 50), ("l" + key, lambda x: x < 50)):
                if key == "BABIP" and name.startswith("h"):
                    filt = lambda x: x >= 20          # workbook hBABIP spans the whole sample
                dev = [(r - xbar, y - gtr, w) for r, y, w in pts if filt(r)]
                if name == "lBABIP":
                    continue                          # pooled SP+RP fit — checked separately
                b, m = C.wls(dev)
                want = role_c[name]
                worst = max(worst, abs(b - want[0]) / (1e-9 + abs(want[0])),
                            abs(m - want[1]) / (1e-9 + abs(want[1])))
    print(f"  pool fidelity vs constants-latest.json: worst rel diff {worst:.2e} "
          f"({'OK' if worst < 1e-6 else '!! POOLS DIFFER FROM LIVE FIT'})")

    print(f"== {lg}: live metadata (25 Metadata.xlsx, read-only) ==")
    meta = read_metadata(lg)

    result = {"league": lg, "support": list(SUPPORT), "roles": {}}
    for role in ("SP", "RP"):
        pool = pools[role]
        mrole = meta[role]
        result["roles"][role] = {"blocks": {}, "share_vR": mrole["share"],
                                 "league_rates": mrole["rates"], "pool_n": pool["n"]}
        print(f"\n---- {lg} {role} (pool n={pool['n']}, live share vR={mrole['share']:.4f}) ----")
        for blk, (xcol, xcol_vl, direction) in BLOCKS.items():
            pts = pool["blocks"][blk]
            buckets = bucket_table(pts)
            fit = fit_logistic(pts, direction, buckets)
            f_logi = lambda r: logistic(r, fit["A"], fit["B"], fit["k"], fit["m"])
            # elite-K knot: SO block only (the audited 75->80 jump). All other
            # blocks keep the pure logistic.
            knot = fit_elite_knot(pts, f_logi, buckets) if blk == "SO" else None
            if knot:
                f_arc = lambda r: f_logi(r) + knot["s"] * max(0.0, r - knot["r"])
            else:
                f_arc = f_logi

            # two-line reference in the archive frame
            key = seg_key[blk]
            role_c = consts[f"pitching_{role.lower()}"]
            seg = {"hi": role_c["h" + key], "lo": role_c["l" + key]}
            xbar, gtr = pool["xbar"][blk], pool["gt_rate"][blk]

            # live rating-scale drift (audit D1: transport calibrated against the
            # live-league buckets). An affine rating map composed with a logistic
            # is still a logistic: k* = k*a, m* = (m-50-c)/a + 50 — so the final
            # curve stays 4-parameter; A/B (the asymptote span) stay the clone's.
            lpts = live_points(mrole, blk, xcol, xcol_vl)
            # the drift map is fitted on the PURE logistic: it is a mid-range
            # affine fit, and keeping the hinge out leaves (a,c) — and with them
            # the whole 40-70 curve — identical to the pre-knot fit (gate: the
            # knot must not move the mid-range).
            a_sc, c_sc, sse_live, sse_id = calibrate_scale(f_logi, lpts)
            fit_k = fit["k"] * a_sc
            fit_m = (fit["m"] - 50.0 - c_sc) / a_sc + 50.0
            # verifier's deep-tail clamp, applied to the FINAL curve too: the
            # archive-equivalent rating a*(r-50)+50+c is clamped to the clone
            # support [20,80] before the clone-shape logistic is evaluated.
            # On the composed live-frame logistic (a>0, monotone map) that is
            # exactly a NARROWED live support interval — ratings past these
            # bounds evaluate at the bound. Emitted as the block's "support",
            # so pitchers.py sc_rate needs no change.
            sup_lo = max(SUPPORT[0], (SUPPORT[0] - 50.0 - c_sc) / a_sc + 50.0)
            sup_hi = min(SUPPORT[1], (SUPPORT[1] - 50.0 - c_sc) / a_sc + 50.0)

            def f_base(r, lo=sup_lo, hi=sup_hi, A=fit["A"], B=fit["B"],
                       kk=fit_k, mm=fit_m):
                return logistic(min(max(r, lo), hi), A, B, kk, mm)

            # ---- elite-K hinge, live frame (SO only) ----
            # The hinge arm is  s * max(0, min(r, hinge_hi) - K0_live).  Two
            # fitted variants, per the audit's own transport rule ("calibrate
            # the transport against the live-league K%-by-STU buckets"):
            #   * LIVE-FITTED (preferred, whenever live arms sit above the
            #     drift-mapped knot): s = BF-weighted through-origin WLS of the
            #     live residuals (y - logistic - offset0) on the hinge arm,
            #     hinge_hi = 80 (the live support top). s >= 0 keeps it
            #     monotone; a fitted 0 is kept as a measured live zero.
            #   * ARCHIVE-TRANSPORT (fallback when the live league has no arms
            #     above the knot to measure): s = s_arc * a, hinge_hi = sup_hi,
            #     i.e. exactly the archive-certified 5-point jump under the
            #     drift map, never extrapolated past archive support.
            # Rationale: with a > 1 the drift map compresses live 75/80 onto
            # archive 80 and would BOTH flatten the live elite ladder and
            # overshoot its level (verified +6 pts at live STU-75, TGS RP);
            # the live elite buckets are the out-of-sample truth the audit
            # says to calibrate against, so they win when they exist.
            kn_live = None
            if knot:
                K0_live = (knot["r"] - 50.0 - c_sc) / a_sc + 50.0
                E0 = (mrole["share"] * live_mean(f_base, mrole["vr"], xcol)
                      + (1 - mrole["share"]) * live_mean(f_base, mrole["vl"], xcol_vl))
                off0 = mrole["rates"][blk] - E0
                # ENGINE-FRAME tail points: the engine evaluates the curve at
                # each side's RAW rating and share-blends the statlines, so the
                # regressor is the share-blended hinge ARM (not the arm of the
                # blended rating) and the residual is vs the share-blended base.
                share = mrole["share"]
                rat_vr = {x["ID"]: x.get(xcol) for x in mrole["vr"]}
                rat_vl = {x["ID"]: x.get(xcol_vl) for x in mrole["vl"]}
                arm = lambda r: max(0.0, min(r, SUPPORT[1]) - K0_live)
                tail = []
                for pid, st in mrole["per_id"].items():
                    rv, rl = rat_vr.get(pid), rat_vl.get(pid)
                    if rv is None and rl is None:
                        continue
                    rv = rv if rv is not None else rl
                    rl = rl if rl is not None else rv
                    dn = st["BF"] - st["HP"] - st["BB"]
                    if dn <= 0:
                        continue
                    armb = share * arm(rv) + (1 - share) * arm(rl)
                    if armb <= 0:
                        continue
                    basey = share * f_base(rv) + (1 - share) * f_base(rl) + off0
                    tail.append((armb, st["K"] / dn - basey, dn))
                if tail:
                    num = sum(w * res * a for a, res, w in tail)
                    den = sum(w * a * a for a, res, w in tail)
                    s_live = max(num / den, 0.0) if den > 0 else 0.0
                    kn_live = {"r": K0_live, "s": s_live, "hi": SUPPORT[1],
                               "fit": "live", "n_live_tail": len(tail),
                               "bf_live_tail": sum(w for _, _, w in tail)}
                else:
                    kn_live = {"r": K0_live, "s": knot["s"] * a_sc, "hi": sup_hi,
                               "fit": "archive-transport", "n_live_tail": 0}

            def f(r, kn=kn_live):
                v = f_base(r)
                if kn:
                    v += kn["s"] * max(0.0, min(r, kn["hi"]) - kn["r"])
                return v

            # transport: exact-mean renormalization over the live population
            E_live = (mrole["share"] * live_mean(f, mrole["vr"], xcol)
                      + (1 - mrole["share"]) * live_mean(f, mrole["vl"], xcol_vl))
            offset = mrole["rates"][blk] - E_live

            # bucket residual table (ARCHIVE frame: the clone-shape fit vs the
            # empirical bucket means vs the current two-line fit)
            rows = []
            for rung, (n, bf, emp) in buckets.items():
                if not (SUPPORT[0] <= rung <= SUPPORT[1]):
                    continue
                sig = f_arc(rung)
                two = two_line(rung, seg, xbar, gtr)
                rows.append((rung, n, bf, emp, sig, two))
            wr_sig = math.sqrt(sum(bf * (emp - sig) ** 2 for _, _, bf, emp, sig, _ in rows)
                               / sum(bf for _, _, bf, *_ in rows))
            wr_two = math.sqrt(sum(bf * (emp - two) ** 2 for _, _, bf, emp, _, two in rows)
                               / sum(bf for _, _, bf, *_ in rows))
            print(f"  {blk:4} ({xcol:9}) arc A={fit['A']:.5f} B={fit['B']:+.5f} "
                  f"k={fit['k']:.4f} m={fit['m']:.2f}  "
                  f"live-scale a={a_sc:.3f} c={c_sc:+.1f} "
                  f"(sse {sse_live/sse_id:.4f} of identity)  "
                  f"final k={fit_k:.4f} m={fit_m:.2f} offset={offset:+.5f}")
            if knot:
                print(f"       elite-K knot: archive K0={knot['r']:.0f} "
                      f"s={knot['s']:+.6f}/pt (tail sse {knot['sse_tail']:.5f} vs "
                      f"{knot['sse_tail_no_knot']:.5f} without)  live K0={kn_live['r']:.2f} "
                      f"s={kn_live['s']:+.6f}/pt hi={kn_live['hi']:.2f} "
                      f"[{kn_live['fit']}, n_tail={kn_live['n_live_tail']}]")
            elif blk == "SO":
                print("       elite-K knot: NO KNOT (fitted slope <= 0 or no tail improvement)")
            print(f"       archive bucketRMSE: sigmoid={wr_sig:.5f} twoline={wr_two:.5f} "
                  f"{'SIGMOID' if wr_sig <= wr_two else '** TWO-LINE BETTER **'}")
            for rung, n, bf, emp, sig, two in rows:
                print(f"      r={rung:3.0f} n={n:4d}  emp={emp:.4f}  sig={sig:.4f} "
                      f"({sig-emp:+.4f})  two={two:.4f} ({two-emp:+.4f})")

            result["roles"][role]["blocks"][blk] = {
                "A": fit["A"], "B": fit["B"], "k": fit_k, "m": fit_m,
                "offset": offset, "support": [sup_lo, sup_hi],
                "knot": ({"r": kn_live["r"], "s": kn_live["s"], "hi": kn_live["hi"],
                          "fit": kn_live["fit"], "n_live_tail": kn_live["n_live_tail"]}
                         if kn_live else None),
                "direction": direction, "x_vR": xcol, "x_vL": xcol_vl,
                "league_rate": mrole["rates"][blk], "E_live_fitted": E_live,
                "archive_fit": {"k": fit["k"], "m": fit["m"], "sse": fit["sse"],
                                "elite_knot": knot},
                "live_scale": {"a": a_sc, "c": c_sc, "sse_ratio": sse_live / sse_id,
                               "n_live": len(lpts)},
                "bucket_rmse_sigmoid": wr_sig, "bucket_rmse_twoline": wr_two,
                "buckets": {str(int(rung)): {"n": n, "bf": bf, "emp": emp,
                                             "sig": sig, "two": two}
                            for rung, n, bf, emp, sig, two in rows},
            }

            # monotonicity scan (gate iii)
            worst_step = 0.0
            bad = 0
            prev = f(SUPPORT[0]) + offset
            r = SUPPORT[0] + 0.1
            while r <= SUPPORT[1] + 1e-9:
                cur = f(r) + offset
                step = (cur - prev) * direction
                if step < -1e-12:
                    bad += 1
                worst_step = max(worst_step, abs(cur - prev))
                prev = cur
                r += 0.1
            result["roles"][role]["blocks"][blk]["monotone_ok"] = (bad == 0)
            result["roles"][role]["blocks"][blk]["max_step_0p1"] = worst_step
            if bad:
                print(f"      !! MONOTONICITY VIOLATIONS: {bad}")

        # ---- live out-of-sample truth ladder: K% by STU vR bucket (gate ii) ----
        blkP = result["roles"][role]["blocks"]["SO"]

        def fK(r, b=blkP):
            rr = min(max(r, b["support"][0]), b["support"][1])
            v = logistic(rr, b["A"], b["B"], b["k"], b["m"]) + b["offset"]
            kn = b.get("knot")
            if kn:
                v += kn["s"] * max(0.0, min(r, kn["hi"]) - kn["r"])
            return v
        share = result["roles"][role]["share_vR"]
        rat_vl_lad = {x["ID"]: x.get(result["roles"][role]["blocks"]["SO"]["x_vL"])
                      for x in mrole["vl"]}
        lb = defaultdict(lambda: [0, 0.0, 0.0, 0.0])  # rung -> [n, K, den, den*proj_eng]
        for x in mrole["vr"]:
            stat = mrole["per_id"].get(x["ID"])
            r = x.get("STU vR")
            if not stat or not isinstance(r, (int, float)):
                continue
            rung = round(float(r) / 5) * 5
            den = stat["BF"] - stat["HP"] - stat["BB"]
            if den <= 0:
                continue
            rl = rat_vl_lad.get(x["ID"])
            rl = float(rl) if isinstance(rl, (int, float)) else float(r)
            # engine-frame projection: curve at each side's RAW rating, share-blended
            peng = share * fK(float(r)) + (1 - share) * fK(rl)
            lb[rung][0] += 1
            lb[rung][1] += stat["K"]
            lb[rung][2] += den
            lb[rung][3] += den * peng
        ladder = {}
        print(f"  live K%-by-STU ladder ({role}):  [proj=curve@rung, "
              f"projEng=den-wtd engine-frame projection of the rung's arms]")
        for rung in sorted(lb):
            n, kk, den, dpe = lb[rung]
            live = kk / den
            proj = fK(rung)
            peng = dpe / den
            ladder[str(int(rung))] = {"n": n, "live": live, "proj": proj,
                                      "proj_eng": peng}
            print(f"      STU={rung:3.0f} n={n:3d}  live={live:.4f}  proj={proj:.4f} "
                  f"({proj-live:+.4f})  projEng={peng:.4f} ({peng-live:+.4f})")
        result["roles"][role]["live_k_by_stu"] = ladder
        gap = fK(50) - fK(45)
        result["roles"][role]["k_gap_45_50"] = gap
        print(f"  projected 45->50 K% gap ({role}): {gap*100:+.2f} pts")

    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(result, fh, indent=1)
    print(f"\nwrote {out_path}")


if __name__ == "__main__":
    main()
