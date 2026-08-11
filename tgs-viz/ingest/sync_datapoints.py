"""
sync_datapoints.py  -  Push regression + metadata constants into the projection sheets.

The projection workbooks ("The Sheet Hitters/Pitchers.xlsx") read a "Data Points"
tab of constants that the engine + web-app pipeline consume. Those constants are
maintained in two other workbooks:

    25 Regressions.xlsx  ->  the rating->stat regression slopes/intercepts
    25 Metadata.xlsx     ->  the anchors, league rates, fielding ratings, pos-adj

This script copies those numbers into each league's projection sheets' Data Points,
matching cells by their LABELS (not fixed addresses), so it survives layout tweaks.
It NEVER writes to the source workbooks and edits the projection files as raw XML so
their PivotTables and charts are preserved byte-for-byte.

Workflow:
    1. Refresh + recalculate 25 Regressions.xlsx and 25 Metadata.xlsx in Excel, save.
    2. Run this (dry-run):   python tgs-viz/ingest/sync_datapoints.py
       Review the planned changes, then:   ... --write
    3. Next StatsPlus pull (Get StatsPlus Ratings.bat) rebuilds the web-app data.

Usage:
    python tgs-viz/ingest/sync_datapoints.py [--league BLM|TGS|both] [--write]

    # Python-calibration mode: take the regression constants from a calibrate.py
    # JSON instead of the 25 Regressions.xlsx values (the workbook is still read,
    # but only for its LABEL LAYOUT; every regression value comes from the JSON):
    python tgs-viz/ingest/sync_datapoints.py --league TGS --calib tgs-viz/engine/calib/TGS/constants-260seasons.json
"""
import openpyxl, zipfile, re, sys, os, json, shutil, datetime, argparse
from collections import defaultdict, deque
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LEAGUE_DIRS = {'BLM': 'The Sheets BLM', 'TGS': 'The Sheets TGS'}
ERR_STRINGS = {'#DIV/0!', '#N/A', '#VALUE!', '#REF!', '#NAME?', '#NULL!', '#NUM!'}

_col = lambda c: re.match(r'[A-Z]+', c).group()
_row = lambda c: int(re.match(r'[A-Z]+(\d+)', c).group(1))
_cell = lambda c, r: f'{c}{r}'


def load_vals(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    d = {}
    r = 0
    for row in ws.iter_rows(values_only=True):
        r += 1
        for c, v in enumerate(row, start=1):
            if v is not None and (not isinstance(v, str) or v.strip() != ''):
                d[f'{get_column_letter(c)}{r}'] = v
    wb.close()
    return d


def labels_in_col(cells, colL, rmin, rmax):
    out = [(_row(k), v.strip()) for k, v in cells.items()
           if isinstance(v, str) and _col(k) == colL and rmin <= _row(k) <= rmax]
    return sorted(out)


def ordered_src(cells, labelcol, valcol, rmin, rmax):
    q = defaultdict(deque)
    for r, lab in labels_in_col(cells, labelcol, rmin, rmax):
        if _cell(valcol, r) in cells:
            q[lab].append(r)
    return q


def fielding_blocks(cells, poscol, metcol, valcols):
    posrows = {r: v for r, v in labels_in_col(cells, poscol, 1, 60)}
    res = []
    for r, lab in labels_in_col(cells, metcol, 1, 60):
        p = None
        for pr in sorted(posrows):
            if pr <= r:
                p = posrows[pr]
        vr = r + 1
        vc = [_cell(v, vr) for v in valcols if _cell(v, vr) in cells]
        if p and vc and lab not in ('X', 'X2', 'X3'):
            res.append((p, lab, vc))
    return res


def override_reg_values(REG, calib):
    """Replace the regression VALUES in the loaded REG cell-dict with the constants from a
    calibrate.py JSON, keeping the workbook's label layout. After this, the normal mapping
    machinery pushes the Python-calibrated numbers exactly where the workbook's went."""
    hit = calib.get('hitting', {})
    sp, rp = calib.get('pitching_sp', {}), calib.get('pitching_rp', {})
    fld = calib.get('fielding', {})
    pair = {'hEYE': 'lEye', 'hPOW': 'lPOW', "hK's": "lK's", 'hBABIP': 'lBABIP',
            'hGAP': 'lGAP', 'hSPE': 'lSPE'}
    n = 0

    # hitting block: labels col B; values row+1 in B/C (hi) and D/E (lo)
    for r, lab in labels_in_col(REG, 'B', 1, 25):
        if lab in hit:
            b, m = hit[lab][0], hit[lab][1]
            REG[_cell('B', r + 1)], REG[_cell('C', r + 1)] = b, m
            n += 2
            lo = hit.get(pair.get(lab, ''))
            if lo and _cell('D', r + 1) in REG:
                REG[_cell('D', r + 1)], REG[_cell('E', r + 1)] = lo[0], lo[1]
                n += 2

    # pitching blocks: labels col N; SP rows <=11, RP rows >=12; SB%/SBA always from SP sheet
    for r, lab in labels_in_col(REG, 'N', 1, 26):
        src = sp if (r <= 11 or lab in ('SB%', 'SBA')) else rp
        if lab not in src:
            continue
        b, m = src[lab][0], src[lab][1]
        REG[_cell('N', r + 1)], REG[_cell('O', r + 1)] = b, m
        n += 2
        if lab.startswith('h') and _cell('P', r + 1) in REG:
            # lo pair: real l-block. The workbook aliased lBABIP to the hi fit (audit B7);
            # calibrate.py now emits a real pooled lBABIP — prefer it, fall back to the
            # alias only for old JSONs that lack it.
            lo = src.get('l' + lab[1:]) or (src[lab] if lab == 'hBABIP' else None)
            if lo:
                REG[_cell('P', r + 1)], REG[_cell('Q', r + 1)] = lo[0], lo[1]
                n += 2

    # fielding blocks: pos col H / metric col I / values I,J,K at metric row+1
    for p, mname, vcells in fielding_blocks(REG, 'H', 'I', ['I', 'J', 'K']):
        key = f'{p} {mname}'
        if key not in fld:
            continue
        v = fld[key]
        vals = [v.get('intercept'), v.get('x'), v.get('x2')]
        for cell, val in zip(vcells, vals):
            if val is not None:
                REG[cell] = val
                n += 1
    return n


def build_mapping(H, P, REG, MET):
    """Return list of dicts: ts('H'/'P'), tc(target cell), src('REG'/'MET'), sc, label, kind."""
    M = []
    add = lambda ts, tc, src, sc, label, kind: M.append(
        dict(ts=ts, tc=tc, src=src, sc=sc, label=label, kind=kind))

    # ---- HITTERS ----
    HITREG = ['hEYE', 'hPOW', "hK's", 'hBABIP', 'hGAP', 'hSPE', 'hSBA', 'SB%', 'UBR']
    tH = {lab: r for r, lab in labels_in_col(H, 'B', 1, 25) if lab in HITREG}
    tR = {lab: r for r, lab in labels_in_col(REG, 'B', 1, 25) if lab in HITREG}
    for lab in HITREG:
        for cc in ['B', 'C', 'D', 'E']:
            if _cell(cc, tH[lab] + 1) in H:
                add('H', _cell(cc, tH[lab] + 1), 'REG', _cell(cc, tR[lab] + 1), lab, 'reg')

    SKIP = {'Hitting Ratings', 'Matchups', 'League Measurements'}
    metEq = ordered_src(MET, 'E', 'F', 1, 45)
    for r, lab in labels_in_col(H, 'G', 1, 45):
        if lab in SKIP or _cell('H', r) not in H:
            continue
        if metEq[lab]:
            add('H', _cell('H', r), 'MET', _cell('F', metEq[lab].popleft()), 'anchor:' + lab, 'meta')

    RATE = ['BB%', 'HR%', 'SO%', 'BABIP', 'XBH%', '3B%', 'SB%', 'UBR', 'SBA%']
    metA = {lab: r for r, lab in labels_in_col(MET, 'A', 30, 45) if lab in RATE}
    for r, lab in labels_in_col(H, 'B', 32, 45):
        if lab in RATE and _cell('C', r) in H and lab in metA:
            add('H', _cell('C', r), 'MET', _cell('B', metA[lab]), 'rate:' + lab, 'meta')

    # fielding regressions  tgt J/K -> K,L,M   src REG H/I -> I,J,K
    sFidx = {}
    for p, m, vc in fielding_blocks(REG, 'H', 'I', ['I', 'J', 'K']):
        sFidx.setdefault((p, m), vc)
    for p, m, vc in fielding_blocks(H, 'J', 'K', ['K', 'L', 'M']):
        if (p, m) in sFidx:
            svc = sFidx[(p, m)]
            for i in range(min(len(vc), len(svc))):
                add('H', vc[i], 'REG', svc[i], f'fld:{p} {m}', 'reg')

    # fielding ratings  tgt O/P -> P,Q   src MET H/I -> I,J
    sFR = defaultdict(list)
    for p, m, vc in fielding_blocks(MET, 'H', 'I', ['I', 'J']):
        sFR[(p, m)].append(vc)
    usedFR = defaultdict(int)
    for p, m, vc in fielding_blocks(H, 'O', 'P', ['P', 'Q']):
        lst = sFR[(p, m)]
        idx = usedFR[(p, m)]
        if idx < len(lst):
            svc = lst[idx]
            usedFR[(p, m)] += 1
            for i in range(min(len(vc), len(svc))):
                add('H', vc[i], 'MET', svc[i], f'fldrat:{p} {m}', 'meta')

    metL = {lab: r for r, lab in labels_in_col(MET, 'L', 1, 45)}
    for r, lab in labels_in_col(H, 'S', 1, 45):
        if _cell('T', r) in H and lab in metL:
            add('H', _cell('T', r), 'MET', _cell('M', metL[lab]), 'fldstat:' + lab, 'meta')

    metO = {lab: r for r, lab in labels_in_col(MET, 'O', 1, 45)}
    for r, lab in labels_in_col(H, 'V', 1, 45):
        if _cell('W', r) in H and lab in metO:
            add('H', _cell('W', r), 'MET', _cell('P', metO[lab]), 'posadj:' + lab, 'meta')

    # ---- PITCHERS ----
    PREG = ['hCON', 'hHRR', 'hSTU', 'hBABIP', 'SB%', 'SBA']
    for (tmin, tmax, smin, smax, blk) in [(1, 9, 1, 11, 'SP'), (12, 26, 12, 26, 'RP')]:
        tPidx = {lab: r for r, lab in labels_in_col(P, 'B', tmin, tmax) if lab in PREG}
        sPidx = {lab: r for r, lab in labels_in_col(REG, 'N', smin, smax) if lab in PREG}
        for lab in PREG:
            if lab in tPidx and lab in sPidx:
                for tc, sc in [('B', 'N'), ('C', 'O'), ('D', 'P'), ('E', 'Q')]:
                    if _cell(tc, tPidx[lab] + 1) in P:
                        add('P', _cell(tc, tPidx[lab] + 1), 'REG', _cell(sc, sPidx[lab] + 1), f'p{blk}:' + lab, 'reg')

    sSq = ordered_src(MET, 'S', 'T', 1, 45)
    for r, lab in labels_in_col(P, 'G', 1, 45):
        if _cell('H', r) not in P or lab in ('Starter Ratings', 'Reliever Ratings', 'Matchups', 'League Stats'):
            continue
        if sSq[lab]:
            add('P', _cell('H', r), 'MET', _cell('T', sSq[lab].popleft()), 'panchor:' + lab, 'meta')

    sVq = ordered_src(MET, 'V', 'W', 1, 45)
    for r, lab in labels_in_col(P, 'J', 1, 45):
        if _cell('K', r) not in P or lab in ('SP', 'RP', 'Pitching Stats'):
            continue
        if sVq[lab]:
            add('P', _cell('K', r), 'MET', _cell('W', sVq[lab].popleft()), 'pstat:' + lab, 'meta')

    for r, lab in labels_in_col(P, 'B', 32, 45):
        if lab in RATE and _cell('C', r) in P and lab in metA:
            add('P', _cell('C', r), 'MET', _cell('B', metA[lab]), 'prate:' + lab, 'meta')

    return M


def sheet_xml_path(zin, sheet_name):
    wb = zin.read('xl/workbook.xml').decode('utf-8')
    rels = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    relmap = dict(re.findall(r'Id="([^"]*)"[^>]*Target="([^"]*)"', rels))
    for nm, rid in re.findall(r'<sheet[^>]*name="([^"]*)"[^>]*r:id="([^"]*)"', wb):
        if nm == sheet_name:
            tgt = relmap[rid]
            return 'xl/' + tgt if not tgt.startswith('xl/') else tgt
    raise KeyError(sheet_name)


def fmt(v):
    if isinstance(v, bool):
        return '1' if v else '0'
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return repr(v)
    return str(v)


def write_cells(wbpath, sheet_name, cell_values):
    """Set constant numeric values for given cells in one sheet, preserving all else."""
    zin = zipfile.ZipFile(wbpath, 'r')
    sp = sheet_xml_path(zin, sheet_name)
    xml = zin.read(sp).decode('utf-8')
    skipped = []
    for ref, val in cell_values.items():
        m = re.search(r'<c r="' + ref + r'"([^>]*?)(?:/>|>(.*?)</c>)', xml, re.S)
        if not m:
            # cell doesn't exist yet - would need row insertion; report instead
            skipped.append((ref, 'missing-cell'))
            continue
        attrs, body = m.group(1), m.group(2)
        if body is not None and '<f' in body:
            skipped.append((ref, 'is-formula'))
            continue
        new_attrs = re.sub(r'\s+t="[^"]*"', '', attrs)  # numeric default type
        newcell = f'<c r="{ref}"{new_attrs}><v>{fmt(val)}</v></c>'
        xml = xml[:m.start()] + newcell + xml[m.end():]
    data = xml.encode('utf-8')
    # rewrite zip preserving every other part byte-for-byte
    tmp = wbpath + '.tmp'
    zout = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED)
    for item in zin.infolist():
        payload = data if item.filename == sp else zin.read(item.filename)
        zi = zipfile.ZipInfo(item.filename, date_time=item.date_time)
        zi.compress_type = item.compress_type
        zi.external_attr = item.external_attr
        zi.internal_attr = item.internal_attr
        zi.create_system = item.create_system
        zout.writestr(zi, payload)
    zout.close()
    zin.close()
    os.replace(tmp, wbpath)
    return skipped


def force_recalc_on_load(wbpath):
    """Set fullCalcOnLoad so Excel recomputes the projection sheet's own formulas."""
    zin = zipfile.ZipFile(wbpath, 'r')
    wb = zin.read('xl/workbook.xml').decode('utf-8')
    if 'fullCalcOnLoad="1"' in wb:
        zin.close(); return
    if '<calcPr' in wb:
        wb2 = re.sub(r'<calcPr ', '<calcPr fullCalcOnLoad="1" ', wb, count=1)
    else:
        wb2 = wb.replace('</workbook>', '<calcPr calcId="191029" fullCalcOnLoad="1"/></workbook>')
    tmp = wbpath + '.tmp'
    zout = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED)
    for item in zin.infolist():
        payload = wb2.encode('utf-8') if item.filename == 'xl/workbook.xml' else zin.read(item.filename)
        zi = zipfile.ZipInfo(item.filename, date_time=item.date_time)
        zi.compress_type = item.compress_type; zi.external_attr = item.external_attr
        zi.internal_attr = item.internal_attr; zi.create_system = item.create_system
        zout.writestr(zi, payload)
    zout.close(); zin.close(); os.replace(tmp, wbpath)


def sync_league(league_dir, write=False, yes=False, calib=None):
    base = os.path.join(REPO, league_dir)
    hpath = os.path.join(base, 'The Sheet Hitters.xlsx')
    ppath = os.path.join(base, 'The Sheet Pitchers.xlsx')
    regpath = os.path.join(base, '25 Regressions.xlsx')
    metpath = os.path.join(base, '25 Metadata.xlsx')
    for p in (hpath, ppath, regpath, metpath):
        if not os.path.exists(p):
            print(f'  [{league_dir}] MISSING {os.path.basename(p)} - skipping league'); return

    H = load_vals(hpath, 'Data Points')
    P = load_vals(ppath, 'Data Points')
    REG = load_vals(regpath, 'Data Points')
    MET = load_vals(metpath, 'Data Points')
    if calib:
        n = override_reg_values(REG, calib)
        print(f'  [calib] regression values overridden from JSON ({n} cells) - workbook used for layout only')
    mapping = build_mapping(H, P, REG, MET)

    SRC = {'REG': REG, 'MET': MET}
    TGT = {'H': H, 'P': P}
    changes = {'H': {}, 'P': {}}
    errors = []
    unchanged = 0
    for m in mapping:
        sval = SRC[m['src']].get(m['sc'])
        if sval is None or (isinstance(sval, str) and sval in ERR_STRINGS):
            errors.append(m); continue
        tval = TGT[m['ts']].get(m['tc'])
        # treat values equal to ~6 significant figures as already-current (skip float-copy noise)
        if isinstance(tval, (int, float)) and isinstance(sval, (int, float)) and abs(tval - sval) <= 1e-6 * (1 + abs(tval)):
            unchanged += 1; continue
        changes[m['ts']][m['tc']] = (sval, tval, m)

    print(f'\n===== {league_dir} =====')
    if errors:
        print(f'  !! ABORT: {len(errors)} source cell(s) are ERROR/blank -> recalc the source workbook(s) and save first.')
        for m in errors[:8]:
            print(f'     {m["src"]}!{m["sc"]}  [{m["label"]}]')
        return
    total = len(changes['H']) + len(changes['P'])
    print(f'  {total} cell(s) will change ({unchanged} already current). Hitters {len(changes["H"])}, Pitchers {len(changes["P"])}.')
    for ts, name in [('H', 'Hitters'), ('P', 'Pitchers')]:
        rows = sorted(changes[ts].items(), key=lambda kv: (kv[1][2]['label'], kv[0]))
        for tc, (sval, tval, m) in rows:
            ov = 'blank' if tval is None else (f'{tval:.6g}' if isinstance(tval, (int, float)) else tval)
            nv = f'{sval:.6g}' if isinstance(sval, (int, float)) else sval
            print(f'    {name:8} {tc:4} {m["label"]:20} {str(ov):>12}  ->  {nv}')

    if write and total and not yes:
        resp = input(f'  Apply these {total} change(s) to {league_dir}? [y/N] ').strip().lower()
        if resp not in ('y', 'yes'):
            print('  skipped (not applied).'); return
    if write and total:
        stamp = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
        for ts, wbpath, name in [('H', hpath, 'Hitters'), ('P', ppath, 'Pitchers')]:
            if not changes[ts]:
                continue
            bak = wbpath + f'.bak-{stamp}'
            shutil.copy2(wbpath, bak)
            print(f'  backup: {os.path.basename(bak)}')
            cv = {tc: sval for tc, (sval, tval, m) in changes[ts].items()}
            skipped = write_cells(wbpath, 'Data Points', cv)
            force_recalc_on_load(wbpath)
            print(f'  wrote {len(cv) - len(skipped)} cells to {name}' + (f'  (skipped {skipped})' if skipped else ''))
    elif write:
        print('  nothing to write.')
    else:
        print('  (dry-run - re-run with --write to apply)')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--league', default='both', choices=['BLM', 'TGS', 'both'])
    ap.add_argument('--write', action='store_true', help='apply changes (otherwise dry-run)')
    ap.add_argument('--yes', action='store_true', help='skip the per-league confirmation prompt')
    ap.add_argument('--calib', help='calibrate.py constants JSON: use ITS regression values '
                    '(the 25 Regressions workbook is then only read for label layout)')
    a = ap.parse_args()
    calib = json.load(open(a.calib)) if a.calib else None
    if calib and a.league == 'both':
        ap.error('--calib is per-league; pass --league TGS or --league BLM with it')
    leagues = ['BLM', 'TGS'] if a.league == 'both' else [a.league]
    for lg in leagues:
        sync_league(LEAGUE_DIRS[lg], write=a.write, yes=a.yes, calib=calib)
    if not a.write:
        print('\nDry-run only. Recalculate 25 Regressions.xlsx + 25 Metadata.xlsx in Excel first,')
        print('then re-run with --write to apply.')


if __name__ == '__main__':
    main()
