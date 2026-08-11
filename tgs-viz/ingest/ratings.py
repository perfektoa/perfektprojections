"""
OOTP ratings-export reader — the piece that kills the manual paste.

The user's OOTP "export to file" produces a table whose columns are the sheet's
`Player List` columns ID..R5 (~99 cols). This reads that export (HTML table OR
delimited text), derives the few helper fields the sheet computes (HT Sort =
height in cm), and hands engine-ready records to the projection engine — no
Excel paste, no Excel formulas.

  python tgs-viz/ingest/ratings.py <export-file> [LEAGUE] [--pitchers]
  python tgs-viz/ingest/ratings.py --selftest [LEAGUE]   # simulate from sheet

Read-only. Stdlib + the engine.
"""
import os, sys, csv, io, re, json
from html.parser import HTMLParser

ENGINE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "engine")
sys.path.insert(0, ENGINE)
import hitters as H  # noqa: E402
import pitchers as P  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def ht_to_sort(ht):
    """'6\\' 4\"' -> 193.04 (cm). The sheet's HT Sort = inches * 2.54."""
    m = re.match(r"\s*(\d+)\s*'\s*(\d+)", str(ht or ""))
    if not m:
        return None
    inches = int(m.group(1)) * 12 + int(m.group(2))
    return inches * 2.54


# Pitch-count fields the Pitchers sheet derives from raw pitch grades (which
# ARE in the export). Current grades for SP Pitch; potential grades for the rest.
_SP_CUR = [("FB", 1), ("CH", 1), ("CB", 1), ("SL", 1), ("SI", 1), ("SP", 1),
           ("CT", 1), ("FO", 1), ("CC", 1), ("SC", 1), ("KC", 1), ("KN", 2)]
_POT = [("KNP", 2), ("FBP", 1), ("CHP", 1), ("CBP", 1), ("SLP", 1), ("SIP", 1),
        ("SPP", 1), ("CTP", 1), ("FOP", 1), ("CCP", 1), ("SCP", 1), ("KCP", 1)]


def _grade(rec, k):
    v = rec.get(k)
    if v in (None, "", "-"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def derive_pitcher_counts(rec):
    # A pitch counts only if it's actually thrown (grade > 0). StatsPlus fills pitches a
    # player DOESN'T have with 0 (not blank), so the old `is not None` test counted all 13
    # slots — making `Pitches` ~13 for everyone and the "enough pitches to start" gate
    # always true (a 2-pitch arm read as a starter). `> 0` counts only his real arsenal.
    pitches = sum(w for k, w in _POT if (_grade(rec, k) or 0) > 0)
    sp = sum(w for k, w in _SP_CUR if (_grade(rec, k) or 0) > 25)
    spp = sum(w for k, w in _POT if (_grade(rec, k) or 0) > 25)
    return pitches, sp, spp


class _TableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows, self._row, self._cell, self._in = [], None, None, False
    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th"):
            self._cell, self._in = "", True
    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._row is not None:
            self._row.append((self._cell or "").strip()); self._in = False
        elif tag == "tr" and self._row is not None:
            if self._row:
                self.rows.append(self._row)
            self._row = None
    def handle_data(self, data):
        if self._in:
            self._cell += data


def parse_text(raw):
    """Parse an export blob (HTML table or delimited) -> list of dict rows."""
    low = raw.lower()
    if "<tr" in low or "<table" in low:
        p = _TableParser(); p.feed(raw); rows = p.rows
    else:
        delim = "\t" if raw.count("\t") >= raw.count(",") else ","
        rows = [r for r in csv.reader(io.StringIO(raw), delimiter=delim)]
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        return []
    headers = [h.strip() for h in rows[0]]
    return [dict(zip(headers, r)) for r in rows[1:]]


def read_export(path):
    with open(path, encoding="utf-8", errors="replace") as f:
        return parse_text(f.read())


def prep(rec):
    """Add sheet-derived helper fields the engine needs but the export lacks.
    Height comes either as the sheet's 6' 1\" string or StatsPlus's cm integer."""
    rec = dict(rec)
    if not rec.get("HT Sort"):
        ht = rec.get("HT")
        s = "" if ht is None else str(ht).strip()
        if "'" in s:
            rec["HT Sort"] = ht_to_sort(s)        # 6' 1"  -> cm
        else:
            try:
                rec["HT Sort"] = float(s)           # StatsPlus already gives cm
            except (TypeError, ValueError):
                rec["HT Sort"] = None
    return rec


def run_hitters(records, league, currency=None, tails=None, fielding=None):
    """Export records -> engine -> full computed records.

    currency (audit D2/D9): a currency_fit.py dict — callers pass
    live_currency(league). Opt-in for the same reason as run_pitchers'
    scurves: the validators must keep matching the workbook.
    tails (audit D4): calib/<LG>/hitter_tails.json — callers pass
    live_hitter_tails(league). Same opt-in contract.
    fielding (audit D3): calib/<LG>/fielding_curves.json — callers pass
    live_fielding(league). Same opt-in contract."""
    hpath = os.path.join(REPO, f"The Sheets {league}", "The Sheet Hitters.xlsx")
    dp, filt, park = H.scan_consts(hpath)
    out = []
    for raw in records:
        rec = prep(raw)
        p, ok = {}, True
        for col in H.RATING_COLS:
            v = rec.get(col)
            if col in ("B", "T", "Name", "POS"):
                p[col] = v
            else:
                nv = 20.0 if isinstance(v, str) and v.strip() == "-" else H.num(v)
                p[col] = nv
                if nv is None and col not in H.OPTIONAL:
                    ok = False
        if not ok:
            continue
        try:
            comp = H.compute(p, dp, filt, park, league, currency=currency,
                             tails=tails, fielding=fielding)
        except Exception:
            continue
        merged = dict(rec)
        merged.update({k: v for k, v in comp.items() if not k.startswith("_")})
        out.append(merged)
    return out


def live_scurves(league):
    """audit D1 — which pitching rate-curve model is LIVE per league.

    TGS: the fitted+transported logistic S-curves, calib/TGS/scurves.json
    (promoted 2026-08-05 from the verified preview after the verifier's
    deep-tail clamp). Pipeline entry points (refresh.py, draft.py) call this
    and pass the result into run_pitchers; run_pitchers itself stays opt-in so
    the sheet-fidelity validators keep producing two-segment numbers.

    BLM: None — the two-segment lines stay live.
    TODO(D1 BLM — DO NOT FLIP YET): BLM's ratings were re-scouted mid-season
    AFTER its 25 Metadata.xlsx anchors were computed, so the S-curve live
    transport (fitted against those anchors) can't be trusted until BLM's
    season-end rebuild. When that rebuild lands: re-run scurve_fit.py --league
    BLM, verify the D1 gates, promote calib/BLM/scurves-preview.json ->
    scurves.json, and delete the league gate below.
    """
    if league != "TGS":
        return None
    path = os.path.join(REPO, "tgs-viz", "engine", "calib", league, "scurves.json")
    return P.load_scurves(path) if os.path.exists(path) else None


def live_hitter_tails(league):
    """audit D4 — the archive-MEASURED hitter tail-region corrections, LIVE for
    BOTH leagues (league policy: fitted from each league's own clone archive,
    not from its metadata anchors, so BLM's mid-season re-scout does not gate
    it). Produced by engine/hitter_tails_fit.py; deltas are 0 outside the four
    audited tail regions. Returns None (sheet behavior) when missing."""
    path = os.path.join(REPO, "tgs-viz", "engine", "calib", league, "hitter_tails.json")
    return json.load(open(path, encoding="utf-8")) if os.path.exists(path) else None


def live_fielding(league):
    """audit D3 — the monotone piecewise PM% curves, LIVE for BOTH leagues.
    Produced by engine/fielding_curves_fit.py (archive isotonic fit + live
    ip-weighted mean-zero transport offsets from 25 Metadata.xlsx). NOTE: the
    offsets bake in the CURRENT metadata population — refit after any metadata
    rebuild (the Recalibrate bats re-run the fit tools). Returns None (sheet
    linear PM%) when missing."""
    path = os.path.join(REPO, "tgs-viz", "engine", "calib", league, "fielding_curves.json")
    return json.load(open(path, encoding="utf-8")) if os.path.exists(path) else None


def live_currency(league):
    """audit D2 + D9 — the archive-FITTED currency layer, LIVE for BOTH leagues
    (league policy: every fix applies to both; only the D1 S-curve flip is
    TGS-gated, because it depends on BLM's mid-season-re-scouted anchors —
    the currency constants below do NOT: the RA/9 exponents and RPW are fitted
    from each league's clone archive, not from its metadata anchors).

    calib/<LEAGUE>/currency.json is produced by engine/currency_fit.py and
    carries: per-role wOBA->RA/9 exponents (D2), the fitted runs-per-win H30
    override (D9), and report-only fit results (run-value regression + gates,
    luck floor, workload shares, RP shrink) consumed by the JS side via
    src/lib/leagueCalib.js. Returns None (sheet behavior) when missing."""
    path = os.path.join(REPO, "tgs-viz", "engine", "calib", league, "currency.json")
    return P.load_currency(path) if os.path.exists(path) else None


# ---- audit M2 (INTERIM): pull-history smoothing ------------------------------
# refresh.py archives every successful live StatsPlus pull under
# .cache/history/statsplus_<slug>_<YYYYMMDD-HHMM>.json (raw schema); recovered
# vintages carry a "recovered-<YYYYMMDD>" stamp (sheet-name schema, ratings
# columns only). smoothed_pull() below averages scouting ratings across the
# last few vintages. PREVIEW-ONLY: nothing in the live refresh path calls it —
# wiring it in is an explicit user decision.

HIST_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".cache", "history")

# Scouting-rating columns eligible for cross-pull averaging, in SHEET-name
# space (i.e. after statsplus.translate_rows). Identity/bio/assignment fields
# (ID, Name, POS, B, T, Age, HT, League, Lev, Ovr/Pot, personality...) are
# NEVER averaged — they always pass through from the newest pull.
_SMOOTH_PIT = (["STU", "HRR", "PBABIP", "CON"]
               + [f"{c} {s}" for c in ("STU", "HRR", "PBABIP", "CON") for s in ("vR", "vL", "P")]
               + ["STM", "HLD"]
               + [k for k, _ in _SP_CUR] + [k for k, _ in _POT])   # pitch grades
_SMOOTH_HIT = ([f"{c} {s}" for c in ("BA", "GAP", "POW", "EYE", "K") for s in ("vR", "vL")]
               + ["HT P", "GAP P", "POW P", "EYE P", "K P",
                  "SPE", "SR", "STE", "RUN",
                  "IF RNG", "IF ERR", "IF ARM", "TDP", "OF RNG", "OF ERR", "OF ARM",
                  "C ABI", "C FRM", "C ARM"])
SMOOTH_COLS = _SMOOTH_PIT + _SMOOTH_HIT


def _pull_stamp(fname):
    """Sortable (date, hhmm) stamp from a history filename; a bare YYYYMMDD
    (recovered vintages) sorts as that day 00:00."""
    m = re.search(r"(\d{8})(?:-(\d{4}))?\.json$", fname)
    return (m.group(1), m.group(2) or "0000") if m else ("", "")


def load_pull_history(slug, history_dir=None):
    """All archived pulls for `slug`, oldest -> newest, translated to sheet
    column names (translate_rows is a no-op on already-translated recovered
    vintages). Returns [(filename, rows), ...]."""
    d = history_dir or HIST_DIR
    if not os.path.isdir(d):
        return []
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    import statsplus as S
    names = sorted((fn for fn in os.listdir(d)
                    if fn.startswith(f"statsplus_{slug}_") and fn.endswith(".json")),
                   key=_pull_stamp)
    return [(fn, S.translate_rows(json.load(open(os.path.join(d, fn), encoding="utf-8"))))
            for fn in names]


def smoothed_pull(slug, n=3, history_dir=None):
    """audit M2 INTERIM RULE — ratings-input smoothing over archived pulls.

    Returns the NEWEST archived pull's rows (sheet-name schema, engine-ready)
    with each scouting-rating column in SMOOTH_COLS replaced by the plain
    average of that player's last `n` available values across the archive
    (default n=3; fewer when fewer pulls carry the player/column — a single
    pull passes through unchanged). '-' reads as 20, matching the engine's
    convention in run_hitters/run_pitchers.

    INTERIM: the equal-weight last-3 average is the audit's M2 stopgap for
    snapshot churn (pull-to-pull scout noise moves projections ~.0346 wOBA
    mean, ~3x model error). The intended final rule is an EWMA whose decay
    weight is FITTED from backtest data (smoothed-vs-raw projections scored
    against season actuals) once enough vintages exist — do not hand-tune n.

    PREVIEW-ONLY until the user flips it live: refresh.py does not call this.
    """
    pulls = load_pull_history(slug, history_dir)
    if not pulls:
        return []
    byid = [{str(r.get("ID")): r for r in rows} for _, rows in pulls]  # oldest->newest
    out = []
    for r in pulls[-1][1]:
        pid = str(r.get("ID"))
        rec = dict(r)
        for col in SMOOTH_COLS:
            vals = []
            for m in byid:
                rr = m.get(pid)
                if rr is None or col not in rr:
                    continue
                v = rr[col]
                nv = 20.0 if isinstance(v, str) and v.strip() == "-" else H.num(v)
                if nv is not None:
                    vals.append(nv)
            if vals:
                vals = vals[-n:]
                rec[col] = sum(vals) / len(vals)
        out.append(rec)
    return out


def run_pitchers(records, league, scurves=None, currency=None):
    """scurves (audit D1): a pitchers.load_scurves() dict — when given, the four
    pitching rate blocks use the fitted logistic curves instead of the
    two-segment lines. Deliberately opt-in (callers pass live_scurves(league)):
    the validators in this file and engine/pitchers.py must keep matching the
    workbook's cached two-segment values.
    currency (audit D2/D9): a currency_fit.py dict — callers pass
    live_currency(league); same opt-in contract."""
    ppath = os.path.join(REPO, f"The Sheets {league}", "The Sheet Pitchers.xlsx")
    dp, filt, park_aa = P.scan_consts(ppath)
    out = []
    for raw in records:
        rec = dict(raw)
        rec["Pitches"], rec["SP Pitch"], rec["SP P Pitch"] = derive_pitcher_counts(rec)
        p, ok = {}, True
        for col in P.RATING_COLS:
            v = rec.get(col)
            if col in P.META:
                p[col] = v
            else:
                nv = 20.0 if isinstance(v, str) and v.strip() == "-" else H.num(v)
                p[col] = nv
                if nv is None and col not in P.OPTIONAL:
                    ok = False
        if not ok:
            continue
        try:
            comp = P.compute(p, dp, filt, park_aa, scurves=scurves, currency=currency)
        except Exception:
            continue
        merged = dict(rec)
        merged.update({k: v for k, v in comp.items() if not str(k).startswith("_")})
        out.append(merged)
    return out


def _simulate_export(league, kind):
    fname = "The Sheet Pitchers.xlsx" if kind == "pitchers" else "The Sheet Hitters.xlsx"
    wb = load_workbook(os.path.join(REPO, f"The Sheets {league}", fname), read_only=True, data_only=True)
    ws = wb["Player List"]
    header = None
    out_rows = []
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        if header is None:
            header = [str(v).strip() if v is not None else "" for v in vals]
            idx = {h: i for i, h in enumerate(header)}
            lo, hi = idx["ID"], idx["R5"]
            out_rows.append(header[lo:hi + 1])
            continue
        if idx.get("Name") is None or not vals[idx["Name"]]:
            continue
        out_rows.append(["" if vals[i] is None else str(vals[i]) for i in range(lo, hi + 1)])
    wb.close()
    return "\n".join("\t".join(r) for r in out_rows)


def _selftest(league, kind="hitters"):
    """Simulate an OOTP export from the sheet's own Player List (ID..R5), parse
    it back, run the engine, and confirm parity with the current JSON — proving
    the export->reader->engine->JSON path on the REAL schema, no real file needed."""
    records = parse_text(_simulate_export(league, kind))
    # The self-test proves the export->reader->engine->JSON path, so it must run
    # the LIVE engine configuration (D1 S-curves + D2/D9 currency where live) —
    # the JSON legitimately diverges from the workbook by exactly those layers.
    if kind == "pitchers":
        computed = run_pitchers(records, league, scurves=live_scurves(league),
                                currency=live_currency(league))
        jname, cols = "pitchers.json", ["WAA wtd", "WAA wtd RP", "WAP", "RA/9 wtd"]
    else:
        computed = run_hitters(records, league, currency=live_currency(league),
                               tails=live_hitter_tails(league),
                               fielding=live_fielding(league))
        jname, cols = "hitters.json", ["Max WAA wtd", "wOBA wtd", "MAX WAA P", "2B WAA wtd", "C WAA wtd", "HT Sort"]
    cur = {str(r.get("ID")): r for r in json.load(open(os.path.join(REPO, "tgs-viz", "public", "data", league, jname), encoding="utf-8"))}
    diffs = {c: 0.0 for c in cols}
    matched = 0
    for r in computed:
        j = cur.get(str(r.get("ID")))
        if not j:
            continue
        matched += 1
        for c in cols:
            mine, sheet = r.get(c), H.num(j.get(c))
            if mine is None or sheet is None:
                continue
            diffs[c] = max(diffs[c], abs(float(mine) - float(sheet)))
    print(f"{league} {kind} self-test: parsed {len(records)} export rows -> {len(computed)} computed, matched {matched} to JSON")
    for c in cols:
        print(f"    {c:14} {diffs[c]:.3e}")
    if kind == "pitchers":
        proj = diffs.get("RA/9 wtd", 0)
        print(f"  => RA/9 (the projection) exact to {proj:.1e}. WAA carries the sheet's "
              f"ROW()/1e7 ranking tie-breaker (~{max(diffs.values()):.0e}) — a sort nudge, not the projection.")
    else:
        worst = max(diffs.values()) if diffs else 0
        print(f"  => {'OK — export reader + engine reproduce the sheet exactly' if worst < 1e-6 else 'CHECK'} (worst {worst:.2e})")


def main():
    if "--selftest" in sys.argv:
        league = next((a for a in sys.argv[1:] if not a.startswith("--")), "TGS")
        _selftest(league, "pitchers" if "--pitchers" in sys.argv else "hitters")
        return
    if len(sys.argv) < 2:
        print(__doc__)
        return
    path = sys.argv[1]
    league = next((a for a in sys.argv[2:] if not a.startswith("--")), "TGS")
    records = read_export(path)
    print(f"read {len(records)} rows from {path}")
    if records:
        print("columns:", list(records[0].keys())[:6], "...")
        computed = run_hitters(records, league)
        print(f"engine computed {len(computed)} hitters")


if __name__ == "__main__":
    main()
